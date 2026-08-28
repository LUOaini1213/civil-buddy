"""
真实尺寸覆盖：优先 dims_override.json / Excel，再估算。

dims_override.json 格式:
{
  "by_keyword": {
    "预埋": {"length_mm": 800, "width_mm": 400, "height_mm": 300},
    "钢通": {"length_mm": 5800, "width_mm": 200, "height_mm": 200}
  },
  "by_id": {
    "M001": {"length_mm": 4200, "width_mm": 350, "height_mm": 175}
  }
}
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Optional


def load_override(path: Optional[str | Path] = None) -> Dict[str, Any]:
    if path is None:
        path = Path(__file__).resolve().parents[2] / "knowledge" / "dims_override.json"
    path = Path(path)
    if not path.exists():
        return {"by_keyword": {}, "by_id": {}}
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def apply_dims_override(
    materials: List[Dict[str, Any]],
    override: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    """返回新列表，命中覆盖则 dims_estimated=False。"""
    ov = override if override is not None else load_override()
    by_id = ov.get("by_id") or {}
    by_kw = ov.get("by_keyword") or {}
    out = []
    for m in materials:
        mm = dict(m)
        mid = mm.get("id") or ""
        name = mm.get("name") or ""
        hit = None
        if mid in by_id:
            hit = by_id[mid]
        else:
            for kw, dims in by_kw.items():
                if kw and kw in name:
                    hit = dims
                    break
        if hit:
            mm["length_mm"] = float(hit.get("length_mm") or mm.get("length_mm") or 0)
            mm["width_mm"] = float(hit.get("width_mm") or mm.get("width_mm") or 0)
            mm["height_mm"] = float(hit.get("height_mm") or mm.get("height_mm") or 0)
            mm["dims_estimated"] = False
            mm["dims_source"] = hit.get("source") or "override"
        out.append(mm)
    return out


def try_load_excel_dims(excel_path: Path) -> Dict[str, Dict[str, float]]:
    """
    从装货单类 Excel 抽 名称/编号 → 尺寸。
    宽松匹配列名：长/宽/高/重量。
    """
    try:
        import openpyxl
    except ImportError:
        return {}
    if not excel_path.exists():
        return {}
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    mapping: Dict[str, Dict[str, float]] = {}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(max_row=5, max_col=20, values_only=True))
        header_row = None
        headers: List[str] = []
        for r in rows:
            cells = [str(c or "") for c in r]
            joined = "".join(cells)
            if any(k in joined for k in ("长", "宽", "高", "重量", "名称")):
                header_row = cells
                headers = cells
                break
        if not header_row:
            continue
        # map col index
        def find_col(*keys):
            for i, h in enumerate(headers):
                for k in keys:
                    if k in h:
                        return i
            return None

        c_name = find_col("名称", "品名", "产品")
        c_part = find_col("编号", "图号", "加工")
        c_l = find_col("长")
        c_w = find_col("宽")
        c_h = find_col("高")
        if c_l is None and c_w is None:
            continue
        for row in ws.iter_rows(min_row=2, max_col=20, values_only=True):
            try:
                name = str(row[c_name] or "") if c_name is not None else ""
                part = str(row[c_part] or "") if c_part is not None else ""
                L = float(row[c_l] or 0) if c_l is not None else 0
                W = float(row[c_w] or 0) if c_w is not None else 0
                H = float(row[c_h] or 0) if c_h is not None else 0
                if L <= 0 and W <= 0:
                    continue
                # 有些表 长宽高 单位是 mm
                dims = {
                    "length_mm": L,
                    "width_mm": W,
                    "height_mm": H,
                    "source": f"excel:{excel_path.name}",
                }
                if name:
                    mapping[name[:40]] = dims
                if part:
                    mapping[part[:40]] = dims
            except Exception:
                continue
    return mapping
