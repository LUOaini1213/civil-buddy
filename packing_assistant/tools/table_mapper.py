"""通用材料表 → MaterialTableIR / materials[]。

任意 Excel/CSV/字典行：中英列名同义词 + 单位归一 → 标准 materials 字段。
行业无关；钢材仅为 profile 提示之一。
"""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple, Union

PathLike = Union[str, Path]

# 标准字段 → 同义词（小写匹配）
COLUMN_SYNONYMS: Dict[str, Tuple[str, ...]] = {
    "id": ("id", "编号", "行号", "line_id", "row_id", "line_no", "no", "序号"),
    "name": (
        "name",
        "名称",
        "品名",
        "货物名称",
        "货名",
        "item",
        "item description",
        "product",
        "description",
        "desc",
        "sku_name",
        "物料名称",
        "品名规格",
        "article",
        "article_name",
        "货品名称",
        "商品名称",
    ),
    "quantity": (
        "quantity",
        "数量",
        "件数",
        "箱数",
        "qty",
        "qty.",
        "q'ty",
        "count",
        "pcs",
        "pc",
        "件",
        "装箱数",
    ),
    "length_mm": (
        "length_mm",
        "length",
        "len",
        "l",
        "长",
        "长度",
        "外长",
        "length_cm",
        "length_m",
        "length(m)",
        "length (m)",
        "l_mm",
        "l_cm",
        "l_m",
        "长度_m",
        "长度_mm",
        "长(mm)",
        "长(m)",
        "长mm",
        "长m",
    ),
    "width_mm": (
        "width_mm",
        "width",
        "w",
        "宽",
        "宽度",
        "外宽",
        "width_cm",
        "width_m",
        "width(cm)",
        "width (cm)",
        "w_mm",
        "w_cm",
        "宽度_mm",
        "宽(mm)",
        "宽mm",
    ),
    "height_mm": (
        "height_mm",
        "height",
        "h",
        "高",
        "高度",
        "外高",
        "height_cm",
        "height_m",
        "height(mm)",
        "height (mm)",
        "h_mm",
        "h_cm",
        "高度_mm",
        "高(mm)",
        "高mm",
        "厚",
        "厚度",
    ),
    "weight_kg": (
        "weight_kg",
        "weight",
        "单重",
        "重量",
        "毛重",
        "净重",
        "gross_weight",
        "gross_kg",
        "net_weight",
        "net weight(kg)",
        "net weight (kg)",
        "kg",
        "unit_weight",
        "单重kg",
        "单重(kg)",
        "weight_t",
        "单重t",
        "吨",
    ),
    "total_weight_kg": (
        "total_weight_kg",
        "total_weight",
        "总重",
        "合计重量",
        "total_kg",
        "gross_total",
        "总重kg",
        "总重(kg)",
        "total_t",
    ),
    "part_no": (
        "part_no",
        "part",
        "件号",
        "料号",
        "图号",
        "sku",
        "sku code",
        "item_no",
        "item_code",
        "drawing_no",
        "物料编码",
    ),
    "category": ("category", "类别", "类型", "品类", "type", "class", "分类"),
    "spec": ("spec", "规格", "型号", "model", "规格型号"),
    "note": ("note", "备注", "说明", "remark", "comments", "comment"),
}

CATEGORY_ALIASES: Dict[str, str] = {
    "carton": "carton",
    "纸箱": "carton",
    "纸盒": "carton",
    "box": "carton",
    "crate": "crate",
    "木箱": "crate",
    "铁架": "crate",
    "铁笼": "crate",
    "long_item": "long_item",
    "超长件": "long_item",
    "长材": "long_item",
    "管材": "long_item",
    "型材": "long_item",
    "pallet": "pallet",
    "托盘": "pallet",
    "bulk_bag": "bulk_bag",
    "吨袋": "bulk_bag",
    "集装袋": "bulk_bag",
    "fragile": "fragile",
    "易碎": "fragile",
    "玻璃": "fragile",
    "liquid_unit": "liquid_unit",
    "液体": "liquid_unit",
    "generic": "generic",
    "普通件": "generic",
    "重件": "generic",
}


def _norm_header(h: Any) -> str:
    s = str(h or "").strip().lower()
    s = s.replace("（", "(").replace("）", ")")
    s = re.sub(r"\s+", "", s)
    return s


def build_column_map(headers: Sequence[Any]) -> Dict[str, str]:
    """原表头 → 标准字段。"""
    inv: Dict[str, str] = {}
    for std, syns in COLUMN_SYNONYMS.items():
        for s in syns:
            inv[_norm_header(s)] = std

    mapping: Dict[str, str] = {}
    used_std: set[str] = set()
    for h in headers:
        raw = str(h or "").strip()
        if not raw:
            continue
        key = _norm_header(raw)
        std = inv.get(key)
        if not std:
            # 模糊：含 length/长 等
            for cand, field in (
                ("length", "length_mm"),
                ("width", "width_mm"),
                ("height", "height_mm"),
                ("weight", "weight_kg"),
                ("qty", "quantity"),
                ("数量", "quantity"),
                ("长", "length_mm"),
                ("宽", "width_mm"),
                ("高", "height_mm"),
            ):
                if cand in key and field not in used_std:
                    std = field
                    break
        if std and std not in used_std:
            mapping[raw] = std
            used_std.add(std)
    return mapping


def _to_float(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    s = re.sub(r"[^\d.\-eE]", "", s)
    if not s or s in ("-", "."):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _infer_length_scale(header: str, values: List[Optional[float]]) -> float:
    """返回乘到 mm 的系数。"""
    h = _norm_header(header)
    if "mm" in h or "(mm)" in h:
        return 1.0
    if h.endswith("_cm") or "cm" in h or "(cm)" in h:
        return 10.0
    if h.endswith("_m") or "(m)" in h or h in ("length_m", "width_m", "height_m", "长m", "长(m)"):
        return 1000.0
    # 启发式：中位值 < 30 → 可能是 m；< 300 且字段叫 length → cm 少见，按 mm
    nums = [x for x in values if x is not None and x > 0]
    if not nums:
        return 1.0
    nums_sorted = sorted(nums)
    mid = nums_sorted[len(nums_sorted) // 2]
    if mid <= 25:
        return 1000.0  # meters
    if mid <= 300 and ("m" in h and "mm" not in h):
        return 1000.0
    return 1.0


def _infer_weight_scale(header: str, values: List[Optional[float]]) -> float:
    """返回乘到 kg 的系数。注意：不可用 `'t' in header`（weight 含字母 t）。"""
    h = _norm_header(header)
    raw = str(header or "")
    # 明确吨：_t / (t) / 吨 / weight_t / 单重t —— 排除 weight/net_weight 等
    if "吨" in raw:
        return 1000.0
    if re.search(r"(^|_)(t)($|[^a-z])", h) or h.endswith("_t") or "(t)" in h:
        if "kg" not in h and "weight" not in h:
            return 1000.0
    if (
        h in ("weight_t", "total_t", "单重t", "总重t", "吨")
        or h.endswith("weight_t")
        or (h.endswith("_t") and "weight" in h)
    ):
        return 1000.0
    # 克（非 kg）— 避免 "kg" 被误判
    if "kg" not in h and (h.endswith("_g") or h.endswith("(g)") or "(g)" in h):
        return 0.001
    nums = [x for x in values if x is not None and x > 0]
    if nums:
        mid = sorted(nums)[len(nums) // 2]
        if mid > 50000:
            return 0.001
    return 1.0


def normalize_category(raw: Any) -> str:
    s = str(raw or "").strip()
    if not s:
        return "generic"
    key = s.lower()
    if key in CATEGORY_ALIASES:
        return CATEGORY_ALIASES[key]
    if s in CATEGORY_ALIASES:
        return CATEGORY_ALIASES[s]
    for k, v in CATEGORY_ALIASES.items():
        if k in s or k in key:
            return v
    return s  # keep free text; caller may still use as-is


# 最近一次 rows_to_ir 清洗统计（parse_table_* 读取，避免改动返回类型）
_LAST_CLEAN_STATS: Dict[str, Any] = {}


def last_clean_stats() -> Dict[str, Any]:
    """返回最近一次 rows_to_ir 的清洗计数。"""
    return dict(_LAST_CLEAN_STATS)


def rows_to_ir(
    rows: Sequence[Dict[str, Any]],
    *,
    headers: Optional[Sequence[Any]] = None,
    source: str = "dict",
    source_path: str = "",
    profile_hint: str = "generic_table",
) -> List[Dict[str, Any]]:
    """字典行列表 → MaterialTableIR（同时兼容现有 materials API）。"""
    global _LAST_CLEAN_STATS
    clean_stats: Dict[str, int] = {
        "n_input_rows": 0,
        "n_kept": 0,
        "n_skip_empty_name": 0,
        "n_skip_noise_name": 0,
        "n_skip_zero_qty": 0,
        "n_skip_zero_placeholder": 0,
        "n_skip_summary_row": 0,
    }
    if not rows:
        _LAST_CLEAN_STATS = {**clean_stats, "n_skipped_total": 0}
        return []

    if headers is None:
        # 保留首次出现顺序
        seen: List[str] = []
        for r in rows:
            for k in r.keys():
                if k not in seen:
                    seen.append(str(k))
        headers = seen

    colmap = build_column_map(list(headers))
    # reverse: std -> original header
    std_to_raw = {std: raw for raw, std in colmap.items()}

    # collect raw numeric series for unit inference
    def series(std: str) -> List[Optional[float]]:
        raw_h = std_to_raw.get(std)
        if not raw_h:
            return []
        return [_to_float(r.get(raw_h)) for r in rows]

    len_scale = {
        "length_mm": _infer_length_scale(std_to_raw.get("length_mm", "length_mm"), series("length_mm")),
        "width_mm": _infer_length_scale(std_to_raw.get("width_mm", "width_mm"), series("width_mm")),
        "height_mm": _infer_length_scale(std_to_raw.get("height_mm", "height_mm"), series("height_mm")),
    }
    wt_scale = _infer_weight_scale(std_to_raw.get("weight_kg", "weight_kg"), series("weight_kg"))
    tw_scale = _infer_weight_scale(
        std_to_raw.get("total_weight_kg", "total_weight_kg"), series("total_weight_kg")
    )

    out: List[Dict[str, Any]] = []
    # 汇总/小计行（整行品名，不误杀「合计架」类真货子串尾）
    _SUMMARY_EXACT = {
        "合计",
        "小计",
        "总计",
        "汇总",
        "总合计",
        "本页合计",
        "grand total",
        "subtotal",
        "total",
        "sum",
    }
    for i, r in enumerate(rows, 1):
        clean_stats["n_input_rows"] += 1
        got: Dict[str, Any] = {}
        for raw_h, std in colmap.items():
            got[std] = r.get(raw_h)

        name = got.get("name")
        if name is None or str(name).strip() == "" or str(name).strip("\u3000 \t") == "":
            clean_stats["n_skip_empty_name"] += 1
            continue
        name_s = str(name).strip().strip("\u3000")
        # 噪声行：整行注释/占位（禁止用「表头」「跳过」等子串误杀真货名）
        if name_s.startswith("#") or name_s.startswith("//"):
            clean_stats["n_skip_noise_name"] += 1
            continue
        low_name = name_s.lower().strip()
        if low_name in (
            "nan",
            "none",
            "null",
            "-",
            "—",
            "n/a",
            "na",
            "header",
            "headers",
            "# comment",
            "comment",
            "备注",
            "备注行",
            "说明",
            "placeholder",
        ):
            clean_stats["n_skip_noise_name"] += 1
            continue
        # 仅匹配整行或「注释:」类前缀，不匹配品名中含「表头/跳过」的真货
        if name_s in ("注释", "这是注释", "表头", "无效空行", "无效空行应跳过", "跳过"):
            clean_stats["n_skip_noise_name"] += 1
            continue
        if name_s.startswith("注释") and (
            len(name_s) <= 2 or name_s[2] in (":", "：", " ", "\t", "-")
        ):
            clean_stats["n_skip_noise_name"] += 1
            continue
        if name_s.startswith("这是注释"):
            clean_stats["n_skip_noise_name"] += 1
            continue
        if low_name in _SUMMARY_EXACT or name_s in _SUMMARY_EXACT:
            clean_stats["n_skip_summary_row"] += 1
            continue
        # 仅标点/空白品名
        if all(not ch.isalnum() and ord(ch) < 128 for ch in name_s.replace(" ", "")):
            # allow CJK product names (non-ascii alnum check fails for CJK)
            if not any("\u4e00" <= ch <= "\u9fff" for ch in name_s):
                clean_stats["n_skip_noise_name"] += 1
                continue

        qty_f = _to_float(got.get("quantity"))
        # 显式 0 数量：噪声，不升成 1
        if qty_f is not None and qty_f <= 0:
            clean_stats["n_skip_zero_qty"] += 1
            continue
        qty = max(1, int(qty_f or 1))

        def dim(std: str) -> float:
            v = _to_float(got.get(std))
            if v is None:
                return 0.0
            return round(v * len_scale[std], 3)

        L, W, H = dim("length_mm"), dim("width_mm"), dim("height_mm")
        dims_estimated = L <= 0 or W <= 0 or H <= 0

        unit_w = _to_float(got.get("weight_kg"))
        if unit_w is not None:
            unit_w = unit_w * wt_scale
        total_w = _to_float(got.get("total_weight_kg"))
        if total_w is not None:
            total_w = total_w * tw_scale
        if total_w is None and unit_w is not None:
            total_w = unit_w * qty
        if unit_w is None and total_w is not None:
            unit_w = total_w / qty
        unit_w = float(unit_w or 0.0)
        total_w = float(total_w or 0.0)
        # 全零占位行（无尺寸无重量）丢弃
        if L <= 0 and W <= 0 and H <= 0 and unit_w <= 0 and total_w <= 0:
            clean_stats["n_skip_zero_placeholder"] += 1
            continue

        cat_raw = got.get("category") or ""
        cat = normalize_category(cat_raw) if cat_raw else _guess_category(L, W, H, unit_w, name_s)

        conf = 0.95
        if dims_estimated:
            conf -= 0.35
        if unit_w <= 0 and total_w <= 0:
            conf -= 0.2
        conf = max(0.1, min(1.0, conf))

        mid = str(got.get("id") or got.get("part_no") or f"M{i:03d}").strip()
        item = {
            "id": mid,
            "name": name_s,
            "spec": str(got.get("spec") or ""),
            "quantity": qty,
            "weight_kg": round(unit_w, 4),
            "total_weight_kg": round(total_w, 4),
            "length_mm": L,
            "width_mm": W,
            "height_mm": H,
            "part_no": str(got.get("part_no") or ""),
            "category": cat,
            "note": str(got.get("note") or ""),
            "meta": {
                "source": source,
                "source_path": source_path,
                "column_map": dict(colmap),
                "units_in": {
                    "length_scale_to_mm": len_scale,
                    "weight_scale_to_kg": {"weight_kg": wt_scale, "total_weight_kg": tw_scale},
                },
                "confidence": round(conf, 3),
                "dims_estimated": dims_estimated,
                "profile_hint": profile_hint,
            },
        }
        out.append(item)
        clean_stats["n_kept"] += 1
    clean_stats["n_skipped_total"] = int(clean_stats["n_input_rows"]) - int(
        clean_stats["n_kept"]
    )
    _LAST_CLEAN_STATS = dict(clean_stats)
    return out


def _guess_category(L: float, W: float, H: float, weight: float, name: str) -> str:
    n = name.lower()
    if any(k in name or k in n for k in ("玻璃", "fragile", "易碎")):
        return "fragile"
    if any(k in name or k in n for k in ("吨袋", "集装袋", "bulk")):
        return "bulk_bag"
    if any(k in name or k in n for k in ("托盘", "pallet")):
        return "pallet"
    if any(k in name or k in n for k in ("纸箱", "carton")):
        return "carton"
    if L >= 4000:
        return "long_item"
    if max(L, W, H) <= 800 and weight <= 50:
        return "carton"
    return "generic"


def load_csv(path: PathLike, encoding: str = "utf-8-sig") -> List[Dict[str, Any]]:
    path = Path(path)
    with path.open("r", encoding=encoding, newline="") as f:
        sample = f.read(8192)
        f.seek(0)
        # Prefer explicit delimiter counts (Sniffer often fails on short EU files)
        delim = ","
        for candidate in (";", "\t", ","):
            if sample.count(candidate) >= 2:
                # header line vote
                first = sample.splitlines()[0] if sample else ""
                if first.count(candidate) >= 2:
                    delim = candidate
                    break
        else:
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
                delim = dialect.delimiter
            except csv.Error:
                delim = ","
        reader = csv.DictReader(f, delimiter=delim)
        rows = [dict(r) for r in reader]
        headers = list(reader.fieldnames or [])
    return rows_to_ir(rows, headers=headers, source="csv", source_path=str(path))


def load_xlsx(path: PathLike, sheet: Optional[str] = None) -> List[Dict[str, Any]]:
    import openpyxl

    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet and sheet in wb.sheetnames:
        ws = wb[sheet]
    elif "materials" in wb.sheetnames:
        ws = wb["materials"]
    else:
        ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    wb.close()
    if not data:
        return []
    headers = [str(c or "").strip() for c in data[0]]
    rows: List[Dict[str, Any]] = []
    for row in data[1:]:
        d = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        # skip full_flow non-material
        rt = d.get("row_type")
        if rt and str(rt) not in ("material", "materials", ""):
            continue
        rows.append(d)
    return rows_to_ir(rows, headers=headers, source="xlsx", source_path=str(path))


def load_json(path: PathLike) -> List[Dict[str, Any]]:
    path = Path(path)
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    if isinstance(data, dict):
        data = data.get("materials") or data.get("rows") or data.get("items") or []
    if not isinstance(data, list):
        raise ValueError("JSON must be list or {materials|rows|items: list}")
    # already IR?
    if data and isinstance(data[0], dict) and "length_mm" in data[0] and "name" in data[0]:
        out = []
        for i, m in enumerate(data, 1):
            item = dict(m)
            item.setdefault("id", f"M{i:03d}")
            item.setdefault("quantity", 1)
            meta = dict(item.get("meta") or {})
            meta.setdefault("source", "json")
            meta.setdefault("source_path", str(path))
            meta.setdefault("profile_hint", "generic_table")
            meta.setdefault("confidence", 0.99)
            meta.setdefault("dims_estimated", float(item.get("length_mm") or 0) <= 0)
            item["meta"] = meta
            out.append(item)
        return out
    return rows_to_ir(data, source="json", source_path=str(path))


def load_table(path: PathLike, **kwargs: Any) -> List[Dict[str, Any]]:
    path = Path(path)
    suf = path.suffix.lower()
    if suf in (".csv", ".tsv", ".txt"):
        return load_csv(path)
    if suf in (".xlsx", ".xlsm"):
        return load_xlsx(path, sheet=kwargs.get("sheet"))
    if suf == ".json":
        return load_json(path)
    raise ValueError(f"unsupported table type: {suf}")


def ir_to_materials(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """IR → 现有 materials API（去掉 meta 也可保留）。"""
    mats = []
    for m in rows:
        mats.append(
            {
                "id": m.get("id"),
                "name": m.get("name"),
                "spec": m.get("spec") or "",
                "quantity": int(m.get("quantity") or 1),
                "weight_kg": float(m.get("weight_kg") or 0),
                "total_weight_kg": float(m.get("total_weight_kg") or 0),
                "length_mm": float(m.get("length_mm") or 0),
                "width_mm": float(m.get("width_mm") or 0),
                "height_mm": float(m.get("height_mm") or 0),
                "part_no": m.get("part_no") or "",
                "category": m.get("category") or "generic",
                "note": m.get("note") or "",
                "meta": m.get("meta") or {},
            }
        )
    return mats


def parse_table_file(path: PathLike, **kwargs: Any) -> Dict[str, Any]:
    """统一入口：文件 → {materials, ir, stats, column_map}。"""
    path = Path(path)
    ir = load_table(path, **kwargs)
    mats = ir_to_materials(ir)
    n_est = sum(1 for m in ir if (m.get("meta") or {}).get("dims_estimated"))
    confs = [(m.get("meta") or {}).get("confidence", 0) for m in ir]
    colmap = {}
    if ir:
        colmap = dict((ir[0].get("meta") or {}).get("column_map") or {})
    clean = last_clean_stats()
    return {
        "ok": bool(mats),
        "path": str(path),
        "materials": mats,
        "ir": ir,
        "column_map": colmap,
        "stats": {
            "n_rows": len(mats),
            "n_dims_estimated": n_est,
            "avg_confidence": round(sum(confs) / len(confs), 3) if confs else 0.0,
            "total_weight_kg": round(sum(float(m.get("total_weight_kg") or 0) for m in mats), 3),
            "n_input_rows": clean.get("n_input_rows"),
            "n_skipped_total": clean.get("n_skipped_total"),
            "n_skip_noise_name": clean.get("n_skip_noise_name"),
            "n_skip_summary_row": clean.get("n_skip_summary_row"),
            "n_skip_zero_qty": clean.get("n_skip_zero_qty"),
            "n_skip_zero_placeholder": clean.get("n_skip_zero_placeholder"),
            "clean": clean,
        },
        "errors": [] if mats else ["no material rows parsed"],
    }


def parse_table_bytes(
    data: bytes,
    *,
    filename: str = "upload.csv",
    **kwargs: Any,
) -> Dict[str, Any]:
    """网关上传入口：bytes → 临时文件 → parse_table_file（与 CLI 同一路径）。"""
    import tempfile

    suf = Path(filename or "upload.csv").suffix.lower() or ".csv"
    if suf not in (".csv", ".tsv", ".txt", ".xlsx", ".xlsm", ".json"):
        suf = ".csv"
    tmp_path = None
    try:
        fd, tmp_path = tempfile.mkstemp(suffix=suf, prefix="tbl_")
        import os

        with os.fdopen(fd, "wb") as f:
            f.write(data)
        out = parse_table_file(tmp_path, **kwargs)
        out["path"] = filename or out.get("path")
        out["source"] = "upload"
        return out
    except Exception as e:
        return {
            "ok": False,
            "path": filename,
            "materials": [],
            "ir": [],
            "column_map": {},
            "stats": {"n_rows": 0},
            "errors": [f"{type(e).__name__}: {e}"],
        }
    finally:
        if tmp_path:
            try:
                Path(tmp_path).unlink(missing_ok=True)
            except Exception:
                pass


def parse_table_rows(
    rows: Sequence[Dict[str, Any]],
    *,
    headers: Optional[Sequence[Any]] = None,
    source: str = "api_rows",
) -> Dict[str, Any]:
    """已解析的字典行 → 与文件入口相同的返回形状。"""
    ir = rows_to_ir(list(rows), headers=headers, source=source)
    mats = ir_to_materials(ir)
    colmap = {}
    if ir:
        colmap = dict((ir[0].get("meta") or {}).get("column_map") or {})
    n_est = sum(1 for m in ir if (m.get("meta") or {}).get("dims_estimated"))
    clean = last_clean_stats()
    return {
        "ok": bool(mats),
        "path": "",
        "materials": mats,
        "ir": ir,
        "column_map": colmap,
        "stats": {
            "n_rows": len(mats),
            "n_dims_estimated": n_est,
            "total_weight_kg": round(sum(float(m.get("total_weight_kg") or 0) for m in mats), 3),
            "n_input_rows": clean.get("n_input_rows"),
            "n_skipped_total": clean.get("n_skipped_total"),
            "clean": clean,
        },
        "errors": [] if mats else ["no material rows"],
    }


# AUTONOMY_EXTRA_SYNONYMS_V2
for _std, _extra in {
    "name": ("货品名称", "中文品名", "商品名称", "material_name", "article"),
    "quantity": ("装箱数", "包装数量", "包装件数", "pcs"),
    "weight_kg": ("毛重kg", "毛重(kg)", "净重kg", "净重(kg)", "单件重量"),
    "part_no": ("商品编码", "条码", "barcode", "物料号"),
    "length_mm": ("外径", "总长", "全长"),
}.items():
    if _std in COLUMN_SYNONYMS:
        COLUMN_SYNONYMS[_std] = tuple(dict.fromkeys(list(COLUMN_SYNONYMS[_std]) + list(_extra)))
