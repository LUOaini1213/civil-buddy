"""Turn tender excerpts (text + tables + several files) into one blob for parse.

No vision OCR, no invented pages. Output is still matrix + P0, not a generated bid.
"""

from __future__ import annotations

import csv
import io
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

ALLOWED_EXT = (".txt", ".md", ".csv", ".tsv", ".docx", ".xlsx")
W_NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}


def suffix_of(name: str) -> str:
    return Path(name or "").suffix.lower()


def allowed_name(name: str) -> bool:
    return suffix_of(name) in ALLOWED_EXT


def decode_file(filename: str, raw: bytes) -> Dict[str, Any]:
    """One file → text. Tables become `|` rows so parse can copy exact_text."""
    name = Path(filename or "upload.txt").name
    ext = suffix_of(name)
    if ext not in ALLOWED_EXT:
        raise ValueError("只接受 .txt / .md / .csv / .tsv / .docx / .xlsx 节选，不解析扫描 PDF")
    if len(raw) > 2_000_000:
        raise ValueError("文件过大")
    n_tables = 0
    if ext in (".txt", ".md"):
        text = raw.decode("utf-8", errors="replace")
    elif ext == ".csv":
        text, n_tables = _csv_text(raw, ",")
    elif ext == ".tsv":
        text, n_tables = _csv_text(raw, "\t")
    elif ext == ".docx":
        text, n_tables = _docx_text(raw)
    else:
        text, n_tables = _xlsx_text(raw)
    text = text.replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        raise ValueError(f"{name} 抽不出文字")
    return {
        "filename": name,
        "ext": ext.lstrip("."),
        "text": text,
        "n_chars": len(text),
        "n_tables": n_tables,
    }


def merge_sections(parts: Iterable[Dict[str, Any]]) -> str:
    blocks: List[str] = []
    for p in parts:
        name = str(p.get("filename") or p.get("name") or "节选")
        body = str(p.get("text") or "").strip()
        if not body:
            continue
        blocks.append(f"## 文件：{name}\n{body}")
    return "\n\n".join(blocks).strip()


def ingest_files(files: List[Dict[str, Any]]) -> Dict[str, Any]:
    """files: [{filename, bytes|text}] → one blob + provenance."""
    decoded: List[Dict[str, Any]] = []
    for f in files:
        if f.get("text") and not f.get("bytes"):
            decoded.append(
                {
                    "filename": Path(str(f.get("filename") or "paste.txt")).name,
                    "ext": "txt",
                    "text": str(f["text"]),
                    "n_chars": len(str(f["text"])),
                    "n_tables": 0,
                }
            )
            continue
        decoded.append(decode_file(str(f.get("filename") or "upload.txt"), bytes(f.get("bytes") or b"")))
    blob = merge_sections(decoded)
    return {
        "schema": "tender.ingest.v1",
        "text": blob,
        "n_files": len(decoded),
        "n_chars": len(blob),
        "n_tables": sum(int(d.get("n_tables") or 0) for d in decoded),
        "files": [
            {"filename": d["filename"], "ext": d["ext"], "n_chars": d["n_chars"], "n_tables": d["n_tables"]}
            for d in decoded
        ],
    }


def ingest_from_json(body: Optional[Dict[str, Any]]) -> Optional[str]:
    """If the API body carries sections/files, merge them; else None."""
    body = body or {}
    sections = body.get("sections") or body.get("files")
    if isinstance(sections, list) and sections:
        parts = []
        for s in sections:
            if not isinstance(s, dict):
                continue
            text = s.get("text") or s.get("body") or ""
            if not str(text).strip():
                continue
            parts.append({"filename": s.get("name") or s.get("filename") or "节选", "text": text})
        if parts:
            return merge_sections(parts)
    texts = body.get("texts")
    if isinstance(texts, list) and texts:
        return merge_sections(
            {"filename": f"节选{i + 1}", "text": str(t)} for i, t in enumerate(texts) if str(t).strip()
        )
    return None


def _csv_text(raw: bytes, delim: str) -> tuple[str, int]:
    sample = raw.decode("utf-8-sig", errors="replace")
    if delim == "," and "\t" in sample and sample.count("\t") > sample.count(","):
        delim = "\t"
    reader = csv.reader(io.StringIO(sample), delimiter=delim)
    lines: List[str] = []
    n = 0
    for row in reader:
        cells = [str(c).strip() if c is not None else "" for c in row]
        while cells and not cells[-1]:
            cells.pop()
        if not any(cells):
            continue
        n += 1
        lines.append(" | ".join(cells))
    return "\n".join(lines), (1 if n else 0)


def _para_text(el: ET.Element) -> str:
    bits = [t.text or "" for t in el.findall(".//w:t", W_NS)]
    return "".join(bits).strip()


def _docx_text(raw: bytes) -> tuple[str, int]:
    with zipfile.ZipFile(io.BytesIO(raw)) as zf:
        try:
            xml = zf.read("word/document.xml")
        except KeyError as e:
            raise ValueError("docx 缺少 document.xml") from e
    root = ET.fromstring(xml)
    body = root.find("w:body", W_NS)
    if body is None:
        raise ValueError("docx 无正文")
    lines: List[str] = []
    n_tables = 0
    for child in list(body):
        tag = child.tag.rsplit("}", 1)[-1]
        if tag == "tbl":
            n_tables += 1
            for tr in child.findall("./w:tr", W_NS):
                cells = []
                for tc in tr.findall("./w:tc", W_NS):
                    cells.append(_para_text(tc))
                if any(cells):
                    lines.append(" | ".join(cells))
        elif tag == "p":
            t = _para_text(child)
            if t:
                lines.append(t)
    return "\n".join(lines), n_tables


def _xlsx_text(raw: bytes) -> tuple[str, int]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    lines: List[str] = []
    n_tables = 0
    try:
        for ws in wb.worksheets:
            sheet_rows = 0
            for row in ws.iter_rows(values_only=True):
                cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if not cells:
                    continue
                sheet_rows += 1
                lines.append(" | ".join(cells))
            if sheet_rows:
                n_tables += 1
    finally:
        wb.close()
    return "\n".join(lines), n_tables
