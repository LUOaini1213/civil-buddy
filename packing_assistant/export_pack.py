"""导出包：POR 装柜单 + 绑扎工单 → xlsx（+ 侧视路径索引）。"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional


def export_shipment_xlsx(
    state: Dict[str, Any],
    *,
    output_dir: str | Path = "output/exports",
    basename: Optional[str] = None,
) -> Dict[str, Any]:
    """
    写 xlsx，返回 {xlsx_path, sheets, por_rows, secure_rows}。
    """
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError as e:
        raise RuntimeError("需要 openpyxl: pip install openpyxl") from e

    plan = state.get("container_plan") or {}
    boxes = state.get("boxes") or []
    mats = state.get("materials") or []
    swo = state.get("secure_work_order")
    if not swo:
        from packing_assistant.tools.secure_work_order import build_secure_work_order

        swo = build_secure_work_order(plan, boxes)
    por = state.get("por_manifest")
    if not por:
        from packing_assistant.tools.por_manifest import build_por_manifest

        por = build_por_manifest(plan, boxes, mats)

    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    name = basename or f"shipment_{ts}"
    xlsx_path = out_dir / f"{name}.xlsx"

    wb = openpyxl.Workbook()
    # —— 摘要 ——
    ws0 = wb.active
    ws0.title = "摘要"
    bold = Font(bold=True)
    rows0 = [
        ("方案", state.get("packing_plan_id") or name),
        ("柜型", plan.get("container_type") or state.get("container_type") or "40HQ"),
        ("订柜N0", plan.get("n0")),
        ("3D用柜", plan.get("containers_used")),
        ("can_fit", plan.get("can_fit")),
        ("ship_ok", state.get("ship_ok")),
        ("worst_mid50", plan.get("worst_mid50")),
        ("team_mode", state.get("team_mode")),
        ("说明", "订舱看N0；工程看3D用柜；绑扎WARN不拦出运"),
    ]
    for i, (k, v) in enumerate(rows0, 1):
        ws0.cell(i, 1, k).font = bold
        ws0.cell(i, 2, str(v) if v is not None else "")

    # —— POR by part ——
    ws1 = wb.create_sheet("POR_by_part")
    ws1.append(["part_no", "total_kg", "n_boxes", "containers"])
    for p in por.get("by_part") or []:
        ws1.append(
            [
                p.get("part_no"),
                p.get("total_kg"),
                p.get("n_boxes"),
                ",".join(str(c) for c in (p.get("containers") or [])),
            ]
        )

    # —— POR by container ——
    ws2 = wb.create_sheet("POR_by_container")
    ws2.append(["container_no", "box_id", "part_no", "name", "weight_kg"])
    for c in por.get("by_container") or []:
        cno = c.get("container_no")
        for r in c.get("rows") or []:
            ws2.append(
                [
                    cno,
                    r.get("box_id"),
                    r.get("part_no"),
                    r.get("name"),
                    r.get("weight_kg"),
                ]
            )

    # —— 绑扎工单 ——
    ws3 = wb.create_sheet("绑扎空隙工单")
    ws3.append(
        ["seq", "type", "severity", "container_no", "box_id", "action", "material"]
    )
    for it in swo.get("items") or []:
        ws3.append(
            [
                it.get("seq"),
                it.get("type"),
                it.get("severity"),
                it.get("container_no"),
                it.get("box_id"),
                it.get("action"),
                it.get("material"),
            ]
        )

    # —— 侧视路径 ——
    ws4 = wb.create_sheet("侧视路径")
    ws4.append(["kind", "path"])
    img = state.get("image_data") or {}
    if img.get("side") and isinstance(img["side"], dict):
        ws4.append(["side", img["side"].get("path")])
    if img.get("side_overview"):
        ws4.append(["overview", img.get("side_overview")])
    for p in img.get("side_per_container") or []:
        if isinstance(p, dict):
            ws4.append([f"c{p.get('container_no')}", p.get("path")])

    wb.save(xlsx_path)
    meta = {
        "xlsx_path": str(xlsx_path),
        "sheets": ["摘要", "POR_by_part", "POR_by_container", "绑扎空隙工单", "侧视路径"],
        "por_summary": por.get("summary"),
        "secure_summary": swo.get("summary"),
        "n_por_parts": len(por.get("by_part") or []),
        "n_secure_items": len(swo.get("items") or []),
    }
    (out_dir / f"{name}_meta.json").write_text(
        json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return meta
