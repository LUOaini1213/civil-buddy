"""Authorized job folder + Office interchange (WorkBuddy local-file slice).

NL run writes real .xlsx next to table drafts so Excel can open them.
If the user names an existing workbook in CIVIL_JOB_ROOT, patch only CB草稿-*
sheets and leave the owner's sheets alone.
Not a desktop shell. Not D:\\layout. Not COM into an open Excel window.
"""

from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Any, Dict, List, Tuple

JOB_EXTS = {".xlsx", ".csv", ".txt", ".md", ".json", ".docx", ".log"}
JOB_MAX_FILES = 12
JOB_FILE_CHARS = 8_000
JOB_TOTAL_CHARS = 48_000
DRAFT_PREFIX = "CB草稿"

FORBIDDEN_LAYOUT = ("d:\\layout", "d:/layout")


def is_forbidden_layout(path: Path) -> bool:
    n = str(path).replace("/", "\\").rstrip("\\").lower()
    return n == "d:\\layout" or n.startswith("d:\\layout\\")


def job_root() -> Path:
    raw = (os.getenv("CIVIL_JOB_ROOT") or "").strip()
    if raw:
        p = Path(raw).expanduser()
        if not is_forbidden_layout(p):
            return p
    return Path.cwd() / ".civil-buddy" / "out"


def _sheet_name(title: str, used: set) -> str:
    t = re.sub(r'[:\\/?*\[\]]', " ", title or "表").strip() or "表"
    t = t[:31]
    base = t
    i = 2
    while t in used:
        suffix = f"_{i}"
        t = (base[: 31 - len(suffix)] + suffix)
        i += 1
    used.add(t)
    return t


def tables_from_md(md: str) -> List[Tuple[str, List[List[str]]]]:
    """Return (sheet_name, rows) for each markdown table. Caption from last heading."""
    lines = (md or "").splitlines()
    heading = "表"
    used: set = set()
    out: List[Tuple[str, List[List[str]]]] = []
    i = 0
    while i < len(lines):
        raw = lines[i].rstrip()
        hs = raw.lstrip()
        if hs.startswith("#"):
            heading = hs.lstrip("#").strip() or heading
            i += 1
            continue
        if hs.startswith("|") and i + 1 < len(lines) and re.match(r"^\s*\|?\s*:?-{3,}", lines[i + 1]):
            rows: List[List[str]] = []
            while i < len(lines) and lines[i].lstrip().startswith("|"):
                cells = [c.strip() for c in lines[i].strip().strip("|").split("|")]
                if cells and not all(re.match(r"^:?-{3,}:?$", c or "") for c in cells):
                    rows.append(cells)
                i += 1
            if rows:
                out.append((_sheet_name(heading, used), rows))
            continue
        i += 1
    return out


def write_xlsx(path: Path, sheets: List[Tuple[str, List[List[str]]]]) -> Path:
    from io import BytesIO

    from packing_assistant.sandbox import guarded_write_bytes

    import openpyxl

    wb = openpyxl.Workbook()
    first = True
    for name, rows in sheets:
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = name[:31] or "表"
        for r_i, row in enumerate(rows, 1):
            for c_i, val in enumerate(row, 1):
                ws.cell(r_i, c_i, val)
    bio = BytesIO()
    wb.save(bio)
    return guarded_write_bytes(Path(path), bio.getvalue())


def _query_from_md(md: str) -> str:
    if "## 用户原文" in (md or ""):
        return (md or "").split("## 用户原文", 1)[1].split("##", 1)[0].strip()
    return (md or "")[:400]


def pick_job_xlsx(query: str) -> Path | None:
    """Existing job-root workbook the user named. Do not guess the first file."""
    q = (query or "").lower()
    if not q:
        return None
    for f in list_job_files():
        if f.get("suffix") != ".xlsx":
            continue
        name = str(f.get("name") or "")
        stem = Path(name).stem.lower()
        if name.lower() in q or (stem and stem in q):
            return Path(str(f["path"]))
    return None


def patch_xlsx(path: Path, sheets: List[Tuple[str, List[List[str]]]]) -> Path:
    """Replace only CB草稿-* sheets. Owner sheets stay."""
    from io import BytesIO

    from packing_assistant.sandbox import guarded_write_bytes

    import openpyxl

    p = Path(path)
    root = job_root().resolve()
    try:
        resolved = p.resolve()
        resolved.relative_to(root)
    except (OSError, ValueError) as e:
        raise PermissionError("job file outside authorized root") from e
    if is_forbidden_layout(resolved):
        raise PermissionError("D:\\layout denied")
    wb = openpyxl.load_workbook(resolved)
    for name in list(wb.sheetnames):
        if name.startswith(DRAFT_PREFIX):
            del wb[name]
    used = set(wb.sheetnames)
    for title, rows in sheets:
        ws = wb.create_sheet(_sheet_name(f"{DRAFT_PREFIX}-{title}", used))
        for r_i, row in enumerate(rows, 1):
            for c_i, val in enumerate(row, 1):
                ws.cell(r_i, c_i, val)
    bio = BytesIO()
    wb.save(bio)
    return guarded_write_bytes(resolved, bio.getvalue())


def export_md_to_xlsx(md_path: Path, query: str = "") -> List[Path]:
    """Sibling xlsx always. If the user named a job-root workbook, patch it too."""
    p = Path(md_path)
    if not p.is_file() or p.suffix.lower() != ".md":
        return []
    text = p.read_text(encoding="utf-8", errors="ignore")
    sheets = tables_from_md(text)
    if not sheets:
        return []
    written: List[Path] = []
    sibling = p.with_suffix(".xlsx")
    written.append(write_xlsx(sibling, sheets))
    if not job_root_granted():
        return written
    q = query or _query_from_md(text)
    target = pick_job_xlsx(q)
    try:
        if target is not None:
            written.append(patch_xlsx(target, sheets))
        else:
            dest = job_root() / sibling.name
            if dest.resolve() != sibling.resolve():
                written.append(write_xlsx(dest, sheets))
    except (OSError, PermissionError, RuntimeError):
        pass
    return written


def job_root_granted() -> bool:
    raw = (os.getenv("CIVIL_JOB_ROOT") or "").strip()
    if not raw:
        return False
    p = Path(raw).expanduser()
    return p.is_dir() and not is_forbidden_layout(p)


def list_job_files() -> List[Dict[str, Any]]:
    """Files in the authorized job folder. Empty if CIVIL_JOB_ROOT is unset."""
    if not job_root_granted():
        return []
    root = job_root()
    rows: List[Dict[str, Any]] = []
    try:
        names = sorted(root.iterdir(), key=lambda p: p.name.lower())
    except OSError:
        return []
    for p in names:
        if not p.is_file() or p.suffix.lower() not in JOB_EXTS:
            continue
        rows.append(
            {
                "name": p.name,
                "path": str(p),
                "suffix": p.suffix.lower(),
                "bytes": p.stat().st_size if p.exists() else 0,
            }
        )
        if len(rows) >= JOB_MAX_FILES:
            break
    return rows


def _read_xlsx_text(path: Path, limit: int) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    lines: List[str] = []
    for ws in wb.worksheets:
        lines.append(f"# {ws.title}")
        for row in ws.iter_rows(max_row=80, max_col=16, values_only=True):
            cells = ["" if c is None else str(c).strip() for c in row]
            if any(cells):
                lines.append(" | ".join(cells))
        if sum(len(x) for x in lines) >= limit:
            break
    wb.close()
    return "\n".join(lines)[:limit]


def _read_docx_text(path: Path, limit: int) -> str:
    import zipfile
    from xml.etree import ElementTree as ET

    with zipfile.ZipFile(path) as z:
        xml = z.read("word/document.xml")
    root_el = ET.fromstring(xml)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    texts = [t.text or "" for t in root_el.findall(".//w:t", ns)]
    return " ".join(t.strip() for t in texts if t.strip())[:limit]


def read_job_file(path: Path, limit: int = JOB_FILE_CHARS) -> str:
    p = Path(path)
    root = job_root().resolve()
    try:
        p = p.resolve()
        p.relative_to(root)
    except (OSError, ValueError) as e:
        raise PermissionError("job file outside authorized root") from e
    if is_forbidden_layout(p):
        raise PermissionError("D:\\layout denied")
    suf = p.suffix.lower()
    if suf == ".xlsx":
        return _read_xlsx_text(p, limit)
    if suf == ".docx":
        return _read_docx_text(p, limit)
    return p.read_text(encoding="utf-8", errors="ignore")[:limit]


def job_files_blob(query: str = "") -> str:
    """Text of job-root files to prepend on run. Prefer names mentioned in query."""
    files = list_job_files()
    if not files:
        return ""
    q = (query or "").lower()
    named = [f for f in files if f["name"].lower() in q or Path(f["name"]).stem.lower() in q]
    pick = named or files
    chunks: List[str] = ["## 作业根文件（授权文件夹，未再上传）"]
    used = 0
    for f in pick:
        room = JOB_TOTAL_CHARS - used
        if room < 80:
            chunks.append(f"（还有 {f['name']} 未贴全文）")
            continue
        try:
            body = read_job_file(Path(f["path"]), min(JOB_FILE_CHARS, room))
        except (OSError, PermissionError, RuntimeError):
            chunks.append(f"### {f['name']}\n（读失败）")
            continue
        block = f"### {f['name']}\n{body}"
        chunks.append(block)
        used += len(block)
    return "\n\n".join(chunks)
