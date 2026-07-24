#!/usr/bin/env python3
"""
将 Material_Summary 进度表转为材料行（尺寸经验估算），用本项目流水线粗算柜数。

用法:
  python scripts/run_material_summary_pack.py
  python scripts/run_material_summary_pack.py --xlsx "C:/Users/.../Material_Summary (3).xlsx"
"""

from __future__ import annotations

import argparse
import json
import sys
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

# 物料组 → 代表件尺寸/单重（幕墙实务粗估，仅用于程序试拼，非正式装箱尺寸）
# (length_mm, width_mm, height_mm, weight_kg, unit_label)
GROUP_PROFILE: Dict[str, Tuple[float, float, float, float, str]] = {
    "11—铝板": (1500, 900, 30, 12.0, "铝板片"),
    "22—铝材": (5800, 80, 60, 9.5, "铝型材支"),
    "13—铁件": (2000, 200, 150, 18.0, "铁件"),
    "14—不锈钢": (800, 200, 80, 3.5, "不锈钢件"),
    "19—胶条、垫块、胶皮": (600, 400, 300, 12.0, "胶条箱当量"),
    "23—紧固件/螺丝": (400, 300, 250, 8.0, "五金箱当量"),
    "24—Glass 玻璃": (1600, 1200, 80, 48.0, "玻璃片含垫"),
    "25—门窗锁配件": (500, 350, 250, 6.0, "五金箱当量"),
    "28—杂项配件": (1200, 800, 100, 15.0, "杂项板件"),
    "18—结构胶/耐候胶": (400, 300, 300, 20.0, "胶桶箱"),
    "17—拉爆螺栓，化学螺栓": (400, 300, 250, 10.0, "螺栓箱"),
}

# 胶条/螺丝按「箱当量」折算件数，避免上万件压爆合箱
BULK_PER_CARTON: Dict[str, int] = {
    "19—胶条、垫块、胶皮": 50,
    "23—紧固件/螺丝": 200,
    "25—门窗锁配件": 40,
    "17—拉爆螺栓，化学螺栓": 100,
}


def _f(x: Any) -> float:
    try:
        if x is None or x == "":
            return 0.0
        return float(x)
    except Exception:
        return 0.0


def _s(x: Any) -> str:
    return str(x or "").strip()


def should_include(row: Dict[str, Any]) -> Tuple[bool, str]:
    """筛 8 月待拼：有量、未装柜、或明确待拼柜。已装柜/已出货剔除。"""
    qty = _f(row.get("总数"))
    arr = _f(row.get("已到货数量"))
    pend = _f(row.get("未到货数量"))
    note = _s(row.get("生产情况备注"))
    pack = row.get("实际装柜日期")
    cntr = _s(row.get("货柜号"))
    desc = _s(row.get("項目描述"))
    dest = _s(row.get("收货  目的地"))

    if qty <= 0 and arr <= 0 and pend <= 0:
        return False, "零量"
    if "错误色号" in desc or "不下单" in desc:
        return False, "不下单"
    if cntr:
        return False, f"已装柜:{cntr}"
    if pack and ("已出货" in note or "已生产完成，已出货" in note):
        return False, "已出货"
    if pack and not note:
        # 7 月已有装柜日的木地板/瓦楞板等
        return False, f"已有装柜日:{pack}"

    # 有效数量：优先待发（未到货>0 或 已到货在库）
    eff = max(qty, arr, pend)
    if eff <= 0:
        return False, "无效量"

    # 明确待拼 / 工厂在库 / 未到货待发
    if "待拼柜" in note or "待出货" in note:
        return True, "待拼柜"
    if dest in ("工厂", "工地") and (arr > 0 or pend > 0 or qty > 0):
        return True, "在途或在库"
    if note and "提前一周" in note:
        return False, "胶类待通知"
    return True, "有量"


def effective_qty(row: Dict[str, Any]) -> float:
    qty = _f(row.get("总数"))
    arr = _f(row.get("已到货数量"))
    pend = _f(row.get("未到货数量"))
    # 超收用已到货
    if arr > qty > 0:
        return arr
    if arr > 0 and pend <= 0:
        return arr
    if pend > 0:
        return pend if pend > 0 else qty
    return max(qty, arr, pend)


def to_materials(xlsx: Path) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # row0 双表头, row1 列名
    headers = [_s(h) for h in rows[1]]
    mats: List[Dict[str, Any]] = []
    skipped: List[Dict[str, Any]] = []
    mid = 0

    for raw in rows[2:]:
        if raw[0] is None:
            continue
        d = {headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers))}
        ok, reason = should_include(d)
        desc = _s(d.get("項目描述"))
        group = _s(d.get("物料组描述"))
        por = _s(d.get("訂貨單 / 加工圖號")) or _s(d.get("施工批次"))
        if not ok:
            skipped.append(
                {
                    "seq": d.get("序號"),
                    "por": por,
                    "group": group,
                    "reason": reason,
                    "qty": _f(d.get("总数")),
                }
            )
            continue

        n = effective_qty(d)
        if n <= 0:
            continue

        # 玻璃/铝板等按件；螺丝/胶条按箱当量
        per = BULK_PER_CARTON.get(group, 1)
        if per > 1:
            units = max(1, int(round(n / per)))
            piece_name_suffix = f"×{per}/箱当量"
        else:
            units = max(1, int(round(n)))
            piece_name_suffix = ""

        prof = GROUP_PROFILE.get(group)
        if not prof:
            # 默认中件
            L, W, H, wt, unit = 1000.0, 500.0, 200.0, 10.0, "通用件"
        else:
            L, W, H, wt, unit = prof

        # 吊具略放大
        if "吊具" in desc:
            L, W, H, wt = 2000.0, 800.0, 600.0, 45.0
            unit = "吊具"
            units = max(1, int(round(n)))

        # 木板/展板
        if "木板" in desc or "展板" in desc:
            L, W, H, wt = 2400.0, 1200.0, 18.0, 22.0
            unit = "木板"
            units = max(1, int(round(n)))

        mid += 1
        dest = _s(d.get("收货  目的地")) or "未填"
        name = f"{unit}{piece_name_suffix} | {por or desc[:40]}"
        mats.append(
            {
                "id": f"M{mid:03d}",
                "name": name[:80],
                "quantity": units,
                "weight_kg": wt,
                "total_weight_kg": round(wt * units, 2),
                "length_mm": L,
                "width_mm": W,
                "height_mm": H,
                "spec": group,
                "part_no": por,
                "note": f"dest={dest}; raw_qty={n}; {_s(d.get('生产情况备注'))[:40]}; dims=estimated",
                "source_sheet": "Material_Summary",
                "destination": dest,
            }
        )

    wb.close()
    return mats, skipped


def save_materials_xlsx(mats: List[Dict[str, Any]], path: Path) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "materials"
    cols = [
        "id",
        "name",
        "spec",
        "quantity",
        "weight_kg",
        "total_weight_kg",
        "length_mm",
        "width_mm",
        "height_mm",
        "part_no",
        "source_sheet",
        "note",
    ]
    ws.append(cols)
    for m in mats:
        ws.append([m.get(c) for c in cols])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def run_pack(mats: List[Dict[str, Any]], container: str = "40HQ") -> Dict[str, Any]:
    from packing_assistant.harness import apply_user_confirmation, run_team_a, run_team_b

    net = sum(float(m.get("total_weight_kg") or 0) for m in mats)
    # 体积累加粗估柜数下限
    vol = 0.0
    for m in mats:
        q = float(m.get("quantity") or 1)
        vol += (
            float(m.get("length_mm") or 0)
            * float(m.get("width_mm") or 0)
            * float(m.get("height_mm") or 0)
            * q
            / 1e9
        )
    guess = min(max(int(vol / 28) + 1, int(net / 18000) + 1, 1), 16)

    t0 = time.time()
    sa = run_team_a("Material_Summary 8月待拼粗算", materials=mats)
    state = None
    used_mc = guess
    for mc in range(guess, 17):
        sb = apply_user_confirmation(
            sa, action="confirm", container_type=container, max_containers=mc
        )
        state = run_team_b(sb)
        used_mc = mc
        if (state.get("container_plan") or {}).get("can_fit"):
            break
    ms = int((time.time() - t0) * 1000)
    plan = (state or {}).get("container_plan") or {}
    risk = (state or {}).get("risk_report") or {}
    boxes = (state or {}).get("boxes") or []
    return {
        "ok": True,
        "ms": ms,
        "materials": len(mats),
        "net_kg": round(net, 1),
        "est_raw_volume_m3": round(vol, 2),
        "guess_containers": guess,
        "max_containers_tried": used_mc,
        "boxes": len(boxes),
        "box_types": [b.get("box_type") for b in boxes],
        "box_summary": [
            {
                "box_id": b.get("box_id"),
                "box_type": b.get("box_type"),
                "outer": b.get("outer_size_mm"),
                "net_kg": b.get("net_kg") or b.get("net_weight_kg"),
                "gross_kg": b.get("gross_kg") or b.get("gross_weight_kg"),
                "structure": b.get("structure_conclusion"),
            }
            for b in boxes[:40]
        ],
        "can_fit": plan.get("can_fit"),
        "containers_used": plan.get("containers_used"),
        "space_utilization": plan.get("space_utilization"),
        "space_best": plan.get("space_utilization_best_container"),
        "floor_avg": plan.get("floor_utilization_avg"),
        "weight_utilization": plan.get("weight_utilization"),
        "cargo_solid_volume_m3": plan.get("cargo_solid_volume_m3"),
        "engine": plan.get("engine"),
        "risk_level": risk.get("level"),
        "risk_score": risk.get("compliance_score"),
        "container_type": container,
        "packing_plan_id": (state or {}).get("packing_plan_id"),
        "final_response": (state or {}).get("final_response") or "",
        "image_data": (state or {}).get("image_data") or {},
        "layout_n": len((plan.get("layout") or [])),
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--xlsx",
        default=r"C:\Users\wenjie.luo\Downloads\Material_Summary (3).xlsx",
    )
    ap.add_argument("--container", default="40HQ", choices=["20GP", "40GP", "40HQ"])
    ap.add_argument("--also-40gp", action="store_true", help="再跑一版 40GP 对比")
    args = ap.parse_args()

    src = Path(args.xlsx)
    if not src.exists():
        print("文件不存在:", src)
        return 1

    mats, skipped = to_materials(src)
    out_dir = ROOT / "output" / "aug_material_summary"
    out_dir.mkdir(parents=True, exist_ok=True)
    mat_xlsx = out_dir / "materials_estimated.xlsx"
    save_materials_xlsx(mats, mat_xlsx)

    print(f"源表: {src.name}")
    print(f"纳入材料行: {len(mats)}  | 跳过: {len(skipped)}")
    print(f"估算材料表: {mat_xlsx}")
    net = sum(m["total_weight_kg"] for m in mats)
    print(f"估算总净重: {net:.1f} kg")
    print("--- 纳入明细 ---")
    for m in mats:
        print(
            f"  {m['id']} q={m['quantity']:5d}  {m['length_mm']:.0f}x{m['width_mm']:.0f}x{m['height_mm']:.0f}  "
            f"{m['total_weight_kg']:8.1f}kg  {m['name'][:60]}"
        )

    results = {}
    for ct in ([args.container] + (["40GP"] if args.also_40gp and args.container != "40GP" else [])):
        print(f"\n=== 拼柜试算 container={ct} ===")
        r = run_pack(mats, container=ct)
        results[ct] = {k: v for k, v in r.items() if k not in ("final_response", "image_data", "box_summary")}
        results[ct]["box_summary"] = r.get("box_summary")
        results[ct]["final_preview"] = (r.get("final_response") or "")[:800]
        print(
            f"  can_fit={r['can_fit']}  containers={r['containers_used']}  "
            f"boxes={r['boxes']}  space={r.get('space_utilization')}  "
            f"floor={r.get('floor_avg')}  weight={r.get('weight_utilization')}  "
            f"risk={r.get('risk_level')}/{r.get('risk_score')}  ms={r['ms']}"
        )
        print((r.get("final_response") or "")[:600])

    report = {
        "source": str(src),
        "note": "尺寸/单重为物料组经验估算，非正式提料尺寸；用于程序粗算柜数",
        "materials_count": len(mats),
        "skipped_count": len(skipped),
        "skipped": skipped,
        "materials": mats,
        "results": results,
    }
    rep_path = out_dir / "pack_report.json"
    rep_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n报告: {rep_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
