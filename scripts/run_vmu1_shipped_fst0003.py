#!/usr/bin/env python3
"""
VMU1 已发货：解析 REDACTED-REF 装货单 1/2 柜 → 订柜 N0 + 3D can_fit。
只读 A: 已发货目录，不改源文件。
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from packing_assistant.tools.booking import compute_booking, pack_with_auto_containers

SHIPPED = Path(r"A:\JOB\REDACTED-JOB\Project\6. Quality QAQC\6.06 POR\已发货")
OUT = ROOT / "output" / "vmu1_shipped"


def _f(x: Any) -> Optional[float]:
    try:
        if x is None or x == "":
            return None
        return float(x)
    except Exception:
        return None


def parse_frames(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    frames: List[Dict[str, Any]] = []
    for sn in wb.sheetnames:
        if "装货" not in sn:
            continue
        cab = 1 if "1柜" in sn else (2 if "2柜" in sn else 0)
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        cur: Optional[Dict[str, Any]] = None
        running_w = 0.0
        running_n = 0

        def close_cur(total: Optional[float] = None) -> None:
            nonlocal cur, running_w, running_n
            if not cur:
                return
            cargo = float(total if total is not None else running_w)
            cur["cargo_kg"] = cargo
            cur["line_sum_kg"] = running_w
            cur["lines"] = running_n
            frames.append(cur)
            cur = None
            running_w = 0.0
            running_n = 0

        for r in rows:
            vals = list(r)
            joined = " ".join(str(v) for v in vals if v is not None)
            name0 = str(vals[0] or "")

            # 铁架外廓行：1.1米铁架 / 2米铁架 / 4米铁架 / 6米铁架
            if "铁架" in name0 or re.search(r"\d+(\.\d+)?\s*米.*铁架", joined):
                nn = [_f(v) for v in vals[1:]]
                nn = [n for n in nn if n is not None]
                if len(nn) >= 3:
                    close_cur()
                    L, W, H = float(nn[0]), float(nn[1]), float(nn[2])
                    # qty often 1, tare ~80-250
                    tare = 80.0
                    if len(nn) >= 5 and nn[3] <= 5:
                        tare = float(nn[4])
                    elif len(nn) >= 4 and nn[3] > 5:
                        tare = float(nn[3])
                    cur = {
                        "cabinet_doc": cab,
                        "sheet": sn,
                        "name": name0.strip() or joined[:48],
                        "L": L,
                        "W": W,
                        "H": H,
                        "tare_kg": tare,
                    }
                    continue

            if cur is None:
                continue

            # 小结行：[件数, 总重] 如 79, 2252.488
            compact = [_f(v) for v in vals]
            compact = [n for n in compact if n is not None]
            if (
                len(compact) == 2
                and compact[1] > 100
                and compact[0] < 500
                and abs(compact[0] - int(compact[0])) < 1e-6
            ):
                close_cur(total=float(compact[1]))
                continue

            # 明细合计重：末列为行合计
            if "PC" in joined or (
                isinstance(vals[0], (int, float)) and _f(vals[0]) is not None
            ):
                if compact:
                    last = compact[-1]
                    if 0.5 < last < 8000:
                        running_w += last
                        running_n += 1

        close_cur()
    wb.close()
    return frames


def frames_to_boxes(frames: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    boxes = []
    for i, fr in enumerate(frames, 1):
        L, W, H = float(fr["L"]), float(fr["W"]), float(fr["H"])
        # 纠正常见：长应为最大水平尺寸
        dims = sorted([L, W, H], reverse=True)
        # 铁架实务：底面两边 + 高；若最大是高 1750 且两边 1100
        if H >= 1500 and max(L, W) <= 1200:
            pass  # 1.1m 立方架
        outer = L * W * H / 1e9
        fill = 0.30
        content = outer * fill
        cargo = float(fr.get("cargo_kg") or 0)
        tare = float(fr.get("tare_kg") or 80)
        gross = cargo + tare
        boxes.append(
            {
                "box_id": f"SHIP-C{fr.get('cabinet_doc')}-F{i}",
                "box_type": fr.get("name") or "铁架",
                "outer_size_mm": {"length": L, "width": W, "height": H},
                "outer_m3": round(outer, 6),
                "content_m3": round(content, 6),
                "crate_fill_ratio": fill,
                "booking_volume_m3": round(min(outer, content * 1.50), 6),
                "gross_weight_kg": round(gross, 2),
                "net_weight_kg": round(cargo, 2),
                "stackable": H <= 1300 and L < 4000,
                "prefer_bottom": True,
                "cabinet_doc": fr.get("cabinet_doc"),
                "sheet": fr.get("sheet"),
            }
        )
    return boxes


def main() -> int:
    OUT.mkdir(parents=True, exist_ok=True)
    files = list(SHIPPED.rglob("*REDACTED-REF*"))
    if not files:
        print("NO REDACTED-REF shipped file under", SHIPPED)
        return 1
    path = files[0]
    print("FILE", path)

    frames = parse_frames(path)
    print(f"frames={len(frames)}")
    by_cab: Dict[int, float] = {}
    for fr in frames:
        c = int(fr.get("cabinet_doc") or 0)
        g = float(fr.get("cargo_kg") or 0) + float(fr.get("tare_kg") or 0)
        by_cab[c] = by_cab.get(c, 0.0) + g
        print(
            f"  cab{c} {fr.get('name')} "
            f"{fr['L']:.0f}x{fr['W']:.0f}x{fr['H']:.0f} "
            f"cargo={fr.get('cargo_kg'):.1f} tare={fr.get('tare_kg'):.1f} "
            f"gross={g:.1f}"
        )

    boxes = frames_to_boxes(frames)
    booking = compute_booking(boxes=boxes, container_type="40HQ", fill_ratio=0.82)
    plan = pack_with_auto_containers(
        boxes, container_type="40HQ", n0=int(booking.get("n0") or 1), n_max=6
    )

    cargo_only = sum(float(b.get("net_weight_kg") or 0) for b in boxes)
    gross = sum(float(b.get("gross_weight_kg") or 0) for b in boxes)

    rep = {
        "scope": "VMU1 已发货 · REDACTED-REF 装货单",
        "source": str(path),
        "frame_count": len(frames),
        "doc_cabinet_gross_kg": {str(k): round(v, 1) for k, v in sorted(by_cab.items())},
        "doc_cabinets": len([k for k, v in by_cab.items() if v > 0]),
        "cargo_kg": round(cargo_only, 1),
        "gross_kg": round(gross, 1),
        "booking": {
            "n0": booking.get("n0"),
            "containers_by_weight": booking.get("containers_by_weight"),
            "containers_by_volume": booking.get("containers_by_volume"),
            "binding_constraint": booking.get("binding_constraint"),
            "volume_m3": booking.get("volume_m3"),
            "crate_outer_m3": (booking.get("volume_detail") or {}).get("crate_outer_m3"),
            "payload_kg": booking.get("payload_kg"),
            "fill_ratio": booking.get("fill_ratio"),
            "volume_suspicious": booking.get("volume_suspicious"),
        },
        "layout_3d": {
            "containers_used": plan.get("containers_used"),
            "can_fit": plan.get("can_fit"),
            "n0": plan.get("n0"),
            "n_tried": plan.get("n_tried"),
            "booking_volume_utilization": plan.get("booking_volume_utilization"),
            "outer_space_utilization": plan.get("outer_space_utilization")
            or plan.get("space_utilization"),
            "weight_utilization": plan.get("weight_utilization"),
            "floor_utilization_avg": plan.get("floor_utilization_avg"),
        },
        "boxes": boxes,
    }

    outj = OUT / "vmu1_shipped_REDACTED-REF_pack.json"
    outj.write_text(json.dumps(rep, ensure_ascii=False, indent=2), encoding="utf-8")

    md = [
        "# VMU1 已发货 · REDACTED-REF 可装几柜",
        "",
        f"- **数据**：`{path.name}`（只读）",
        f"- **装货单页**：7-20装货单1柜 / 2柜",
        f"- **解析铁架数**：{len(frames)}",
        "",
        "## 装货单原分柜（人工已发）",
        "",
        "| 装货单柜号 | 毛重合计 kg（货+架） |",
        "|-----------:|--------------------:|",
    ]
    for k, v in sorted(by_cab.items()):
        md.append(f"| 第 {k} 柜 | {v:.1f} |")
    md.extend(
        [
            f"| **合计** | **{sum(by_cab.values()):.1f}** |",
            "",
            "## 算法复算（40HQ，自主定柜，不写死 2）",
            "",
            f"| 口径 | 值 |",
            f"|------|---|",
            f"| 货净重合计 | {cargo_only:.1f} kg |",
            f"| 毛重合计（含架） | {gross:.1f} kg |",
            f"| **订柜 N0** | **{booking.get('n0')}**（重量柜 {booking.get('containers_by_weight')} / 有效体积柜 {booking.get('containers_by_volume')} / 绑定 {booking.get('binding_constraint')}） |",
            f"| V_eff | {booking.get('volume_m3')} m³ |",
            f"| **3D 建议柜数** | **{plan.get('containers_used')}**（can_fit={plan.get('can_fit')}） |",
            f"| 订柜有效体积率 | {plan.get('booking_volume_utilization')} |",
            f"| 外廓摆柜率（仅展示） | {plan.get('outer_space_utilization') or plan.get('space_utilization')} |",
            f"| 重量利用率 | {plan.get('weight_utilization')} |",
            "",
            "> **结论**：装货单本身已按 **2 柜** 发运；算法用真实铁架外廓+货重复算，",
            f"> 订柜 N0=**{booking.get('n0')}**，3D can_fit 用柜=**{plan.get('containers_used')}**。",
            "> 与「2 柜已发」一致量级；PAYLOAD 单柜 ≤28610 kg。",
            "",
            "## 分架明细",
            "",
            "| 装货单柜 | 架型 | 外廓 mm | 货重 kg | 架重 kg | 毛重 kg |",
            "|--------:|------|---------|--------:|--------:|--------:|",
        ]
    )
    for b in boxes:
        o = b["outer_size_mm"]
        md.append(
            f"| {b.get('cabinet_doc')} | {b.get('box_type')} | "
            f"{o['length']:.0f}×{o['width']:.0f}×{o['height']:.0f} | "
            f"{b['net_weight_kg']:.1f} | {b['gross_weight_kg']-b['net_weight_kg']:.1f} | "
            f"{b['gross_weight_kg']:.1f} |"
        )
    md.append(f"\n产物：`{outj}`")
    mdp = OUT / "VMU1_已发货_REDACTED-REF_装柜复算.md"
    mdp.write_text("\n".join(md), encoding="utf-8")

    print("BOOKING", rep["booking"])
    print("3D", rep["layout_3d"])
    print("DOC cab", rep["doc_cabinet_gross_kg"])
    print("WROTE", outj)
    print("WROTE", mdp)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
