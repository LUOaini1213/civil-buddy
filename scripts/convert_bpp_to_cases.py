#!/usr/bin/env python3
"""
公开 3D-BPP / 集装箱装载样例 → 本项目 JSON + Excel。

支持:
  - D-Wave sample_data_*.txt
  - 简单 tsv/csv: qty,L,W,H
  - 手写预置 case A/B/C

用法:
  python scripts/convert_bpp_to_cases.py
  python scripts/convert_bpp_to_cases.py --input data/external/sample_data_1.txt
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))


def parse_dwave_txt(text: str) -> Dict[str, Any]:
    """
    解析 D-Wave 3d-bin-packing 的 sample_data_*.txt
    # Max num of bins : 1
    # Bin dimensions (L * W * H): 30 30 50
    case_id quantity length width height
    """
    max_bins = 1
    bin_lwh: Optional[Tuple[float, float, float]] = None
    items: List[Dict[str, Any]] = []

    for line in text.splitlines():
        s = line.strip()
        if not s:
            continue
        m = re.search(r"Max num of bins\s*:\s*(\d+)", s, re.I)
        if m:
            max_bins = int(m.group(1))
            continue
        m = re.search(
            r"Bin dimensions.*?:\s*([\d.]+)\s+([\d.]+)\s+([\d.]+)",
            s,
            re.I,
        )
        if m:
            bin_lwh = (float(m.group(1)), float(m.group(2)), float(m.group(3)))
            continue
        if s.startswith("#") or s.startswith("-") or "case_id" in s.lower():
            continue
        # data row: id qty L W H
        parts = re.split(r"\s+", s)
        if len(parts) >= 5:
            try:
                cid = parts[0]
                qty = int(float(parts[1]))
                L, W, H = float(parts[2]), float(parts[3]), float(parts[4])
                items.append(
                    {
                        "case_id": cid,
                        "quantity": qty,
                        "length": L,
                        "width": W,
                        "height": H,
                    }
                )
            except ValueError:
                continue

    if not bin_lwh:
        raise ValueError("未解析到容器尺寸")
    return {"max_bins": max_bins, "bin": bin_lwh, "items": items}


def scale_to_mm(
    raw: Dict[str, Any],
    *,
    unit_to_mm: float = 100.0,
    default_weight_kg: float = 10.0,
    weight_by_volume: bool = True,
) -> Dict[str, Any]:
    """
    原始单位放大为 mm。
    weight: 默认按体积比例估，或固定 default_weight_kg。
    """
    bl, bw, bh = raw["bin"]
    cont = {
        "type": "TEST_BIN",
        "inner_length_mm": int(round(bl * unit_to_mm)),
        "inner_width_mm": int(round(bw * unit_to_mm)),
        "inner_height_mm": int(round(bh * unit_to_mm)),
        "max_payload_kg": 99999.0,
        "max_containers": int(raw.get("max_bins") or 1),
        "source_unit_scale_mm": unit_to_mm,
    }
    boxes = []
    for it in raw["items"]:
        L = int(round(it["length"] * unit_to_mm))
        W = int(round(it["width"] * unit_to_mm))
        H = int(round(it["height"] * unit_to_mm))
        vol = max(L * W * H, 1)
        if weight_by_volume:
            # ~200 kg/m3 假密度
            w = max(1.0, round(vol / 1e9 * 200, 2))
        else:
            w = default_weight_kg
        boxes.append(
            {
                "box_id": f"B{it['case_id']}",
                "length_mm": L,
                "width_mm": W,
                "height_mm": H,
                "weight_kg": w,
                "quantity": int(it["quantity"]),
                "allow_rotate": True,
            }
        )
    return {
        "schema": "packing_assistant.benchmark_case.v1",
        "name": "",
        "description": "",
        "source": "dwave-examples/3d-bin-packing",
        "stage": "container_load_only",  # 第二阶段：已装箱 → 拼柜
        "container": cont,
        "boxes": boxes,
    }


def expand_boxes_api(case: Dict[str, Any]) -> List[Dict[str, Any]]:
    """展开 quantity → pack_boxes_api 用的 boxes[]。"""
    out = []
    for b in case.get("boxes") or []:
        q = int(b.get("quantity") or 1)
        for i in range(q):
            bid = b.get("box_id") or "B"
            out.append(
                {
                    "box_id": f"{bid}-{i+1:03d}" if q > 1 else bid,
                    "outer_size_mm": {
                        "length": b["length_mm"],
                        "width": b["width_mm"],
                        "height": b["height_mm"],
                    },
                    "gross_weight_kg": float(b.get("weight_kg") or 0),
                    "allowRotate": bool(b.get("allow_rotate", True)),
                    "special_attributes": list(b.get("special_attributes") or []),
                }
            )
    return out


def to_materials(case: Dict[str, Any]) -> List[Dict[str, Any]]:
    """转材料表（第一阶段）；通用箱不当真做结构。"""
    mats = []
    for i, b in enumerate(case.get("boxes") or [], 1):
        mats.append(
            {
                "id": f"M{i:03d}",
                "name": f"benchmark-{b.get('box_id') or i}",
                "quantity": int(b.get("quantity") or 1),
                "weight_kg": float(b.get("weight_kg") or 0),
                "total_weight_kg": float(b.get("weight_kg") or 0) * int(b.get("quantity") or 1),
                "length_mm": b["length_mm"],
                "width_mm": b["width_mm"],
                "height_mm": b["height_mm"],
                "note": "benchmark_not_steel",
            }
        )
    return mats


def write_json(path: Path, case: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(case, ensure_ascii=False, indent=2), encoding="utf-8")


def write_excel(path: Path, case: Dict[str, Any]) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    # boxes
    ws = wb.active
    ws.title = "boxes"
    bh = [
        "box_id",
        "length_mm",
        "width_mm",
        "height_mm",
        "weight_kg",
        "quantity",
        "allow_rotate",
    ]
    ws.append(bh)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
    for b in case.get("boxes") or []:
        ws.append([b.get(h) for h in bh])

    # materials
    ws2 = wb.create_sheet("materials")
    mh = [
        "id",
        "name",
        "quantity",
        "weight_kg",
        "total_weight_kg",
        "length_mm",
        "width_mm",
        "height_mm",
        "note",
    ]
    ws2.append(mh)
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
    for m in to_materials(case):
        ws2.append([m.get(h) for h in mh])

    # container
    ws3 = wb.create_sheet("container")
    ws3.append(["key", "value"])
    c = case.get("container") or {}
    for k, v in c.items():
        ws3.append([k, v])
    ws3.append(["case_name", case.get("name") or ""])
    ws3.append(["stage", case.get("stage") or ""])
    ws3.append(["source", case.get("source") or ""])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def builtin_cases() -> List[Dict[str, Any]]:
    """预置 Case A/B/C + 40HQ 风格 + 超重。"""
    cases = []

    # A 小件电商风 20GP
    cases.append(
        {
            "schema": "packing_assistant.benchmark_case.v1",
            "name": "case_a_small_cartons_20gp",
            "description": "小件纸箱级，易装满 20GP",
            "source": "synthetic/ecommerce-style",
            "stage": "container_load_only",
            "container": {
                "type": "20GP",
                "inner_length_mm": 5898,
                "inner_width_mm": 2352,
                "inner_height_mm": 2385,
                "max_payload_kg": 21000,
                "max_containers": 1,
            },
            "boxes": [
                {
                    "box_id": "CTN400",
                    "length_mm": 400,
                    "width_mm": 300,
                    "height_mm": 300,
                    "weight_kg": 8,
                    "quantity": 200,
                    "allow_rotate": True,
                },
                {
                    "box_id": "CTN500",
                    "length_mm": 500,
                    "width_mm": 400,
                    "height_mm": 350,
                    "weight_kg": 12,
                    "quantity": 120,
                    "allow_rotate": True,
                },
                {
                    "box_id": "CTN600",
                    "length_mm": 600,
                    "width_mm": 400,
                    "height_mm": 400,
                    "weight_kg": 15,
                    "quantity": 80,
                    "allow_rotate": True,
                },
            ],
        }
    )

    # B 长件/铁架风格 40HQ
    cases.append(
        {
            "schema": "packing_assistant.benchmark_case.v1",
            "name": "case_b_long_frames_40hq",
            "description": "2m/4m/6m 铁架风格长件，偏业务",
            "source": "synthetic/steel-frame-style",
            "stage": "container_load_only",
            "container": {
                "type": "40HQ",
                "inner_length_mm": 12032,
                "inner_width_mm": 2352,
                "inner_height_mm": 2698,
                "max_payload_kg": 26480,
                "max_containers": 2,
            },
            "boxes": [
                {
                    "box_id": "FR2M",
                    "length_mm": 2100,
                    "width_mm": 1100,
                    "height_mm": 1200,
                    "weight_kg": 1800,
                    "quantity": 2,
                    "allow_rotate": False,
                    "special_attributes": ["需加固"],
                },
                {
                    "box_id": "FR4M",
                    "length_mm": 4100,
                    "width_mm": 1100,
                    "height_mm": 1300,
                    "weight_kg": 3200,
                    "quantity": 3,
                    "allow_rotate": False,
                    "special_attributes": ["需加固"],
                },
                {
                    "box_id": "FR6M",
                    "length_mm": 6000,
                    "width_mm": 1100,
                    "height_mm": 1200,
                    "weight_kg": 4500,
                    "quantity": 2,
                    "allow_rotate": False,
                    "special_attributes": ["超长", "需加固"],
                },
                {
                    "box_id": "CAGE",
                    "length_mm": 2200,
                    "width_mm": 1100,
                    "height_mm": 1100,
                    "weight_kg": 800,
                    "quantity": 2,
                    "allow_rotate": True,
                },
            ],
        }
    )

    # C 压力/风险：限重 8t
    cases.append(
        {
            "schema": "packing_assistant.benchmark_case.v1",
            "name": "case_c_payload_stress_40hq",
            "description": "40HQ 但限重压到 8 吨 + 大重箱，测超重/装不下/风险",
            "source": "synthetic/risk",
            "stage": "container_load_only",
            "container": {
                "type": "40HQ_LIMITED",
                "inner_length_mm": 12032,
                "inner_width_mm": 2352,
                "inner_height_mm": 2698,
                "max_payload_kg": 8000,
                "max_containers": 2,
            },
            "boxes": [
                {
                    "box_id": "HEAVY1",
                    "length_mm": 3000,
                    "width_mm": 1500,
                    "height_mm": 1500,
                    "weight_kg": 3500,
                    "quantity": 2,
                    "allow_rotate": False,
                },
                {
                    "box_id": "HEAVY2",
                    "length_mm": 4000,
                    "width_mm": 1200,
                    "height_mm": 1200,
                    "weight_kg": 2800,
                    "quantity": 2,
                    "allow_rotate": False,
                },
                {
                    "box_id": "MID",
                    "length_mm": 2000,
                    "width_mm": 1000,
                    "height_mm": 1000,
                    "weight_kg": 1200,
                    "quantity": 3,
                    "allow_rotate": True,
                },
            ],
        }
    )

    # 40HQ style mixed
    cases.append(
        {
            "schema": "packing_assistant.benchmark_case.v1",
            "name": "test_case_40hq_style",
            "description": "接近铁架/40HQ 混装",
            "source": "synthetic/40hq-style",
            "stage": "container_load_only",
            "container": {
                "type": "40HQ",
                "inner_length_mm": 12032,
                "inner_width_mm": 2352,
                "inner_height_mm": 2698,
                "max_payload_kg": 26480,
                "max_containers": 1,
            },
            "boxes": [
                {
                    "box_id": "B6",
                    "length_mm": 5800,
                    "width_mm": 1100,
                    "height_mm": 1100,
                    "weight_kg": 4000,
                    "quantity": 2,
                    "allow_rotate": False,
                    "special_attributes": ["超长"],
                },
                {
                    "box_id": "B4",
                    "length_mm": 4000,
                    "width_mm": 1100,
                    "height_mm": 1400,
                    "weight_kg": 2500,
                    "quantity": 2,
                    "allow_rotate": False,
                },
                {
                    "box_id": "B11",
                    "length_mm": 1200,
                    "width_mm": 1100,
                    "height_mm": 1100,
                    "weight_kg": 900,
                    "quantity": 3,
                    "allow_rotate": True,
                },
            ],
        }
    )

    cases.append(
        {
            "schema": "packing_assistant.benchmark_case.v1",
            "name": "test_case_overweight",
            "description": "故意超重，测风险合规",
            "source": "synthetic/overweight",
            "stage": "container_load_only",
            "container": {
                "type": "20GP",
                "inner_length_mm": 5898,
                "inner_width_mm": 2352,
                "inner_height_mm": 2385,
                "max_payload_kg": 5000,
                "max_containers": 1,
            },
            "boxes": [
                {
                    "box_id": "OW1",
                    "length_mm": 2000,
                    "width_mm": 1000,
                    "height_mm": 1000,
                    "weight_kg": 3000,
                    "quantity": 3,
                    "allow_rotate": False,
                }
            ],
        }
    )
    return cases


def register_custom_container(case: Dict[str, Any]) -> None:
    """把自定义柜型注册进 CONTAINER_INNER（进程内）。"""
    from packing_assistant.tools import bin3d

    c = case.get("container") or {}
    ctype = c.get("type") or "TEST_BIN"
    if ctype not in bin3d.CONTAINER_INNER:
        bin3d.CONTAINER_INNER[ctype] = {
            "L": float(c.get("inner_length_mm") or 12000),
            "W": float(c.get("inner_width_mm") or 2350),
            "H": float(c.get("inner_height_mm") or 2690),
            "max_load_kg": float(c.get("max_payload_kg") or 26000),
        }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", action="append", default=[], help="D-Wave txt 路径，可多次")
    ap.add_argument("--out", default=str(ROOT / "test" / "benchmarks"))
    ap.add_argument("--unit-mm", type=float, default=100.0, help="原始单位→mm 倍率，默认×100")
    args = ap.parse_args()

    out = Path(args.out)
    if not out.is_absolute():
        out = ROOT / out
    out.mkdir(parents=True, exist_ok=True)
    excel_dir = out / "excel"
    excel_dir.mkdir(exist_ok=True)

    generated: List[str] = []

    # D-Wave files
    inputs = [Path(p) for p in args.input]
    if not inputs:
        for name in ("sample_data_1.txt", "sample_data_2.txt"):
            p = ROOT / "data" / "external" / name
            if p.exists():
                inputs.append(p)

    for ip in inputs:
        if not ip.exists():
            print("SKIP missing", ip)
            continue
        raw = parse_dwave_txt(ip.read_text(encoding="utf-8"))
        case = scale_to_mm(raw, unit_to_mm=args.unit_mm)
        case["name"] = ip.stem
        case["description"] = f"From {ip.name}, scale={args.unit_mm}mm per unit"
        case["source"] = f"dwave-examples:{ip.name}"
        jpath = out / f"{ip.stem}.json"
        write_json(jpath, case)
        write_excel(excel_dir / f"{ip.stem}.xlsx", case)
        generated.append(str(jpath))
        print(f"OK {jpath.name} boxes_sku={len(case['boxes'])} max_bins={case['container']['max_containers']}")

    # builtins
    for case in builtin_cases():
        jpath = out / f"{case['name']}.json"
        write_json(jpath, case)
        write_excel(excel_dir / f"{case['name']}.xlsx", case)
        generated.append(str(jpath))
        print(f"OK {jpath.name}")

    # index
    index = {
        "dir": str(out),
        "cases": [Path(p).name for p in generated],
        "usage": {
            "stage2_load": "scripts/run_benchmark_cases.py",
            "convert": "scripts/convert_bpp_to_cases.py",
            "business": "test/excel 远东真实数据",
        },
        "notes": [
            "通用 BPP 数据只测拼柜引擎，不做结构计算当真",
            "业务正确性仍以远东 Excel 为主",
        ],
    }
    (out / "INDEX.json").write_text(
        json.dumps(index, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (out / "README.md").write_text(
        """# 公开 3D-BPP / 拼柜基准用例

| 来源 | 文件 | 用途 |
|------|------|------|
| D-Wave sample_data_1/2 | `sample_data_*.json` | 单柜/多柜冒烟 |
| Case A | `case_a_small_cartons_20gp.json` | 小件易装 |
| Case B | `case_b_long_frames_40hq.json` | 长件/铁架风格 |
| Case C | `case_c_payload_stress_40hq.json` | 限重压力 |
| 40HQ style / overweight | 对应 json | 混装 / 超重风险 |

```bash
python scripts/convert_bpp_to_cases.py
python scripts/run_benchmark_cases.py
```

Excel 在 `excel/` 子目录。
""",
        encoding="utf-8",
    )
    print("=" * 50)
    print(f"DONE n={len(generated)} → {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
