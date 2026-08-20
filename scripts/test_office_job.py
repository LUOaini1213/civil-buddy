#!/usr/bin/env python3
"""Job-root Office interchange: tables → xlsx; never D:\\layout."""

from __future__ import annotations

import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))


def main() -> int:
    from packing_assistant.office_job import (
        export_md_to_xlsx,
        is_forbidden_layout,
        job_root,
        tables_from_md,
    )

    assert is_forbidden_layout(Path(r"D:\layout"))
    assert is_forbidden_layout(Path(r"D:\layout\foo"))
    assert not is_forbidden_layout(Path(r"C:\Users\LW\civil-buddy"))

    os.environ["CIVIL_JOB_ROOT"] = r"D:\layout"
    assert not is_forbidden_layout(job_root())
    assert str(job_root()).replace("\\", "/").lower().endswith(".civil-buddy/out")

    job = ROOT / "output" / "office-job-test"
    job.mkdir(parents=True, exist_ok=True)
    os.environ["CIVIL_JOB_ROOT"] = str(job)
    os.environ["CIVIL_SANDBOX_ROOTS"] = str(job)
    assert job_root().resolve() == job.resolve()

    md = (
        "# 核算\n\n## 6 报销勾选\n\n"
        "| 检查项 | 本稿 |\n| --- | --- |\n| 发票查验 | 待核 |\n| 审批链 | 待核 |\n\n"
        "## 8 对账缺口\n\n"
        "| 台账 | 本稿 |\n| --- | --- |\n| 物资收发存 | 缺口待列 |\n"
    )
    sheets = tables_from_md(md)
    assert len(sheets) == 2, sheets
    assert sheets[0][0] == "6 报销勾选"
    assert sheets[0][1][0] == ["检查项", "本稿"]
    assert ["发票查验", "待核"] in sheets[0][1]

    src = job / "finance-book__check.md"
    src.write_text(md, encoding="utf-8")
    paths = export_md_to_xlsx(src)
    assert paths, paths
    xlsx = [p for p in paths if p.suffix == ".xlsx"]
    assert xlsx
    import openpyxl

    wb = openpyxl.load_workbook(xlsx[0])
    assert "6 报销勾选" in wb.sheetnames
    assert wb["6 报销勾选"]["A2"].value == "发票查验"
    assert wb["6 报销勾选"]["B2"].value == "待核"
    ledger = job / "现场台账.xlsx"
    import openpyxl

    wb2 = openpyxl.Workbook()
    ws = wb2.active
    ws.title = "收发"
    ws["A1"] = "物资"
    ws["B1"] = "入库"
    ws["A2"] = "钢筋"
    ws["B2"] = "12吨"
    wb2.save(ledger)
    from packing_assistant.office_job import job_files_blob, list_job_files

    listed = list_job_files()
    assert any(f["name"] == "现场台账.xlsx" for f in listed), listed
    blob = job_files_blob("写一份收发存 现场台账")
    assert "作业根文件" in blob
    assert "钢筋" in blob
    assert "12吨" in blob
    from packing_assistant.expert_turn import run_expert_turn

    wh = run_expert_turn(
        "写一份收发存 现场台账",
        "warehouse",
        force_intent="run",
        session_id="t065-wh-job",
    )
    assert wh["wrote"] is True
    wt = Path(wh["files"][0]["path"]).read_text(encoding="utf-8")
    assert "钢筋" in wt
    assert "12吨" in wt
    patched = openpyxl.load_workbook(ledger)
    assert "收发" in patched.sheetnames
    assert patched["收发"]["A2"].value == "钢筋"
    assert patched["收发"]["B2"].value == "12吨"
    drafts = [n for n in patched.sheetnames if n.startswith("CB草稿")]
    assert drafts, patched.sheetnames
    found = False
    for n in drafts:
        for row in patched[n].iter_rows(max_row=20, max_col=8, values_only=True):
            if row and any(cell and "钢筋" in str(cell) for cell in row):
                found = True
    assert found, [list(patched[n].iter_rows(max_row=8, values_only=True)) for n in drafts]
    print("PASS office_job")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
