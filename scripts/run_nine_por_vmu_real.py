#!/usr/bin/env python3
"""
用 POR/VMU 真实提料尺寸 + FAC0008 样例估算，跑 9 智能体。
- 排除 已发货
- 不修改 A: 项目文件
"""
from __future__ import annotations

import importlib.util
import json
import re
import sys
from pathlib import Path
from typing import Any, Dict, List

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook, load_workbook
from pypdf import PdfReader

from packing_assistant.harness import apply_user_confirmation, make_initial_state
from packing_assistant.agents import (
    agent_box_scheme,
    agent_evaluator,
    agent_finalize,
    agent_loader,
    agent_material_parser,
    agent_orchestrator,
    agent_planner,
    agent_present_team_a,
    agent_risk_compliance,
    agent_structure,
    agent_visualizer,
)

VMU1 = Path(r"A:\JOB\REDACTED-JOB\Project\6. Quality QAQC\6.06 POR\VMU\VMU1")
OUT = ROOT / "output" / "por_vmu_nine"
OUT.mkdir(parents=True, exist_ok=True)

# 铝型材截面估算（提料仅有 L）：幕墙型材实务量级
PROFILE_SECTION = {
    "default": (80.0, 60.0, 2.8),  # W, H, kg/m approx for density
    "louver": (100.0, 40.0, 2.2),
    "gasket": (40.0, 20.0, 0.8),
}


def _al_weight(L_mm: float, kg_per_m: float) -> float:
    return round(kg_per_m * (L_mm / 1000.0), 3)


def _panel_weight(W: float, H: float, T: float, density: float = 2.7e-6) -> float:
    # density kg/mm3 for aluminum ~ 2.7e-6; honeycomb lighter factor later
    return round(W * H * T * density, 3)


def load_bgl0003() -> List[Dict[str, Any]]:
    x = Path(
        r"A:\JOB\REDACTED-JOB\InterDepartment\3.05 Action\User Working files"
        r"\Simon Ng\VMU\MATERIAL ORDER\Glass"
        r"\REDACTED-CODE-VMU-0001-BGL0003 (REDACTED-CODE 玻璃提料单-送工厂）.xlsx"
    )
    mats = []
    if not x.exists():
        return mats
    wb = load_workbook(x, data_only=True)
    ws = wb.active
    mid = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i <= 4 or row[0] is None:
            continue
        try:
            W, H, qty = float(row[7]), float(row[8]), int(float(row[13]))
        except Exception:
            continue
        T = 40.0  # 中空夹胶实务估计 mm
        mid += 1
        # glass ~ 2.5*area*thickness_factor; ~25 kg/m2 for IGU rough
        area = W * H / 1e6
        wt = round(area * 28.0, 2)  # kg/pc estimate
        mats.append(
            {
                "id": f"BGL3-{mid:03d}",
                "name": f"玻璃 {row[1] or 'IGU'} {row[4] or ''}".strip()[:60],
                "quantity": max(qty, 1),
                "weight_kg": wt,
                "total_weight_kg": round(wt * max(qty, 1), 2),
                "length_mm": max(W, H),
                "width_mm": min(W, H),
                "height_mm": T,
                "spec": "BGL0003",
                "part_no": str(row[4] or ""),
                "note": "real POR dims; T=40mm estimated IGU",
            }
        )
    wb.close()
    return mats


def load_fac0007() -> List[Dict[str, Any]]:
    x = Path(
        r"A:\JOB\REDACTED-JOB\InterDepartment\3.05 Action\User Working files"
        r"\Simon Ng\VMU\MATERIAL ORDER\Honeycomb"
        r"\007-REDACTED-CODE-REDACTED-CODE-FAC0007(REDACTED-CODEF HONEYCOMB MATERIAL TAKE OFF-SEND TO FACTORY.xlsx"
    )
    mats = []
    if not x.exists():
        return mats
    wb = load_workbook(x, data_only=True)
    ws = wb.active
    mid = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i <= 4 or row[0] is None:
            continue
        try:
            W, H, T, qty = float(row[6]), float(row[7]), float(row[8]), int(float(row[13]))
        except Exception:
            continue
        mid += 1
        # honeycomb ~ 5-8 kg/m2 for 25mm
        wt = round((W * H / 1e6) * 6.5, 2)
        mats.append(
            {
                "id": f"FAC7-{mid:03d}",
                "name": f"蜂窝铝板 {row[3] or mid}",
                "quantity": max(qty, 1),
                "weight_kg": wt,
                "total_weight_kg": round(wt * max(qty, 1), 2),
                "length_mm": max(W, H),
                "width_mm": min(W, H),
                "height_mm": T,
                "spec": "FAC0007",
                "part_no": str(row[3] or ""),
                "note": "real POR/excel W×H×T=25",
            }
        )
    wb.close()
    return mats


def load_fac0012() -> List[Dict[str, Any]]:
    d = next(VMU1.glob("*FAC0012*"), None)
    if not d:
        return []
    pdf = next(d.glob("*.pdf"), None)
    if not pdf:
        return []
    r = PdfReader(str(pdf))
    text = "\n".join((p.extract_text() or "") for p in r.pages)
    mats = []
    mid = 0
    for line in text.splitlines():
        m = re.search(
            r"^(\d+)\s+(.+?)\s+(\d{6,})\s+\S+\s+\S+\s+\S*\s*"
            r"(\d+(?:\.\d+)?)\s+(\d+(?:\.\d+)?)\s+(\d+(?:\.\d+)?)\s+.*?"
            r"(\d+(?:\.\d+)?)\s*PC",
            line.strip(),
        )
        if not m:
            continue
        W, H, T, qty = float(m.group(4)), float(m.group(5)), float(m.group(6)), int(float(m.group(7)))
        mid += 1
        wt = _panel_weight(W, H, T)
        mats.append(
            {
                "id": f"FAC12-{mid:03d}",
                "name": f"3mm铝板 FAC0012-{mid}",
                "quantity": max(qty, 1),
                "weight_kg": wt,
                "total_weight_kg": round(wt * max(qty, 1), 2),
                "length_mm": max(W, H),
                "width_mm": min(W, H),
                "height_mm": T,
                "spec": "FAC0012",
                "part_no": "",
                "note": "real POR PDF",
            }
        )
    return mats


def load_bal_pdf(glob_pat: str, code: str) -> List[Dict[str, Any]]:
    fs = list(VMU1.glob(glob_pat))
    if not fs:
        return []
    r = PdfReader(str(fs[0]))
    text = "\n".join((p.extract_text() or "") for p in r.pages)
    mats = []
    mid = 0
    for line in text.splitlines():
        m = re.search(
            r"^(\d+)\s+(\S+)\s+(\d{6,})\s+(\d+(?:\.\d+)?)\s+\S+\s+(\d+(?:\.\d+)?)\s*PC",
            line.strip(),
        )
        if not m:
            m = re.search(
                r"^(\d+)\s+.*?(\d{6,})\s+(\d+(?:\.\d+)?)\s+\S+\s+(\d+(?:\.\d+)?)\s*PC",
                line.strip(),
            )
            if not m:
                continue
            name, L, qty = "profile", float(m.group(3)), int(float(m.group(4)))
        else:
            name, L, qty = m.group(2), float(m.group(4)), int(float(m.group(5)))
        mid += 1
        sec = PROFILE_SECTION["louver"] if "百叶" in name or "louver" in name.lower() else PROFILE_SECTION["default"]
        if "垫片" in name:
            sec = PROFILE_SECTION["gasket"]
        W, H, kg_m = sec
        wt = _al_weight(L, kg_m)
        mats.append(
            {
                "id": f"{code}-{mid:03d}",
                "name": f"铝料 {name}"[:60],
                "quantity": max(qty, 1),
                "weight_kg": wt,
                "total_weight_kg": round(wt * max(qty, 1), 2),
                "length_mm": L,
                "width_mm": W,
                "height_mm": H,
                "spec": code,
                "part_no": name,
                "note": "real L from POR; W×H×kg/m estimated section",
            }
        )
    return mats


def estimate_fac0008(total_pcs: int = 906) -> List[Dict[str, Any]]:
    """
    FAC0008 主票：PDF 多为扫描，用文字页样例尺寸分布估算 906 件。
    样例（3mm 铝板）：W≈1411–2285, H≈818–1496, T=3。
    按若干代表规格分摊数量。
    """
    # representative bins from real samples on text pages
    bins = [
        # (W, H, T, share)
        (2235.0, 820.0, 3.0, 0.18),
        (2235.0, 1300.0, 3.0, 0.20),
        (2235.0, 1450.0, 3.0, 0.12),
        (2033.0, 820.0, 3.0, 0.12),
        (2033.0, 1310.0, 3.0, 0.12),
        (1411.0, 840.0, 3.0, 0.10),
        (1411.0, 1340.0, 3.0, 0.08),
        (2285.0, 1320.0, 3.0, 0.08),
    ]
    # normalize shares
    s = sum(b[3] for b in bins)
    bins = [(W, H, T, sh / s) for W, H, T, sh in bins]
    mats = []
    assigned = 0
    for i, (W, H, T, sh) in enumerate(bins, 1):
        qty = int(round(total_pcs * sh))
        if i == len(bins):
            qty = max(0, total_pcs - assigned)
        assigned += qty
        if qty <= 0:
            continue
        wt = _panel_weight(W, H, T)
        mats.append(
            {
                "id": f"FAC8-E{i:02d}",
                "name": f"3mm铝板 FAC0008估 {W:.0f}x{H:.0f}",
                "quantity": qty,
                "weight_kg": wt,
                "total_weight_kg": round(wt * qty, 2),
                "length_mm": max(W, H),
                "width_mm": min(W, H),
                "height_mm": T,
                "spec": "FAC0008_estimated",
                "part_no": f"EST-{i}",
                "note": f"estimated from FAC0008 text-page samples; total_target={total_pcs}",
            }
        )
    return mats


def build_materials() -> List[Dict[str, Any]]:
    mats: List[Dict[str, Any]] = []
    mats += estimate_fac0008(906)  # 待拼柜主票
    mats += load_fac0007()
    mats += load_fac0012()
    mats += load_bgl0003()
    mats += load_bal_pdf("*BAL0004*.pdf", "BAL0004")
    mats += load_bal_pdf("*BAL0005*.pdf", "BAL0005")
    mats += load_bal_pdf("*BAL0020*.pdf", "BAL0020")
    # renumber ids unique
    for i, m in enumerate(mats, 1):
        m["id"] = f"M{i:03d}"
    return mats


def save_xlsx(mats: List[Dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "materials"
    cols = [
        "id",
        "name",
        "spec",
        "quantity",
        "weight_kg",
        "total_weight_kg",
        "length_mm",
        "width_mm",
        "height_mm",
        "part_no",
        "note",
    ]
    ws.append(cols)
    for m in mats:
        ws.append([m.get(c) for c in cols])
    wb.save(path)


def _apply(state, upd):
    for k, v in (upd or {}).items():
        if k in ("messages", "traces", "errors", "validation_warnings") and isinstance(v, list):
            state[k] = list(state.get(k) or []) + v
        else:
            state[k] = v
    return state


def run_nine(
    mats,
    container="40HQ",
    max_containers=12,
    max_box_net=3200.0,
    revision_mode: bool = False,
):
    agents = [
        (1, "主控智能体", "orchestrator", agent_orchestrator),
        (2, "材料解析智能体", "material_parser", agent_material_parser),
        (3, "结构计算智能体", "structure", agent_structure),
        (4, "装箱方案智能体", "box_scheme", agent_box_scheme),
        (0, "确认闸门", "present_team_a", agent_present_team_a),
        (5, "规划智能体", "planner", agent_planner),
        (6, "装载执行智能体", "loader", agent_loader),
        (7, "评估优化智能体", "evaluator", agent_evaluator),
        (8, "风险合规智能体", "risk_compliance", agent_risk_compliance),
        (9, "可视化智能体", "visualizer", agent_visualizer),
        (0, "主控汇总", "finalize", agent_finalize),
    ]
    state = make_initial_state(
        user_input="POR/VMU 真实尺寸+FAC0008估算 9智能体",
        materials=mats,
        container_type=container,
        enable_auto_confirm=True,
        max_containers=max_containers,
        session_id="por-vmu-real-nine",
    )
    state["packing_options"] = {
        "max_box_net_kg": max_box_net,
        "revision_mode": revision_mode,
    }
    if revision_mode:
        state["revision"] = {
            "active": True,
            "max_box_net_kg": max_box_net,
            "reason": "structure_reject",
        }
    steps = []
    for num, name, node, fn in agents:
        state = _apply(state, fn(state) or {})
        if node == "orchestrator":
            # 装箱模块高度按用户柜型，勿被主控误推 40GP 拉垮结构
            state["container_type"] = container
        if node == "present_team_a":
            state = apply_user_confirmation(
                state, action="confirm", container_type=container, max_containers=max_containers
            )
        last = ""
        for m in reversed(state.get("messages") or []):
            if m.get("content"):
                last = str(m["content"])
                break
        print("=" * 64)
        print(f"### {num or ''} {name} [{node}]")
        print(last[:900])
        steps.append({"n": num, "name": name, "node": node, "message": last[:2000], "phase": state.get("phase")})
        if node == "box_scheme":
            ta = state.get("team_a_summary") or {}
            print("  team_a", ta)
        if node == "loader":
            cp = state.get("container_plan") or {}
            print(
                f"  can_fit={cp.get('can_fit')} used={cp.get('containers_used')} "
                f"space={cp.get('space_utilization')} floor={cp.get('floor_utilization_avg')} "
                f"wt={cp.get('weight_utilization')}"
            )
        if node == "risk_compliance":
            rr = state.get("risk_report") or {}
            print(
                f"  risk decision={rr.get('decision')} level={rr.get('level')} "
                f"blockers={len(rr.get('blockers') or [])} passed={rr.get('passed')}"
            )
        if node == "finalize":
            print("ship_ok", state.get("ship_ok"), "status", state.get("status"))
    return state, steps


def main() -> int:
    print("构建材料（真实+FAC0008估）...")
    mats = build_materials()
    net = sum(float(m["total_weight_kg"]) for m in mats)
    print(f"材料行 {len(mats)}  估算净重 {net:.1f} kg")
    by_spec: Dict[str, int] = {}
    for m in mats:
        by_spec[m["spec"]] = by_spec.get(m["spec"], 0) + int(m["quantity"])
    print("件数 by spec:", by_spec)
    xlsx = OUT / "materials_por_vmu_real_fac8est.xlsx"
    save_xlsx(mats, xlsx)
    print("材料表:", xlsx)

    # 体积粗估 → 柜数下界
    vol = sum(
        m["length_mm"] * m["width_mm"] * m["height_mm"] * m["quantity"] / 1e9 for m in mats
    )
    guess = min(max(int(vol / 25) + 1, int(net / 18000) + 1, 2), 20)
    print(f"原料理论体积~{vol:.1f}m3  起始 max_containers={guess}")

    mc = max(guess, 4)
    # 铝料+长件：实测 max_box_net≈800 结构可全过
    cap = 800.0
    all_rounds = []
    for rnd in range(8):
        print("\n" + "#" * 64)
        print(f"ROUND {rnd} max_containers={mc} max_box_net={cap}")
        state, steps = run_nine(
            mats,
            container="40HQ",
            max_containers=mc,
            max_box_net=cap,
            revision_mode=(rnd > 0),
        )
        snap = {
            "round": rnd,
            "max_containers": mc,
            "max_box_net_kg": cap,
            "packing_plan_id": state.get("packing_plan_id"),
            "boxes": len(state.get("boxes") or []),
            "struct_fail": sum(
                1 for b in (state.get("boxes") or []) if b.get("structure_conclusion") == "不通过"
            ),
            "can_fit": (state.get("container_plan") or {}).get("can_fit"),
            "containers_used": (state.get("container_plan") or {}).get("containers_used"),
            "space": (state.get("container_plan") or {}).get("space_utilization"),
            "floor": (state.get("container_plan") or {}).get("floor_utilization_avg"),
            "weight": (state.get("container_plan") or {}).get("weight_utilization"),
            "risk_decision": (state.get("risk_report") or {}).get("decision"),
            "risk_level": (state.get("risk_report") or {}).get("level"),
            "ship_ok": state.get("ship_ok"),
            "status": state.get("status"),
            "phase": state.get("phase"),
        }
        print("SNAPSHOT", json.dumps(snap, ensure_ascii=False))
        all_rounds.append(
            {
                "snapshot": snap,
                "steps": steps,
                "final": (state.get("final_response") or "")[:2500],
            }
        )
        if snap.get("ship_ok") or (
            snap.get("can_fit")
            and snap.get("risk_decision") in ("PASS", "WARN")
            and snap.get("struct_fail") == 0
        ):
            print("成功收敛")
            break
        if snap.get("struct_fail", 0) > 0:
            cap = max(600.0, min(cap * 0.65, cap - 200))
            print(f"结构打回 -> 降 max_box_net={cap:.0f}")
            continue
        if not snap.get("can_fit") and snap.get("struct_fail") == 0:
            mc = min(mc + 3, 30)
            print(f"加柜 -> {mc}")
            continue
        break

    rep = {
        "note": "BGL/FAC0007/FAC0012/BAL 真实尺寸；FAC0008 按文字页样例估算 906 件；排除已发货",
        "materials_file": str(xlsx),
        "materials_count": len(mats),
        "net_kg": net,
        "by_spec_qty": by_spec,
        "rounds": all_rounds,
        "final_snapshot": all_rounds[-1]["snapshot"] if all_rounds else {},
    }
    outj = OUT / "agent_sequence_9_por_vmu.json"
    outj.write_text(json.dumps(rep, ensure_ascii=False, indent=2), encoding="utf-8")
    print("WROTE", outj)
    print("FINAL", json.dumps(rep["final_snapshot"], ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
