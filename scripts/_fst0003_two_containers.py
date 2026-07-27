#!/usr/bin/env python3
"""Confirm FST0003 site packing is already 2 cabinets in shipped loading lists."""
from pathlib import Path
import openpyxl

p = Path(
    r"A:\JOB\2517SLTO\Project\6. Quality QAQC\6.06 POR\已发货"
    r"\SLT0-VMU-0001-FST0003-02-4-8PR2 远东 新加坡陆路交通局办公楼项目 铁件(1)(1)(2).xlsx"
)
wb = openpyxl.load_workbook(p, data_only=True)
print("sheets:", wb.sheetnames)
for sn in wb.sheetnames:
    if "装货" not in sn and "柜" not in sn:
        continue
    ws = wb[sn]
    rows = list(ws.iter_rows(values_only=True))
    print("\n====", sn, "rows", len(rows))
    # find header with 单件重量
    hdr_i = None
    for i, r in enumerate(rows[:30]):
        vals = [str(c or "") for c in r]
        if any("单件重量" in v or "重量" in v for v in vals) and any(
            "名称" in v or "長" in v or "长" in v for v in vals
        ):
            hdr_i = i
            print("header", r[:15])
            break
    if hdr_i is None:
        for r in rows[:8]:
            print(r[:12])
        continue
    # sum weights
    # columns: try find weight and qty
    hdr = [str(c or "").replace("\n", "") for c in rows[hdr_i]]
    w_idx = None
    q_idx = None
    for j, h in enumerate(hdr):
        if "单件重量" in h or h.strip() == "单件重量(kg)":
            w_idx = j
        if h.strip() in ("数量", "數量", "qty"):
            q_idx = j
    print("w_idx", w_idx, "q_idx", q_idx, "hdr", hdr[:14])
    total_w = 0.0
    n = 0
    for r in rows[hdr_i + 1 :]:
        if r is None or r[0] is None:
            # may still have data
            pass
        try:
            w = float(r[w_idx]) if w_idx is not None and r[w_idx] not in (None, "") else 0.0
        except Exception:
            w = 0.0
        try:
            q = float(r[q_idx]) if q_idx is not None and r[q_idx] not in (None, "") else 1.0
        except Exception:
            q = 1.0
        if w > 0:
            total_w += w * (q if q > 0 else 1)
            n += 1
    print(f"lines_with_weight={n} total_piece_weight_kg≈{total_w:.1f}")
wb.close()
