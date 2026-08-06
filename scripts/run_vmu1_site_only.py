#!/usr/bin/env python3
"""
VMU1 剩余未发 · 仅「送工地」料单 → 估算装箱柜数（标准箱库+混装）。

数据源优先：
  A:\\...\\POR\\VMU\\Material_Summary_VMU送工地.xlsx
排除：已到货=0 且 未到货=0 的空行；可选对照 已发货 目录。
"""
from __future__ import annotations

import json
import sys
import time
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

VMU_DIR = Path(r"A:\JOB\2517SLTO\Project\6. Quality QAQC\6.06 POR\VMU")
SHIPPED_DIR = Path(r"A:\JOB\2517SLTO\Project\6. Quality QAQC\6.06 POR\已发货")
SITE_XLSX = VMU_DIR / "Material_Summary_VMU送工地.xlsx"
OUT = ROOT / "output" / "vmu1_site_only"

# 物料组 → (L,W,H mm, unit_kg, label) 实务粗估，非正式提料尺寸
GROUP_PROFILE: Dict[str, Tuple[float, float, float, float, str]] = {
    "11—铝板": (1800, 1200, 40, 18.0, "铝板"),
    "22—铝材": (4500, 80, 60, 9.5, "铝型材"),
    "13—铁件": (2000, 200, 150, 22.0, "铁件"),
    "14—不锈钢": (800, 200, 80, 4.0, "不锈钢件"),
    "19—胶条、垫块、胶皮": (600, 400, 300, 12.0, "胶条箱当量"),
    "23—紧固件/螺丝": (400, 300, 250, 8.0, "五金箱当量"),
    "24—Glass 玻璃": (1600, 1200, 80, 48.0, "玻璃"),
    "25—门窗锁配件": (500, 350, 250, 6.0, "五金箱当量"),
    "28—杂项配件": (1200, 800, 100, 15.0, "杂项"),
    "18—结构胶/耐候胶": (400, 300, 300, 20.0, "胶桶箱"),
    "17—拉爆螺栓，化学螺栓": (400, 300, 250, 10.0, "螺栓箱"),
}

BULK_PER_CARTON: Dict[str, int] = {
    "19—胶条、垫块、胶皮": 50,
    "23—紧固件/螺丝": 200,
    "25—门窗锁配件": 40,
    "17—拉爆螺栓，化学螺栓": 100,
}


def _f(x: Any) -> float:
    try:
        if x is None or x == "":
            return 0.0
        return float(x)
    except Exception:
        return 0.0


def _s(x: Any) -> str:
    return str(x or "").strip()


def load_site_rows(path: Path) -> List[Dict[str, Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [_s(h) for h in rows[0]]
    out = []
    for raw in rows[1:]:
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        d = {headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers))}
        out.append(d)
    wb.close()
    return out


def remaining_qty(row: Dict[str, Any]) -> float:
    """未发剩余量：优先未到货；否则已到货（在库待发工地）。双零视为无量。"""
    arr = _f(row.get("已到货数量"))
    pend = _f(row.get("未到货数量"))
    if pend > 0:
        return pend
    if arr > 0:
        return arr
    return 0.0


def is_vmu1(row: Dict[str, Any]) -> bool:
    batch = _s(row.get("施工批次")) + " " + _s(row.get("项目描述")) + " " + _s(
        row.get("訂貨單/加工圖號")
    )
    return "VMU-0001" in batch or "VMU01" in batch.upper() or "VMU1" in batch.upper()


def shipped_por_keys() -> set:
    keys = set()
    if not SHIPPED_DIR.exists():
        return keys
    for p in SHIPPED_DIR.rglob("*"):
        name = p.name.upper()
        if "VMU" not in name and "SLTO" not in name:
            continue
        # 提取 POR 号片段 FAC0011 / FST0003 等
        import re

        for m in re.findall(
            r"(FAC|FST|FSS|FHA|FHU|BBF|BGK|BGL|BSS|BOM|BAL|FST)\d{4}",
            name,
            flags=re.I,
        ):
            keys.add(m.upper())
        keys.add(p.stem[:40])
    return keys


def row_looks_shipped(row: Dict[str, Any], shipped: set) -> bool:
    por = _s(row.get("訂貨單/加工圖號")).upper()
    desc = _s(row.get("项目描述")).upper()
    for k in shipped:
        if len(k) >= 6 and (k in por or k in desc):
            return True
    return False


def to_materials(rows: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    将工地剩余量转为「可拼柜材料行」。
    大票铁件/铝板等按「装木箱/铁架当量」折算，避免 1998 件铁件被当成 1998 个大件虚增柜数。
    """
    mats: List[Dict[str, Any]] = []
    skipped: List[Dict[str, Any]] = []
    mid = 0

    for d in rows:
        por = _s(d.get("訂貨單/加工圖號"))
        desc = _s(d.get("项目描述"))
        group = _s(d.get("物料组描述"))
        n = remaining_qty(d)
        raw_n = n

        if not is_vmu1(d):
            skipped.append({"por": por, "reason": "非VMU1", "qty": n})
            continue
        if n <= 0:
            skipped.append({"por": por, "reason": "零量(已到=未到=0)", "qty": 0})
            continue
        if "工厂" in desc and "工地" not in desc:
            skipped.append({"por": por, "reason": "描述像送工厂", "qty": n})
            continue

        pu = por.upper()
        # —— 按 POR/物料组：件数 → 标准箱当量 + 当量外廓/单重 ——
        # 目标：与现场「若干件合一铁架/木箱」一致，而不是 1 件 1 箱
        if "BBF" in pu or "紧固件" in group or "螺丝" in group or "螺栓" in group:
            # 贴装货单：五金多箱合并，减少碎箱占底面
            per = 400
            units = max(1, int(round(n / per)))
            L, W, H, wt = 800.0, 600.0, 500.0, 55.0  # 稍大木箱当量
            unit, suffix = "五金箱当量", f"×{per}件/箱"
        elif "BGK" in pu or "胶条" in group or "垫块" in group:
            per = 60
            units = max(1, int(round(n / per)))
            L, W, H, wt = 800.0, 600.0, 500.0, 35.0
            unit, suffix = "胶条垫块箱", f"×{per}/箱"
        elif "FAC0011" in pu or ("铝板" in desc and "FAC" in pu):
            # 铝板叠层：略加密，并可上二层（H=800）
            per = 40
            units = max(1, int(round(n / per)))
            L, W, H, wt = 2200.0, 1200.0, 800.0, 40 * 18.0
            unit, suffix = "铝板架", f"×{per}片/架"
        elif "FST0022" in pu or "垫片" in desc:
            per = 90
            units = max(1, int(round(n / per)))
            L, W, H, wt = 1200.0, 800.0, 600.0, 90 * 2.5
            unit, suffix = "钢垫片箱", f"×{per}/箱"
        elif "FST0017" in pu or "吊具" in desc:
            units = max(1, int(round(n)))
            L, W, H, wt = 2000.0, 800.0, 600.0, 45.0
            unit, suffix = "吊具", ""
        elif "FST" in pu or "铁件" in group:
            # 对照 FST0003 装货单：1.1m 为主、2m/4m 点缀；合箱更密贴近现场
            per = 100
            units = max(1, int(round(n / per)))
            # 装货单 1 柜大量 1.1m 架；2/4m 少量
            frame_cycle = [
                (1100.0, 1100.0, 1750.0, 80.0),
                (1100.0, 1100.0, 1750.0, 80.0),
                (1100.0, 1100.0, 1750.0, 80.0),
                (1100.0, 1100.0, 1750.0, 80.0),
                (1100.0, 1100.0, 1750.0, 80.0),
                (2000.0, 1100.0, 1750.0, 140.0),
                (4000.0, 1100.0, 1750.0, 250.0),
            ]
            unit, suffix = "铁件架", f"×{per}件/架 raw={int(n)}"
            L, W, H, frame_tare = frame_cycle[0]
            wt = per * 15.0 + frame_tare
        elif "FSS" in pu or "不锈钢" in group:
            per = 40
            units = max(1, int(round(n / per))) if n >= 10 else max(1, int(round(n)))
            L, W, H, wt = 1500.0, 800.0, 600.0, max(n, 1) / max(units, 1) * 4.0
            unit, suffix = "不锈钢箱", f"×{per}/箱" if n >= 10 else ""
        elif "BOM0019" in pu or "瓦楞" in desc:
            # 密装：半柜宽+可叠高，减少架数，便于龙申 1 柜与铁件拼满
            # 旧：80件/架 × 2400×1200×500 → 9 架常挤出第 2 柜
            # 新：140件/架 × 2200×1100×1100 → ~5 架，双排+可上二层
            per = 140
            units = max(1, int(round(n / per)))
            L, W, H, wt = 2200.0, 1100.0, 1100.0, 140 * 8.0
            unit, suffix = "瓦楞板架密装", f"×{per}/架"
        elif "BOM0016" in pu or "木板" in desc:
            per = 30
            units = max(1, int(round(n / per)))
            L, W, H, wt = 2400.0, 1200.0, 400.0, 30 * 18.0
            unit, suffix = "木板架", f"×{per}/架"
        elif "BSS" in pu or "胶" in group:
            units = max(1, int(round(n / 12))) if n > 12 else max(1, int(round(n)))
            L, W, H, wt = 500.0, 400.0, 400.0, 24.0
            unit, suffix = "胶类箱", ""
        else:
            per = BULK_PER_CARTON.get(group, 1)
            if per > 1:
                units = max(1, int(round(n / per)))
                suffix = f"×{per}/箱当量"
            else:
                units = max(1, int(round(n)))
                suffix = ""
            prof = GROUP_PROFILE.get(group)
            if not prof:
                L, W, H, wt, unit = 1200.0, 800.0, 600.0, 20.0, "通用箱当量"
            else:
                L, W, H, wt, unit = prof

        # 每个当量箱单独一行 quantity=1，避免再被合箱算法并成超重/超跨
        n_units = max(1, int(units))
        is_steel_frame = unit == "铁件架"
        frame_cycle = [
            (1100.0, 1100.0, 1750.0, 80.0),
            (1100.0, 1100.0, 1750.0, 80.0),
            (1100.0, 1100.0, 1750.0, 80.0),
            (1100.0, 1100.0, 1750.0, 80.0),
            (1100.0, 1100.0, 1750.0, 80.0),
            (2000.0, 1100.0, 1750.0, 140.0),
            (4000.0, 1100.0, 1750.0, 250.0),
        ]
        for i in range(n_units):
            mid += 1
            if is_steel_frame:
                Lf, Wf, Hf, tare = frame_cycle[i % len(frame_cycle)]
                unit_wt = float(per) * 15.0 + float(tare)
                Li, Wi, Hi = Lf, Wf, Hf
                frame_tag = {1100.0: "1.1米", 2000.0: "2米", 4000.0: "4米"}.get(Lf, "")
                name = f"{frame_tag}铁件架{suffix} | {por or desc[:40]}#{i+1}"
            else:
                unit_wt = float(wt)
                Li, Wi, Hi = float(L), float(W), float(H)
                name = f"{unit}{suffix} | {por or desc[:40]}#{i+1}"
            mats.append(
                {
                    "id": f"S{mid:03d}",
                    "name": name,
                    "quantity": 1,
                    "weight_kg": round(unit_wt, 2),
                    "total_weight_kg": round(unit_wt, 2),
                    "length_mm": float(Li),
                    "width_mm": float(Wi),
                    "height_mm": float(Hi),
                    "spec": group or "未分组",
                    "part_no": por,
                    "note": (
                        f"dest=工地; raw_qty={raw_n}; crate={i+1}/{n_units}; "
                        f"arr={_f(d.get('已到货数量'))}; pend={_f(d.get('未到货数量'))}; "
                        f"dims=crate_equiv_est"
                    ),
                    "destination": "工地",
                    "source_sheet": "Material_Summary_VMU送工地",
                    "raw_qty": raw_n,
                }
            )
    return mats, skipped


def _crate_fill_hint(m: Dict[str, Any]) -> float:
    """当量箱填充率：铁架空心低、五金箱高；用于订柜 pack_effective。"""
    name = str(m.get("name") or "")
    spec = str(m.get("spec") or "")
    if "铁件架" in name or "铁件" in spec or "米铁" in name:
        return 0.28  # 对照装货单：架内件密、外廓仍空心
    if "铝板架" in name or "铝板" in spec:
        return 0.35
    if "瓦楞" in name or "木板" in name:
        return 0.40
    if "五金" in name or "紧固" in spec or "螺丝" in spec:
        return 0.65
    if "胶" in name or "垫" in name:
        return 0.55
    return 0.35


def run_pack(mats: List[Dict[str, Any]], container: str = "40HQ") -> Dict[str, Any]:
    """
    当量箱 → 自主定柜 N0=max(重量, pack_effective) → 3D 自 N0 递增 can_fit。
    订柜体积用 min(outer, content×k)，禁止把空心架外廓当实心分子。
    """
    from packing_assistant.tools.booking import compute_booking, pack_with_auto_containers

    net = sum(float(m.get("total_weight_kg") or 0) for m in mats)

    # 材料行 → API boxes（1 行 = 1 当量箱）+ 订柜体积字段
    boxes = []
    for i, m in enumerate(mats, 1):
        L = int(round(float(m["length_mm"])))
        W = int(round(float(m["width_mm"])))
        H = int(round(float(m["height_mm"])))
        outer_m3 = L * W * H / 1e9
        fill = _crate_fill_hint(m)
        content_m3 = outer_m3 * fill
        longish = L >= 4000
        net_w = float(m.get("weight_kg") or 0)
        name = str(m.get("name") or "")
        # 矮货/板材可上二层（装货单常见：铁架底、板箱上）；1.1m 架 H=1750 仅底层
        can_stack = (H <= 1200 and not longish) or ("铝板" in name) or ("瓦楞" in name) or ("木板" in name) or ("五金" in name)
        prefer_bottom = longish or ("铁件架" in name) or net_w >= 1200
        boxes.append(
            {
                "box_id": f"CRATE-{i:03d}",
                "box_type": "当量箱",
                "outer_size_mm": {"length": L, "width": W, "height": H},
                "outer_m3": round(outer_m3, 6),
                "content_m3": round(content_m3, 6),
                "crate_fill_ratio": fill,
                "booking_volume_m3": round(min(outer_m3, content_m3 * 1.50), 6),
                "net_weight_kg": net_w,
                "gross_weight_kg": net_w + 40,  # 箱皮粗估，勿过大抬高重量柜
                "stackable": can_stack and not prefer_bottom,
                "prefer_bottom": prefer_bottom,
                "special_attributes": ["超长"] if longish else [],
                "name": name,
                "content": [
                    {
                        "name": m.get("name"),
                        "quantity": 1,
                        "material_id": m.get("id"),
                        "outer_size_mm": {
                            "length": max(1, int(L * 0.9)),
                            "width": max(1, int(W * 0.7)),
                            "height": max(1, int(H * fill / 0.7)) if fill > 0 else max(1, H // 3),
                        },
                    }
                ],
            }
        )

    t0 = time.time()
    booking = compute_booking(boxes=boxes, container_type=container, fill_ratio=0.82)
    n0 = int(booking.get("n0") or booking.get("containers_needed") or 1)
    print(
        "BOOKING",
        json.dumps(
            {
                "n0": n0,
                "n_weight": booking.get("containers_by_weight"),
                "n_volume": booking.get("containers_by_volume"),
                "binding": booking.get("binding_constraint"),
                "V_eff": booking.get("volume_m3"),
                "outer_sum": (booking.get("volume_detail") or {}).get("crate_outer_m3"),
                "gross_kg": booking.get("gross_kg"),
                "payload": booking.get("payload_kg"),
                "eta": booking.get("fill_ratio"),
                "suspicious": booking.get("volume_suspicious"),
                "warning": booking.get("warning"),
            },
            ensure_ascii=False,
        ),
    )

    plan = pack_with_auto_containers(
        boxes,
        container_type=container,
        n0=n0,
        n_max=min(40, max(n0 + 12, 8)),
        fill_ratio=0.82,
    )
    # N0 柜时未装入明细（解释 2 vs 3，不作订柜改数）
    gap_at_n0: Dict[str, Any] = {}
    used_3d = int(plan.get("containers_used") or 0)
    if used_3d > n0:
        from packing_assistant.tools.bin3d import pack_boxes_api

        trial = pack_boxes_api(boxes, container_type=container, max_containers=n0)
        unp = list(trial.get("unpacked_box_ids") or [])
        by_name: Dict[str, int] = {}
        for bid in unp:
            bx = next((b for b in boxes if b.get("box_id") == bid), None)
            tag = str((bx or {}).get("name") or bid).split("|")[0].strip()[:28]
            by_name[tag] = by_name.get(tag, 0) + 1
        gap_at_n0 = {
            "n0": n0,
            "unpacked_count": len(unp),
            "unpacked_by_type": by_name,
            "note": "3D 在 N0 柜差这些箱；属成箱/摆柜上界，不改订柜 N0",
        }
        print("GAP_AT_N0", json.dumps(gap_at_n0, ensure_ascii=False))

    snap = {
        "round": 0,
        "n0": n0,
        "booking_containers": n0,  # 订柜口径
        "layout_containers": plan.get("containers_used"),  # 3D 摆柜口径
        "max_containers": plan.get("n_tried", [{}])[-1].get("n") if plan.get("n_tried") else n0,
        "boxes": len(boxes),
        "can_fit": plan.get("can_fit"),
        "containers_used": plan.get("containers_used"),
        "space": plan.get("outer_space_utilization") or plan.get("space_utilization"),
        "booking_volume_util": plan.get("booking_volume_utilization"),
        "floor": plan.get("floor_utilization_avg"),
        "weight": plan.get("weight_utilization"),
        "n_tried": plan.get("n_tried"),
        "gap_at_n0": gap_at_n0,
        "risk": "N/A_estimate",
        "risk_level": "estimate",
        "ship_ok": bool(plan.get("can_fit")),
        "struct_fail": 0,
        "engine": plan.get("engine"),
        "volume_suspicious": booking.get("volume_suspicious"),
        "binding_constraint": booking.get("binding_constraint"),
    }
    print("SNAPSHOT", json.dumps(snap, ensure_ascii=False))
    best = {
        "snapshot": snap,
        "booking": {
            "n0": n0,
            "containers_by_weight": booking.get("containers_by_weight"),
            "containers_by_volume": booking.get("containers_by_volume"),
            "binding_constraint": booking.get("binding_constraint"),
            "volume_m3": booking.get("volume_m3"),
            "crate_outer_m3": (booking.get("volume_detail") or {}).get("crate_outer_m3"),
            "gross_kg": booking.get("gross_kg"),
            "payload_kg": booking.get("payload_kg"),
            "fill_ratio": booking.get("fill_ratio"),
            "volume_suspicious": booking.get("volume_suspicious"),
            "warning": booking.get("warning"),
        },
        "steps": [
            {
                "node": "booking+bin3d",
                "message": (
                    f"当量箱 {len(boxes)} 只 | N0={n0} "
                    f"(重量柜{booking.get('containers_by_weight')}/"
                    f"有效体积柜{booking.get('containers_by_volume')}/"
                    f"绑定{booking.get('binding_constraint')}) "
                    f"→ 3D used={plan.get('containers_used')} can_fit={plan.get('can_fit')} "
                    f"V_eff={booking.get('volume_m3')}m³"
                ),
            }
        ],
        "final": "",
        "team_a": {"box_count": len(boxes), "note": "crate_equivalent_booking_n0"},
        "boxes": [
            {
                "box_id": b["box_id"],
                "box_type": b["box_type"],
                "outer": b["outer_size_mm"],
                "net": b["net_weight_kg"],
                "gross": b["gross_weight_kg"],
                "content_m3": b.get("content_m3"),
                "crate_fill_ratio": b.get("crate_fill_ratio"),
                "booking_volume_m3": b.get("booking_volume_m3"),
                "struct": "当量跳过",
                "content": [c.get("name") for c in b.get("content") or []],
            }
            for b in boxes[:30]
        ],
        "container_plan": {
            k: plan.get(k)
            for k in (
                "can_fit",
                "containers_used",
                "space_utilization",
                "outer_space_utilization",
                "booking_volume_utilization",
                "floor_utilization_avg",
                "weight_utilization",
                "engine",
                "cargo_solid_volume_m3",
                "container_inner_volume_m3",
                "n0",
                "n_tried",
            )
        },
    }

    ms = int((time.time() - t0) * 1000)
    return best or {}, ms, n0


def main() -> int:
    OUT.mkdir(parents=True, exist_ok=True)
    if not SITE_XLSX.exists():
        # 工地 Excel 不在本机时：跳过而非 fail（CI / 无 A: 盘）。有文件时仍硬跑。
        print("SKIP vmu1_site_only: site workbook not found:", SITE_XLSX)
        print("HINT: mount job drive or pass local xlsx; precommit --quick skips this script.")
        return 0

    rows = load_site_rows(SITE_XLSX)
    print(f"site summary rows: {len(rows)} from {SITE_XLSX.name}")
    mats, skipped = to_materials(rows)
    by_group = Counter(m["spec"] for m in mats)
    by_por = Counter(m["part_no"] for m in mats)
    net = sum(m["total_weight_kg"] for m in mats)
    pcs = sum(m["quantity"] for m in mats)
    print(f"remaining site materials: {len(mats)} lines, qty_units={pcs}, net≈{net:.1f} kg")
    print("by group:", dict(by_group))
    print("by POR:", dict(by_por))
    print(f"skipped: {len(skipped)}")
    for s in skipped[:15]:
        print("  skip", s)

    # save materials
    import openpyxl

    xlsx = OUT / "materials_vmu1_site_remaining.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "materials"
    cols = [
        "id",
        "name",
        "spec",
        "quantity",
        "weight_kg",
        "total_weight_kg",
        "length_mm",
        "width_mm",
        "height_mm",
        "part_no",
        "note",
        "destination",
    ]
    ws.append(cols)
    for m in mats:
        ws.append([m.get(c) for c in cols])
    wb.save(xlsx)
    print("WROTE", xlsx)

    if not mats:
        print("NO MATERIALS — nothing to pack")
        return 0

    best, ms, guess = run_pack(mats, "40HQ")
    rep = {
        "scope": "VMU1 剩余未发 · 仅送工地",
        "source": str(SITE_XLSX),
        "materials_file": str(xlsx),
        "materials_lines": len(mats),
        "qty_units": pcs,
        "net_kg": net,
        "by_group": dict(by_group),
        "by_por": dict(by_por),
        "skipped": skipped,
        "dims_note": "Material_Summary 无逐件尺寸，按物料组/POR 实务估算，非正式提料尺寸",
        "pack": best,
        "ms": ms,
        "initial_guess_containers": guess,
    }
    outj = OUT / "vmu1_site_only_pack.json"
    outj.write_text(json.dumps(rep, ensure_ascii=False, indent=2), encoding="utf-8")
    print("WROTE", outj)

    snap = (best or {}).get("snapshot") or {}
    booking = (best or {}).get("booking") or {}
    # COSCO 40HQ 铭牌：PAYLOAD 28610 kg / CU.CAP 76.4 m3
    payload_kg = float(booking.get("payload_kg") or 28610.0)
    cont_m3 = 76.4
    n0 = int(booking.get("n0") or snap.get("n0") or guess)
    n_wt = int(booking.get("containers_by_weight") or 0)
    n_vol = int(booking.get("containers_by_volume") or 0)
    v_eff = booking.get("volume_m3")
    outer_sum = booking.get("crate_outer_m3")
    prog_c = int(snap.get("layout_containers") or snap.get("containers_used") or 0)
    can_fit = snap.get("can_fit")
    gap = snap.get("gap_at_n0") or {}

    md = [
        "# VMU1 送工地装柜（柜型按 COSCO 40HQ 铭牌）",
        "",
        "## 柜型参数（你提供的 COSCO 铭牌）",
        "",
        "| 项 | 铭牌 | 程序采用 |",
        "|---|---:|---:|",
        "| MAX.WT（最大总重） | 32,500 kg | 32,500 |",
        "| TARE（皮重） | 3,890 kg | 3,890 |",
        "| **PAYLOAD（最大货重）** | **28,610 kg** | **28,610** |",
        "| HIGH | 2.9 m | 内高约 2,698 mm |",
        "| **CU.CAP（容积）** | **76.4 m³** | **76.4** |",
        "",
        "重量与体积**都是硬约束**（自主定柜，**不写死柜数**）：",
        f"- 重量柜数 = ceil(货重 / {payload_kg:.0f})",
        f"- 体积柜数 = ceil(V_eff / (76.4 × η))，η=0.82，V_eff=Σ min(outer, content×k)",
        "- **订柜 N0** = max(重量柜, 有效体积柜) → **给领导订舱/汇报**",
        "- **3D 建议柜数** = 当量外廓摆柜 can_fit 用柜 → **工程上界，可与 N0 不同**",
        "- 两数**不要合成一个硬报**；outer 摆柜率不作订柜依据",
        "",
        f"- **数据源**：`{SITE_XLSX.name}`",
        f"- **范围**：VMU1 送工地",
        "",
        "## 双口径结论（固定两行）",
        "",
        f"| 口径 | 值 | 对外说法 |",
        f"|---|---|---|",
        f"| **订柜 N0** | **{n0}**（重量柜 {n_wt} / 有效体积柜 {n_vol} / 绑定 {booking.get('binding_constraint')}） | **给领导订舱/汇报用这个** |",
        f"| **3D 建议柜数** | **{prog_c}**（can_fit={can_fit}） | 当前成箱模型下的摆柜上界 |",
        f"| FST0003 装货单对照 | 2 柜（约 19.8 t + 12.7 t） | 回归样例，非系统约束 |",
        "",
        f"| 明细 | 值 |",
        f"|---|---|",
        f"| 建议柜型 | **40HQ** |",
        f"| V_eff 订柜体积 | {v_eff} m³（箱外廓合计 {outer_sum} m³，已打折） |",
        f"| 净重（当量） | {net:.1f} kg |",
        f"| 外廓摆柜率（仅展示） | {snap.get('space')} |",
        f"| 订柜有效体积率 | {snap.get('booking_volume_util')} |",
        f"| 底面积 / 重量利用率 | {snap.get('floor')} / {snap.get('weight')} |",
        f"| 体积可疑 | {booking.get('volume_suspicious')} {booking.get('warning') or ''} |",
        f"| 当量箱数 | {snap.get('boxes')} |",
        "",
        f"> 按重量和有效包装体积，剩余工地货约 **{n0} 个 40HQ**；  ",
        f"> 当前自动成箱模型做三维摆柜时约需 **{prog_c} 柜**"
        + (
            f"（N0 柜时差 {gap.get('unpacked_count')} 箱，见下）"
            if gap.get("unpacked_count")
            else ""
        )
        + "；  ",
        f"> 与历史装货单 2 柜同量级；后续可通过贴近装货单的合箱把 3D 往 N0 收。",
        "",
    ]
    if gap.get("unpacked_by_type"):
        md.extend(
            [
                "### N0 柜时未装入（摆柜差距，不改订柜）",
                "",
                "| 当量箱类型 | 件数 |",
                "|---|---:|",
            ]
        )
        for k, v in sorted(
            (gap.get("unpacked_by_type") or {}).items(), key=lambda x: -x[1]
        ):
            md.append(f"| {k} | {v} |")
        md.append("")
    md.extend(
        [
        "## 纳入 POR（VMU1·工地·有剩余量）",
        "",
        "| POR | 物料组 | 当量箱数 | 原料件数 |",
        "|---|---|---:|---:|",
        ]
    )
    por_info = defaultdict(lambda: {"group": "", "qty": 0, "raw": 0.0})
    for m in mats:
        por_info[m["part_no"]]["group"] = m["spec"]
        por_info[m["part_no"]]["qty"] += int(m["quantity"])
        por_info[m["part_no"]]["raw"] = max(
            por_info[m["part_no"]]["raw"], float(m.get("raw_qty") or 0)
        )
    for por, info in sorted(por_info.items()):
        md.append(
            f"| {por} | {info['group']} | {info['qty']} | {info['raw']:.0f} |"
        )
    md.extend(
        [
            "",
            "## 未纳入",
            "",
            "- **送工厂**全部（FAC0008/BAL/BGL/FAC0007 等）— 不在本次范围",
            "- **VMU02/03/04** 送工地行（本表有，但领导只问 VMU1）",
            "- 已到=未到=0 的空量行（BBF0006/BOM0013/BSS0010 等）",
            "",
            "## 说明",
            "",
            "1. **订柜 N0 ≠ 3D 建议柜数**：前者给订舱，后者是当前当量外廓的摆柜上界。",
            "2. 件数→当量箱：铁件约 100 件/架（1.1m 为主混型）、铝板约 40 片/架、紧固件约 400 件/箱（贴装货单加密）。",
            "3. 若 **FST0003 / FAC0011** 已部分发运，请用真实剩余件数替换后重跑。",
            "4. 正式订柜前对照提料单尺寸与已发货装货单再校一版。",
            "",
            f"产物：`{outj}`  /  `{xlsx}`",
        ]
    )
    mdp = OUT / "VMU1_送工地_剩余装柜估算.md"
    mdp.write_text("\n".join(md), encoding="utf-8")
    print("WROTE", mdp)
    print("FINAL_CONTAINERS", snap.get("containers_used"), "boxes", snap.get("boxes"))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
