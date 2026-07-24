#!/usr/bin/env python3
"""生成含侧视图的 VMU1 装柜计划 xlsx。"""
from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "output" / "por_vmu_nine"


def _latest_side_bundle() -> tuple[Path | None, list[Path]]:
    """找最新 overview + 同前缀 c01..cNN 分柜图。"""
    out = ROOT / "output"
    overs = sorted(out.glob("side_*_overview.png"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not overs:
        return None, []
    ov = overs[0]
    # side_YYYYMMDD_HHMMSS_overview.png -> prefix side_YYYYMMDD_HHMMSS
    prefix = ov.name.replace("_overview.png", "")
    per = sorted(out.glob(f"{prefix}_c*.png"))
    return ov, per


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    overview, per_imgs = _latest_side_bundle()
    if overview:
        shutil.copy2(overview, OUT_DIR / overview.name)
        for p in per_imgs:
            shutil.copy2(p, OUT_DIR / p.name)
        print("images overview", overview.name, "per", len(per_imgs))
    else:
        # 回退旧单图
        for name in ("side_20260724_181818.png", "side_20260724_181814.png"):
            p = ROOT / "output" / name
            if p.exists():
                shutil.copy2(p, OUT_DIR / p.name)

    wb = Workbook()
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=16, color="1F4E79")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    ok_fill = PatternFill("solid", fgColor="C6EFCE")

    # 01 摘要
    ws = wb.active
    ws.title = "01-计划摘要"
    ws["A1"] = "VMU1 装柜计划（草案·含侧视图）"
    ws["A1"].font = title_font
    ws.merge_cells("A1:B1")
    ws["A2"] = "9智能体方案 PKG-20260724_185245_dd9cee61 · dense密装 · 联网对照业务计划"
    ws["A2"].font = Font(italic=True, color="666666")

    summary = [
        ("项目", "SLTO / PD218 · VMU01"),
        ("范围", "POR/VMU/VMU1 待发（不含 POR/已发货）"),
        ("方案号", "PKG-20260724_185245_dd9cee61"),
        ("柜型", "40HQ"),
        ("计划柜数", "14（密装；原模块外廓16）"),
        ("成箱数", "41（结构全过；dense）"),
        ("总净重约(kg)", "39497"),
        ("总件数", "2512"),
        ("能否装下（几何）", "是"),
        ("能否出运（合规）", "可讨论出运（WARN）"),
        ("容积/底面积/重量", "约23% / 48% / 13%"),
        ("箱外廓/货件体积", "约249 m³ / 44 m³（箱内填充均~18%）"),
        ("装箱模式", "dense 密装：短件贴货；≥3.5m 保留1150×1200模块"),
        ("装载引擎", "python-laff-3d"),
        ("数据说明", "BGL/FAC7/12/BAL 真实；FAC0008 估算；BAL截面估算"),
        ("编制日期", "2026-07-24"),
    ]
    ws["A4"] = "项"
    ws["B4"] = "内容"
    ws["A4"].fill = header_fill
    ws["B4"].fill = header_fill
    ws["A4"].font = header_font
    ws["B4"].font = header_font
    for i, (a, b) in enumerate(summary, 5):
        ws[f"A{i}"] = a
        ws[f"B{i}"] = b
        ws[f"A{i}"].border = thin
        ws[f"B{i}"].border = thin
        if "可讨论" in b or "WARN" in b:
            ws[f"B{i}"].fill = warn_fill
        if a.startswith("能否装下"):
            ws[f"B{i}"].fill = ok_fill

    ws["A20"] = "分票建议"
    ws["A20"].font = Font(bold=True, size=12)
    ws["A21"] = "票次"
    ws["B21"] = "内容"
    ws["C21"] = "建议柜量"
    for col in "ABC":
        ws[f"{col}21"].fill = header_fill
        ws[f"{col}21"].font = header_font
    tickets = [
        ("A", "铝料 BAL0004 为主（含6-7.2m超长）", "约6-8x40HQ"),
        ("B", "铝板 FAC0008估 + FAC0012 + FAC0007", "约4-6x40HQ"),
        ("C", "玻璃 BGL0003", "约1-2x40HQ"),
        ("D", "拉弯 BAL0005/0020 与尾货", "并入A/B或0-1柜"),
        ("合计", "合票拼柜 dense 密装收敛", "14x40HQ"),
    ]
    for i, row in enumerate(tickets, 22):
        ws[f"A{i}"], ws[f"B{i}"], ws[f"C{i}"] = row
        for col in "ABC":
            ws[f"{col}{i}"].border = thin

    ws["A28"] = "联网评审要点（摘要）"
    ws["A28"].font = Font(bold=True, size=12)
    points = [
        "1. 产业要求装柜计划含柜型/件数/配重绑扎与可视化；本表已嵌侧视图。",
        "2. 幕墙出运：铝料防护+木架；玻璃独立木箱；板材防划——对应WARN作业项。",
        "3. 利用率按箱外廓计；超长+薄板天然低于普货70-85%；密装已减2柜/约13%外廓。",
        "4. FAC0008仍为估算，全量表到位后重算；14柜为dense收敛结果。",
        "5. WARN不打回、REJECT才打回——与可讨论出运一致。",
        "6. packing_options.dense_mode=true 已接入 box_scheme→run_packing。",
    ]
    for i, t in enumerate(points, 29):
        ws[f"A{i}"] = t

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 18

    # 02 货量
    ws2 = wb.create_sheet("02-货量清单")
    headers = ["料单", "品类", "件数", "净重约kg", "尺寸说明", "去向", "数据状态"]
    ws2.append(headers)
    for col in range(1, 8):
        c = ws2.cell(1, col)
        c.fill = header_fill
        c.font = header_font
    for r in [
        ["FAC0008", "3mm铝板", 906, 17100, "估 W1.4-2.3m x H0.8-1.5m x T3mm", "工厂", "估算"],
        ["FAC0007", "25mm蜂窝铝板", 22, 243, "真实 WxHxT=25", "工厂", "真实"],
        ["FAC0012", "3mm铝板小件", 40, 16, "真实 140x357x3", "工厂", "真实"],
        ["BGL0003", "中空夹胶玻璃", 61, 2607, "真实 WxH；厚约40估", "工厂", "真实WxH"],
        ["BAL0004", "铝型材", 1206, 16700, "真实 L=4.0-7.2m；截面估", "工厂", "真实L"],
        ["BAL0005", "拉弯铝料", 265, 2706, "真实 L=3.5-4.5m", "工厂", "真实L"],
        ["BAL0020", "拉弯铝料", 12, 128, "真实 L=3.8m", "工厂", "真实L"],
        ["合计", "", 2512, 39497, "", "", ""],
    ]:
        ws2.append(r)
    for col, w in enumerate([12, 16, 10, 12, 40, 10, 12], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # 03 风险
    ws3 = wb.create_sheet("03-风险与动作")
    ws3.append(["等级", "项", "动作", "是否打回"])
    for col in range(1, 5):
        c = ws3.cell(1, col)
        c.fill = header_fill
        c.font = header_font
    for r in [
        ["必做", "超长铝料沿柜长", "加强绑扎+照片存档", "否(WARN)"],
        ["必做", "玻璃防倾防碎", "专用垫木/隔离", "否(WARN)"],
        ["建议", "单箱毛重偏高", "优先铁架/钢骨", "否(WARN)"],
        ["建议", "VGM前复核毛重", "含箱自重", "-"],
        ["数据", "FAC0008估算", "完整Excel后重算柜数", "-"],
        ["策略", "WARN", "人工确认后订舱", "不打回"],
        ["策略", "REJECT", "回装箱或加柜", "打回"],
    ]:
        ws3.append(r)
    for col, w in enumerate([10, 22, 36, 14], 1):
        ws3.column_dimensions[get_column_letter(col)].width = w

    def add_scaled(ws_sheet, path: Path, anchor: str, width_px: int = 920) -> int:
        if not path.exists():
            return 0
        im = XLImage(str(path))
        ow = float(im.width or width_px)
        oh = float(im.height or 300)
        ratio = width_px / max(ow, 1)
        im.width = width_px
        im.height = int(oh * ratio)
        ws_sheet.add_image(im, anchor)
        return max(8, int(im.height / 18))

    # 04 总览
    ws4 = wb.create_sheet("04-侧视总览")
    ws4["A1"] = "16柜侧视总览（拼版）"
    ws4["A1"].font = title_font
    ws4["A2"] = "终局：16x40HQ | 43箱 | 空间约24% | 重量约12% | ship_ok=true"
    ws4["A3"] = "旧版只有1柜图：因把多柜箱子画进同一条12m轮廓。现已按柜出图。"
    ws4["A3"].font = Font(color="666666", size=10)
    if overview:
        local_ov = OUT_DIR / overview.name
        if local_ov.exists():
            add_scaled(ws4, local_ov, "A5", 1000)
    ws4.column_dimensions["A"].width = 100

    # 05 分柜
    ws5img = wb.create_sheet("05-分柜侧视图")
    ws5img["A1"] = "分柜侧视图（第1-16柜）"
    ws5img["A1"].font = title_font
    ws5img["A2"] = "每张图 = 一个40HQ柜内箱位（沿柜长）"
    row_cursor = 4
    if per_imgs:
        for idx, p in enumerate(per_imgs, 1):
            local = OUT_DIR / p.name
            ws5img[f"A{row_cursor}"] = f"第 {idx} 柜 · {p.name}"
            ws5img[f"A{row_cursor}"].font = Font(bold=True, size=11)
            row_cursor += 1
            used = add_scaled(ws5img, local, f"A{row_cursor}", 880)
            row_cursor += used + 2
    else:
        ws5img["A4"] = "无分柜图"
    ws5img.column_dimensions["A"].width = 100

    # 06 联网评审
    ws5 = wb.create_sheet("06-联网评审")
    ws5["A1"] = "VMU1 装柜计划 · 联网评审"
    ws5["A1"].font = title_font
    review_lines = [
        "【总评】约 8.4/10（含16柜侧视后可视化交付明显改进）",
        "",
        "【为何曾只有1柜图】",
        "- 旧 visualizer 把所有柜的箱子画进同一条 12m 轮廓，多柜重叠，看起来像1柜。",
        "- 现已按 container_no 输出 16 张分柜图 + 1 张总览，并嵌入本 xlsx。",
        "",
        "【与产业对齐】",
        "- 装柜计划应含柜型、件数、配重/利用率、绑扎与可视化：本表已覆盖分柜侧视。",
        "- 幕墙出运：型材防护+木架；玻璃独立木箱；板材防划——与 WARN 作业项一致。",
        "- 40HQ 适合轻泡/高货；利用率 24%/重量 12% 对超长+薄板可解释。",
        "",
        "【WARN 是否打回】不打回（黄灯人工确认）；REJECT 才打回。",
        "",
        "【环境】Node24 + Maven/JDK17 可起 skjolber；PATH 上 java 1.7 勿混用。",
        "",
        "【下一步】FAC0008 全量表；型材截面；分票日程；可选 skjolber A/B。",
    ]
    for i, line in enumerate(review_lines, 3):
        ws5[f"A{i}"] = line
    ws5.column_dimensions["A"].width = 100

    # 07 不纳入
    ws6 = wb.create_sheet("07-不纳入")
    ws6.append(["说明", "示例"])
    ws6["A1"].fill = header_fill
    ws6["B1"].fill = header_fill
    ws6["A1"].font = header_font
    ws6["B1"].font = header_font
    ws6.append(["已发货目录内 POR", "见 POR/已发货"])
    ws6.append(["FAC0011 送工地铝板", "Material_Summary 已出货"])
    ws6.append(["FST0003/0022 铁件垫片", "如 CSNU6547612 等"])
    ws6.column_dimensions["A"].width = 28
    ws6.column_dimensions["B"].width = 40

    out = OUT_DIR / "VMU1_装柜计划_草案_含图.xlsx"
    wb.save(out)
    print("WROTE", out)
    print("size_kb", round(out.stat().st_size / 1024, 1))
    print("per_images", len(per_imgs), "overview", overview)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
