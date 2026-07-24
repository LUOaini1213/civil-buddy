#!/usr/bin/env python3
"""
解析 test/ 装箱单 PDF → 同一项目拼柜（默认）→ DeepSeek 润色汇总 → Excel/HTML。

用法:
  python scripts/run_test_shipments.py
  python scripts/run_test_shipments.py --mode project --deepseek
  python scripts/run_test_shipments.py --mode per_container
  python scripts/run_test_shipments.py --mode both
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
import traceback
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))


def setup_deepseek_env() -> str:
    """配置 DeepSeek；优先 deepseek api.txt / .env。"""
    try:
        from dotenv import load_dotenv

        load_dotenv(ROOT / ".env")
    except Exception:
        pass

    os.environ.setdefault("LLM_MODEL", "deepseek-v4-flash")
    os.environ.setdefault("DEEPSEEK_MODEL", "deepseek-v4-flash")
    os.environ.setdefault("DEEPSEEK_BASE_URL", "https://api.deepseek.com")
    os.environ.setdefault("OPENAI_BASE_URL", "https://api.deepseek.com")
    os.environ.setdefault("LLM_BASE_URL", "https://api.deepseek.com")

    for name in ("deepseek api.txt", "deepseek_api.txt", "deepseek-api.txt"):
        key_file = ROOT / name
        if key_file.exists():
            key = key_file.read_text(encoding="utf-8-sig").strip().splitlines()[0].strip()
            if key.startswith("sk-"):
                os.environ["DEEPSEEK_API_KEY"] = key
                os.environ["OPENAI_API_KEY"] = key
                os.environ["LLM_API_KEY"] = key
                return f"file:{name}"

    if os.getenv("DEEPSEEK_API_KEY") or os.getenv("OPENAI_API_KEY") or os.getenv("LLM_API_KEY"):
        key = os.getenv("DEEPSEEK_API_KEY") or os.getenv("OPENAI_API_KEY") or os.getenv("LLM_API_KEY")
        os.environ["OPENAI_API_KEY"] = key or ""
        os.environ["DEEPSEEK_API_KEY"] = key or ""
        return "env"
    return "none"


def ping_llm() -> Dict[str, Any]:
    from packing_assistant.llm import chat, llm_config

    cfg = llm_config()
    t0 = time.time()
    text = chat(
        system="你是测试助手，只回复一个词：OK",
        user="ping",
        temperature=0,
        max_tokens=32,
    )
    return {
        "model": cfg.get("model"),
        "base_url": cfg.get("base_url"),
        "has_key": bool(cfg.get("api_key")),
        "reply": text,
        "ms": int((time.time() - t0) * 1000),
        "ok": bool(text) and not str(text).startswith("[LLM_ERROR]"),
    }


def try_llm_models(models: List[str]) -> Dict[str, Any]:
    last: Dict[str, Any] = {}
    for m in models:
        os.environ["LLM_MODEL"] = m
        os.environ["DEEPSEEK_MODEL"] = m
        last = ping_llm()
        last["tried_model"] = m
        if last.get("ok"):
            return last
    return last


def _group_by_container(materials: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    groups: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for m in materials:
        ctn = (m.get("container_no") or "").strip() or "_UNKNOWN_"
        groups[ctn].append(m)
    return dict(groups)


def _estimate_max_containers(materials: List[Dict[str, Any]], floor: int = 1) -> int:
    """按净重 + 材料件数粗估柜数上限（40HQ 有效载重约 26t，保守 18t/柜）。"""
    net = sum(float(m.get("total_weight_kg") or 0) for m in materials)
    by_weight = int(net / 18000) + 1
    by_count = (len(materials) + 4) // 5
    guess = max(floor, by_weight, by_count, 1)
    return min(guess, 12)


def _run_one_group(
    materials: List[Dict[str, Any]],
    *,
    label: str,
    container_type: str = "40HQ",
    max_containers: int = 1,
) -> Dict[str, Any]:
    from packing_assistant.harness import (
        apply_user_confirmation,
        run_team_a,
        run_team_b,
    )

    mats = []
    for i, m in enumerate(materials, 1):
        mm = dict(m)
        mm["id"] = f"M{i:03d}"
        mats.append(mm)

    sa = run_team_a(f"项目拼柜:{label}", materials=mats)
    guess = _estimate_max_containers(mats, floor=max_containers)
    guess = max(guess, max_containers)
    state = None
    for mc in range(guess, 13):
        sb = apply_user_confirmation(
            sa,
            action="confirm",
            container_type=container_type,
            max_containers=mc,
        )
        state = run_team_b(sb)
        plan = state.get("container_plan") or {}
        if plan.get("can_fit"):
            break
    return state or {}


def _slim_state(state: Dict[str, Any]) -> Dict[str, Any]:
    boxes = []
    for b in state.get("boxes") or []:
        bb = dict(b)
        sc = bb.get("structure_calc") or {}
        if "calc_report_md" in sc:
            sc = {k: v for k, v in sc.items() if k != "calc_report_md"}
            sc["calc_report_md_len"] = len(b.get("structure_calc", {}).get("calc_report_md") or "")
            bb["structure_calc"] = sc
        if bb.get("structure_detail") and "calc_report_md" in (bb["structure_detail"] or {}):
            sd = dict(bb["structure_detail"])
            sd["calc_report_md_len"] = len(sd.pop("calc_report_md") or "")
            bb["structure_detail"] = sd
        boxes.append(bb)
    final = state.get("final_response") or ""
    return {
        "phase": state.get("phase"),
        "status": state.get("status"),
        "boxes": boxes,
        "container_plan": state.get("container_plan"),
        "risk_report": {
            k: (state.get("risk_report") or {}).get(k)
            for k in (
                "passed",
                "compliance_score",
                "level",
                "risks",
                "blockers",
                "explanation",
                "cog",
            )
        },
        "views_keys": list((state.get("views") or {}).keys()),
        "final_response": final[:3000],
        "llm_used": "LLM:" in final or "deepseek" in final.lower(),
    }


def _detail_from_state(
    state: Dict[str, Any],
    *,
    container_no: str,
    materials: List[Dict[str, Any]],
) -> Dict[str, Any]:
    plan = state.get("container_plan") or {}
    risk = state.get("risk_report") or {}
    n_boxes = len(state.get("boxes") or [])
    return {
        "container_no": container_no,
        "materials": len(materials),
        "net_kg": round(sum(float(m.get("total_weight_kg") or 0) for m in materials), 2),
        "boxes": n_boxes,
        "box_types": [b.get("box_type") for b in (state.get("boxes") or [])],
        "can_fit": bool(plan.get("can_fit")),
        "containers_used": plan.get("containers_used"),
        "space_utilization": plan.get("space_utilization"),
        "space_best": plan.get("space_utilization_best_container"),
        "floor_avg": plan.get("floor_utilization_avg"),
        "weight_utilization": plan.get("weight_utilization"),
        "engine": plan.get("engine"),
        "risk_level": risk.get("level"),
        "risk_score": risk.get("compliance_score"),
        "llm_used": "LLM:" in (state.get("final_response") or ""),
        "final_response_preview": (state.get("final_response") or "")[:500],
        "layout_plan": {
            "container_type": plan.get("container_type"),
            "layout": plan.get("layout"),
            "metrics_note": plan.get("metrics_note"),
            "space_utilization": plan.get("space_utilization"),
            "floor_utilization_avg": plan.get("floor_utilization_avg"),
            "containers_used": plan.get("containers_used"),
        },
    }


def _svg_three_views(plan: Dict[str, Any], colors: Optional[List[str]] = None) -> str:
    """从 layout 生成俯视/侧视/正视 SVG；多柜时按算法柜号分块。"""
    layout_all = plan.get("layout") or []
    if not layout_all:
        return "<p class='muted'>无布局数据</p>"
    colors = colors or [
        "#3B82F6",
        "#06B6D4",
        "#22C55E",
        "#EAB308",
        "#A78BFA",
        "#F97316",
        "#EC4899",
        "#94A3B8",
    ]
    from packing_assistant.tools.bin3d import CONTAINER_INNER

    ctype = plan.get("container_type") or "40HQ"
    spec = CONTAINER_INNER.get(ctype) or CONTAINER_INNER["40HQ"]
    CL, CW, CH = float(spec["L"]), float(spec["W"]), float(spec["H"])

    # 按算法柜号分组
    by_cn: Dict[Any, List[Dict[str, Any]]] = defaultdict(list)
    for p in layout_all:
        by_cn[p.get("container_no") or 1].append(p)
    if not by_cn:
        by_cn[1] = layout_all

    def one(view: str, title: str, layout: List[Dict[str, Any]]) -> str:
        Wsvg, Hsvg, pad = 420, 160, 18
        if view == "top":
            max_a, max_b = CL, CW
        elif view == "side":
            max_a, max_b = CL, CH
        else:
            max_a, max_b = CW, CH
        sx = (Wsvg - 2 * pad) / max_a
        sy = (Hsvg - 2 * pad) / max_b
        scale = min(sx, sy)
        rects = []
        for i, p in enumerate(layout):
            pos, size = p.get("position") or {}, p.get("size") or {}
            if view == "top":
                a, b, da, db = pos.get("x", 0), pos.get("y", 0), size.get("dx", 0), size.get("dy", 0)
            elif view == "side":
                a, b, da, db = pos.get("x", 0), pos.get("z", 0), size.get("dx", 0), size.get("dz", 0)
            else:
                a, b, da, db = pos.get("y", 0), pos.get("z", 0), size.get("dy", 0), size.get("dz", 0)
            x = pad + float(a) * scale
            y = pad + (max_b - float(b) - float(db)) * scale
            w = max(float(da) * scale, 1)
            h = max(float(db) * scale, 1)
            c = colors[i % len(colors)]
            label = p.get("box_id") or ""
            rects.append(
                f'<rect x="{x:.1f}" y="{y:.1f}" width="{w:.1f}" height="{h:.1f}" '
                f'fill="{c}" fill-opacity="0.85" stroke="#e2e8f0" stroke-width="1"/>'
                f'<text x="{x + 3:.1f}" y="{y + 12:.1f}" fill="#fff" font-size="10">{label}</text>'
            )
        border = (
            f'<rect x="{pad}" y="{pad}" width="{max_a * scale:.1f}" height="{max_b * scale:.1f}" '
            f'fill="none" stroke="#64748b" stroke-width="2"/>'
        )
        return (
            f'<div class="view"><h4>{title}</h4>'
            f'<svg width="{Wsvg}" height="{Hsvg}" xmlns="http://www.w3.org/2000/svg">'
            f"{border}{''.join(rects)}</svg></div>"
        )

    blocks = []
    for cn in sorted(by_cn.keys(), key=lambda x: (str(type(x)), str(x))):
        layout = by_cn[cn]
        title_prefix = f"算法柜#{cn} · " if len(by_cn) > 1 else ""
        blocks.append(
            f'<div class="algo-ctn"><p class="muted">{title_prefix}boxes={len(layout)}</p>'
            '<div class="views">'
            + one("top", f"{title_prefix}俯视 top", layout)
            + one("side", f"{title_prefix}侧视 side", layout)
            + one("front", f"{title_prefix}正视 front", layout)
            + "</div></div>"
        )
    return "".join(blocks)


def build_reports(summary: Dict[str, Any], out_dir: Path) -> None:
    """Excel + HTML 汇总（含三视图 + LLM 摘要）。"""
    rows_html = []
    views_blocks = []
    llm_blocks = []
    mode = summary.get("mode") or "project"
    for s in summary.get("shipments") or []:
        if not s.get("ok"):
            rows_html.append(
                f"<tr class='fail'><td>{s.get('file')}</td><td colspan='11'>FAIL: {s.get('error')}</td></tr>"
            )
            continue
        for c in s.get("containers_detail") or []:
            rows_html.append(
                "<tr>"
                f"<td>{s.get('file')}</td>"
                f"<td>{c.get('container_no')}</td>"
                f"<td>{c.get('materials')}</td>"
                f"<td>{c.get('boxes')}</td>"
                f"<td>{c.get('can_fit')}</td>"
                f"<td>{c.get('containers_used')}</td>"
                f"<td>{c.get('space_utilization')}</td>"
                f"<td>{c.get('space_best')}</td>"
                f"<td>{c.get('floor_avg')}</td>"
                f"<td>{c.get('weight_utilization')}</td>"
                f"<td>{c.get('risk_level')}/{c.get('risk_score')}</td>"
                f"<td>{c.get('engine')}</td>"
                f"<td>{'是' if c.get('llm_used') else '-'}</td>"
                "</tr>"
            )
            plan = c.get("layout_plan") or {}
            if plan.get("layout"):
                views_blocks.append(
                    f"<section class='ship'><h3>{s.get('file')} · {c.get('container_no')}</h3>"
                    f"<p class='muted'>{plan.get('metrics_note') or ''}</p>"
                    f"{_svg_three_views(plan)}</section>"
                )
            if c.get("final_response_preview"):
                llm_blocks.append(
                    f"<section class='ship'><h3>LLM · {s.get('file')} · {c.get('container_no')}</h3>"
                    f"<pre class='llm'>{_escape(c.get('final_response_preview') or '')}</pre></section>"
                )

    llm_info = summary.get("llm") or {}
    html = f"""<!DOCTYPE html>
<html lang="zh-CN"><head><meta charset="utf-8"/>
<title>test 装箱单跑批汇总 · {mode}</title>
<style>
body{{font-family:Microsoft YaHei,sans-serif;margin:24px;background:#0f1419;color:#e7ecf3}}
table{{border-collapse:collapse;width:100%;font-size:12px}}
th,td{{border:1px solid #2a3a52;padding:6px 8px;text-align:left}}
th{{background:#1a2332;color:#93c5fd}}
tr:nth-child(even){{background:#151c27}}
.fail{{color:#fca5a5}}
h1{{font-size:1.2rem}}
.badge{{display:inline-block;padding:2px 8px;border-radius:8px;background:#14532d;margin-left:8px}}
.views{{display:flex;flex-wrap:wrap;gap:12px;margin:8px 0 12px}}
.view{{background:#151c27;border:1px solid #2a3a52;border-radius:8px;padding:8px}}
.view h4{{margin:0 0 6px;color:#93c5fd;font-size:13px}}
.ship{{margin:18px 0;padding:12px;border:1px solid #2a3a52;border-radius:10px;background:#1a2332}}
.muted{{color:#94a3b8;font-size:12px}}
.note{{background:#1e293b;padding:10px 12px;border-radius:8px;font-size:13px;line-height:1.5;margin:12px 0}}
.llm{{white-space:pre-wrap;font-size:12px;line-height:1.45;color:#cbd5e1;margin:0}}
.algo-ctn{{margin-bottom:10px}}
</style></head><body>
<h1>test/ 装箱单跑批 · mode={mode}
<span class="badge">ok={summary.get('ok')} fail={summary.get('fail')}</span>
<span class="badge">LLM={'OK' if llm_info.get('ok') else 'N/A'} {llm_info.get('model') or ''}</span>
</h1>
<div class="note">
<strong>拼柜模式：</strong>
<code>project</code>=同一 PDF/项目全部材料合并优化柜数（推荐）；
<code>per_container</code>=按 PDF 柜号拆分（复现现场分柜）；
<code>both</code>=两种都跑便于对比。
<br/>
<strong>利用率：</strong>普货满载约 70–85%；钢结构铁架常更低。指标=箱体外体积/柜内容积。
已按货包络定制外廓。三视图按算法柜分块。
</div>
<table>
<thead><tr>
<th>文件</th><th>组/柜</th><th>材料行</th><th>生成箱</th><th>can_fit</th>
<th>算法用柜</th><th>容积利用率</th><th>最满柜</th><th>底面积均</th><th>重量利用率</th><th>风险</th><th>引擎</th><th>LLM</th>
</tr></thead>
<tbody>
{''.join(rows_html)}
</tbody></table>
<h2>三视图（俯视 / 侧视 / 正视）</h2>
{''.join(views_blocks) if views_blocks else '<p class="muted">无布局可绘</p>'}
<h2>DeepSeek 汇总预览</h2>
{''.join(llm_blocks) if llm_blocks else '<p class="muted">无 LLM 输出（未配置或调用失败）</p>'}
</body></html>"""
    (out_dir / "report.html").write_text(html, encoding="utf-8")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "拼柜汇总"
        headers = [
            "文件",
            "组/柜",
            "材料行",
            "净重kg",
            "生成箱数",
            "can_fit",
            "算法用柜",
            "容积利用率",
            "最满柜容积率",
            "底面积均",
            "重量利用率",
            "风险等级",
            "合规分",
            "引擎",
            "LLM",
            "箱型列表",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E3A5F")
        for s in summary.get("shipments") or []:
            if not s.get("ok"):
                ws.append([s.get("file"), "", "", "", "", False, "", "", "", "", "", "FAIL", s.get("error"), "", "", ""])
                continue
            for c in s.get("containers_detail") or []:
                ws.append(
                    [
                        s.get("file"),
                        c.get("container_no"),
                        c.get("materials"),
                        c.get("net_kg"),
                        c.get("boxes"),
                        c.get("can_fit"),
                        c.get("containers_used"),
                        c.get("space_utilization"),
                        c.get("space_best"),
                        c.get("floor_avg"),
                        c.get("weight_utilization"),
                        c.get("risk_level"),
                        c.get("risk_score"),
                        c.get("engine"),
                        "yes" if c.get("llm_used") else "",
                        ",".join(c.get("box_types") or []),
                    ]
                )
        ws2 = wb.create_sheet("装运级")
        ws2.append(
            [
                "文件",
                "模式",
                "材料行",
                "PDF柜数",
                "处理组数",
                "总箱数",
                "算法用柜合计",
                "全部can_fit",
                "耗时ms",
                "ok",
            ]
        )
        for s in summary.get("shipments") or []:
            used_sum = 0
            for c in s.get("containers_detail") or []:
                try:
                    used_sum += int(c.get("containers_used") or 0)
                except Exception:
                    pass
            ws2.append(
                [
                    s.get("file"),
                    s.get("run_mode") or mode,
                    s.get("materials"),
                    len(s.get("containers_in_pdf") or []),
                    s.get("containers_processed"),
                    s.get("boxes_total"),
                    used_sum if s.get("ok") else "",
                    s.get("all_can_fit"),
                    s.get("ms"),
                    s.get("ok"),
                ]
            )
        xlsx_path = out_dir / "report.xlsx"
        wb.save(xlsx_path)
    except Exception as e:
        (out_dir / "report_excel_error.txt").write_text(str(e), encoding="utf-8")


def _escape(s: str) -> str:
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def _run_shipment_project(
    mats: List[Dict[str, Any]],
    *,
    fname: str,
    container_type: str,
) -> Tuple[List[Dict[str, Any]], Dict[str, Any], int, bool]:
    """同一项目：全部材料合并拼柜。"""
    pdf_ctns = sorted(
        {(m.get("container_no") or "").strip() for m in mats if (m.get("container_no") or "").strip()}
    )
    label = f"{fname}:PROJECT"
    print(
        f"  >> PROJECT merge materials={len(mats)} "
        f"pdf_containers={pdf_ctns or ['(none)']} net≈"
        f"{sum(float(m.get('total_weight_kg') or 0) for m in mats):.0f}kg"
    )
    state = _run_one_group(
        mats,
        label=label,
        container_type=container_type,
        max_containers=_estimate_max_containers(mats, floor=1),
    )
    tag = "PROJECT" + (f"←{','.join(pdf_ctns)}" if pdf_ctns else "")
    det = _detail_from_state(state, container_no=tag, materials=mats)
    print(
        f"     boxes={det['boxes']} fit={det['can_fit']} used={det['containers_used']} "
        f"space={det['space_utilization']} best={det['space_best']} "
        f"weight={det['weight_utilization']} "
        f"risk={det['risk_level']}/{det['risk_score']} llm={det.get('llm_used')}"
    )
    return [det], {tag: _slim_state(state)}, det["boxes"], bool(det["can_fit"])


def _run_shipment_per_container(
    mats: List[Dict[str, Any]],
    *,
    fname: str,
    container_type: str,
) -> Tuple[List[Dict[str, Any]], Dict[str, Any], int, bool]:
    groups = _group_by_container(mats)
    containers_detail = []
    boxes_total = 0
    all_fit = True
    per_results: Dict[str, Any] = {}
    for ctn, gmats in groups.items():
        label = f"{fname}:{ctn}"
        print(f"  >> container {ctn} materials={len(gmats)}")
        state = _run_one_group(
            gmats,
            label=label,
            container_type=container_type,
            max_containers=1,
        )
        det = _detail_from_state(state, container_no=ctn, materials=gmats)
        boxes_total += det["boxes"]
        all_fit = all_fit and det["can_fit"]
        containers_detail.append(det)
        per_results[ctn] = _slim_state(state)
        print(
            f"     boxes={det['boxes']} fit={det['can_fit']} used={det['containers_used']} "
            f"space={det['space_utilization']} risk={det['risk_level']}/{det['risk_score']}"
        )
    return containers_detail, per_results, boxes_total, all_fit


def main() -> int:
    ap = argparse.ArgumentParser(description="test/ 装箱单全量跑批（项目拼柜 + DeepSeek）")
    ap.add_argument("--dir", default=str(ROOT / "test"))
    ap.add_argument("--container", default="40HQ")
    ap.add_argument(
        "--mode",
        choices=("project", "per_container", "both"),
        default="project",
        help="project=同一项目合并拼柜（默认）；per_container=按PDF柜号；both=对比",
    )
    ap.add_argument(
        "--deepseek",
        action="store_true",
        default=True,
        help="接入 DeepSeek（默认开启）",
    )
    ap.add_argument("--no-deepseek", action="store_true", help="关闭 LLM")
    ap.add_argument(
        "--out",
        default="",
        help="输出目录，默认 output/test_shipments",
    )
    args = ap.parse_args()

    llm_meta: Dict[str, Any] = {"enabled": False, "ok": False}
    if args.deepseek and not args.no_deepseek:
        src = setup_deepseek_env()
        print(f"LLM 配置来源: {src}")
        print(f"LLM_MODEL={os.getenv('LLM_MODEL')} BASE={os.getenv('OPENAI_BASE_URL')}")
        print("ping DeepSeek ...")
        ping = try_llm_models(
            [
                os.getenv("LLM_MODEL") or "deepseek-v4-flash",
                "deepseek-v4-flash",
                "deepseek-chat",
                "deepseek-v3",
            ]
        )
        llm_meta = {"enabled": True, "source": src, **ping}
        print(json.dumps({k: ping.get(k) for k in ("ok", "model", "tried_model", "ms", "reply")}, ensure_ascii=False))
        if not ping.get("ok"):
            print("WARNING: DeepSeek 不可用，继续规则引擎；finalize/risk 将无 LLM 润色")
    else:
        print("LLM disabled (--no-deepseek)")

    from packing_assistant.tools.packing_list_parser import parse_all_in_dir
    from packing_assistant.tools.dims_override import apply_dims_override, load_override

    parsed = parse_all_in_dir(args.dir)
    override = load_override()
    out_dir = Path(args.out) if args.out else (ROOT / "output" / "test_shipments")
    out_dir.mkdir(parents=True, exist_ok=True)

    summary: Dict[str, Any] = {
        "shipments": [],
        "ok": 0,
        "fail": 0,
        "mode": args.mode,
        "llm": llm_meta,
    }

    modes_to_run = (
        ["project", "per_container"] if args.mode == "both" else [args.mode]
    )

    for pl in parsed:
        fname = pl.get("source_file") or "unknown"
        print("=" * 60)
        print(f"FILE: {fname}")
        if pl.get("error"):
            print("  PARSE ERROR:", pl["error"])
            summary["fail"] += 1
            summary["shipments"].append({"file": fname, "ok": False, "error": pl["error"]})
            continue

        mats = apply_dims_override(pl.get("materials") or [], override)
        print(
            f"  materials={len(mats)} pieces={pl.get('total_pieces')} "
            f"net_kg={pl.get('total_net_kg')} containers={pl.get('containers')}"
        )
        for m in mats[:6]:
            src = m.get("dims_source") or ("est" if m.get("dims_estimated") else "raw")
            print(
                f"    - {m['id']} {m['name'][:36]} q={m['quantity']} "
                f"{m['weight_kg']}kg L={m['length_mm']} [{src}] ctn={m.get('container_no')}"
            )

        if not mats:
            print("  SKIP: no materials")
            summary["fail"] += 1
            summary["shipments"].append({"file": fname, "ok": False, "error": "no materials"})
            (out_dir / f"{Path(fname).stem}_parse.json").write_text(
                json.dumps(pl, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            continue

        stem = Path(fname).stem
        (out_dir / f"{stem}_parse.json").write_text(
            json.dumps({**pl, "materials": mats}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

        file_ok = True
        for run_mode in modes_to_run:
            t0 = time.time()
            try:
                if run_mode == "project":
                    details, per_results, boxes_total, all_fit = _run_shipment_project(
                        mats, fname=fname, container_type=args.container
                    )
                    suffix = "_project"
                else:
                    details, per_results, boxes_total, all_fit = _run_shipment_per_container(
                        mats, fname=fname, container_type=args.container
                    )
                    suffix = "_by_container"

                ms = int((time.time() - t0) * 1000)
                rec = {
                    "file": fname,
                    "ok": True,
                    "run_mode": run_mode,
                    "ms": ms,
                    "materials": len(mats),
                    "total_net_kg": pl.get("total_net_kg"),
                    "containers_in_pdf": pl.get("containers"),
                    "containers_processed": len(details),
                    "boxes_total": boxes_total,
                    "all_can_fit": all_fit,
                    "containers_detail": details,
                    "algo_containers_sum": sum(int(d.get("containers_used") or 0) for d in details),
                }
                print(
                    f"  SHIPMENT OK mode={run_mode} boxes_total={boxes_total} "
                    f"all_fit={all_fit} algo_ctn={rec['algo_containers_sum']} {ms}ms"
                )
                summary["shipments"].append(rec)
                (out_dir / f"{stem}{suffix}.json").write_text(
                    json.dumps(per_results, ensure_ascii=False, indent=2, default=str),
                    encoding="utf-8",
                )
                # 兼容旧文件名
                if run_mode == "project":
                    (out_dir / f"{stem}_result.json").write_text(
                        json.dumps(per_results, ensure_ascii=False, indent=2, default=str),
                        encoding="utf-8",
                    )
            except Exception as e:
                file_ok = False
                print("  FAIL:", e)
                traceback.print_exc()
                summary["shipments"].append(
                    {
                        "file": fname,
                        "ok": False,
                        "run_mode": run_mode,
                        "error": str(e),
                    }
                )

        if file_ok:
            summary["ok"] += 1
        else:
            summary["fail"] += 1

    (out_dir / "summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    build_reports(summary, out_dir)
    print("=" * 60)
    print(f"DONE mode={args.mode} ok={summary['ok']} fail={summary['fail']}")
    print(f"  LLM ok={llm_meta.get('ok')} model={llm_meta.get('model')}")
    print(f"  HTML: {out_dir / 'report.html'}")
    print(f"  Excel: {out_dir / 'report.xlsx'}")
    print(f"  JSON: {out_dir / 'summary.json'}")
    return 0 if summary["fail"] == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
