#!/usr/bin/env python3
"""
跑 test/excel 下的钢结构业务 Excel 测试集（材料→拼柜）。

用法:
  python scripts/build_steel_test_set.py   # 先生成
  python scripts/run_excel_tests.py
  python scripts/run_excel_tests.py --only syn_short_frames,syn_long_6m
  python scripts/run_excel_tests.py --deepseek
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
from pathlib import Path
from typing import Any, Dict, List, Optional

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))


def setup_deepseek() -> Dict[str, Any]:
    try:
        from dotenv import load_dotenv

        load_dotenv(ROOT / ".env")
    except Exception:
        pass
    os.environ.setdefault("LLM_MODEL", "deepseek-v4-flash")
    os.environ.setdefault("DEEPSEEK_MODEL", "deepseek-v4-flash")
    os.environ.setdefault("OPENAI_BASE_URL", "https://api.deepseek.com")
    for name in ("deepseek api.txt", "deepseek_api.txt"):
        kf = ROOT / name
        if kf.exists():
            key = kf.read_text(encoding="utf-8-sig").strip().splitlines()[0].strip()
            if key.startswith("sk-"):
                os.environ["DEEPSEEK_API_KEY"] = key
                os.environ["OPENAI_API_KEY"] = key
                os.environ["LLM_API_KEY"] = key
                return {"source": name, "ok": True}
    return {"source": "env/none", "ok": bool(os.getenv("DEEPSEEK_API_KEY") or os.getenv("OPENAI_API_KEY"))}


def load_materials_xlsx(path: Path) -> List[Dict[str, Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    # 优先 materials sheet
    ws = wb["materials"] if "materials" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c or "").strip() for c in rows[0]]
    # 兼容 full_flow materials
    mats = []
    for row in rows[1:]:
        d = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        if d.get("row_type") and str(d.get("row_type")) not in ("material", "materials", ""):
            continue
        name = d.get("name") or d.get("名称")
        if not name:
            continue
        qty = int(float(d.get("quantity") or d.get("数量") or 1))
        wt = float(d.get("weight_kg") or d.get("单重") or 0)
        tw = float(d.get("total_weight_kg") or 0) or wt * qty
        mats.append(
            {
                "id": str(d.get("id") or f"M{len(mats)+1:03d}"),
                "name": str(name),
                "quantity": max(qty, 1),
                "weight_kg": wt,
                "total_weight_kg": tw,
                "length_mm": float(d.get("length_mm") or 0),
                "width_mm": float(d.get("width_mm") or 0),
                "height_mm": float(d.get("height_mm") or 0),
                "spec": str(d.get("spec") or ""),
                "part_no": str(d.get("part_no") or ""),
                "note": str(d.get("note") or ""),
            }
        )
    wb.close()
    return mats


def run_one(path: Path, container: str = "40HQ", use_llm: bool = False) -> Dict[str, Any]:
    from packing_assistant.harness import (
        apply_user_confirmation,
        run_team_a,
        run_team_b,
    )

    mats = load_materials_xlsx(path)
    if not mats:
        return {"file": path.name, "ok": False, "error": "no materials"}

    net = sum(float(m.get("total_weight_kg") or 0) for m in mats)
    guess = min(max(int(net / 18000) + 1, 1), 12)

    t0 = time.time()
    sa = run_team_a(f"excel:{path.name}", materials=mats)
    state = None
    for mc in range(guess, 13):
        sb = apply_user_confirmation(
            sa, action="confirm", container_type=container, max_containers=mc
        )
        state = run_team_b(sb)
        if (state.get("container_plan") or {}).get("can_fit"):
            break
    ms = int((time.time() - t0) * 1000)
    plan = (state or {}).get("container_plan") or {}
    risk = (state or {}).get("risk_report") or {}
    boxes = (state or {}).get("boxes") or []
    return {
        "file": path.name,
        "ok": True,
        "ms": ms,
        "materials": len(mats),
        "net_kg": round(net, 1),
        "boxes": len(boxes),
        "box_types": [b.get("box_type") for b in boxes],
        "can_fit": plan.get("can_fit"),
        "containers_used": plan.get("containers_used"),
        "space_utilization": plan.get("space_utilization"),
        "space_best": plan.get("space_utilization_best_container"),
        "weight_utilization": plan.get("weight_utilization"),
        "risk_level": risk.get("level"),
        "risk_score": risk.get("compliance_score"),
        "llm": "LLM:" in ((state or {}).get("final_response") or ""),
        "final_preview": ((state or {}).get("final_response") or "")[:400],
        "engine": plan.get("engine"),
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", default=str(ROOT / "test" / "excel"))
    ap.add_argument("--container", default="40HQ")
    ap.add_argument("--only", default="", help="逗号分隔文件名关键字")
    ap.add_argument("--deepseek", action="store_true")
    ap.add_argument("--include-boxes-as-materials", action="store_true",
                    help="也跑 test_materials_from_boxes")
    args = ap.parse_args()

    if args.deepseek:
        info = setup_deepseek()
        print("DeepSeek:", info)

    d = Path(args.dir)
    if not d.is_absolute():
        d = ROOT / d
    if not d.exists():
        print("先运行: python scripts/build_steel_test_set.py")
        return 1

    files: List[Path] = []
    # 材料级
    for pat in (
        "test_materials_01.xlsx",
        "test_materials_from_boxes.xlsx",
        "synthetic/*.xlsx",
    ):
        files.extend(sorted(d.glob(pat)))
    # 去重
    seen = set()
    uniq = []
    for f in files:
        if f.name in seen:
            continue
        if not args.include_boxes_as_materials and f.name == "test_materials_from_boxes.xlsx":
            # 默认也跑，真实尺寸更好
            pass
        seen.add(f.name)
        uniq.append(f)
    files = uniq

    if args.only:
        keys = [k.strip() for k in args.only.split(",") if k.strip()]
        files = [f for f in files if any(k in f.name for k in keys)]

    out_dir = ROOT / "output" / "excel_tests"
    out_dir.mkdir(parents=True, exist_ok=True)

    results = []
    print("=" * 64)
    print(f"Excel tests dir={d} n={len(files)}")
    for f in files:
        print("-" * 64)
        print(f"FILE {f.relative_to(ROOT) if f.is_relative_to(ROOT) else f}")
        try:
            r = run_one(f, container=args.container, use_llm=args.deepseek)
        except Exception as e:
            r = {"file": f.name, "ok": False, "error": str(e)}
            import traceback
            traceback.print_exc()
        results.append(r)
        if r.get("ok"):
            print(
                f"  mats={r['materials']} net={r['net_kg']}kg boxes={r['boxes']} "
                f"types={r['box_types'][:6]} fit={r['can_fit']} used={r['containers_used']} "
                f"vol={r['space_utilization']} best={r.get('space_best')} "
                f"wt={r['weight_utilization']} risk={r['risk_level']}/{r['risk_score']} {r['ms']}ms"
            )
        else:
            print("  FAIL", r.get("error"))

    summary = {"results": results, "ok": sum(1 for r in results if r.get("ok")), "fail": sum(1 for r in results if not r.get("ok"))}
    (out_dir / "summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    # HTML
    rows = []
    for r in results:
        if not r.get("ok"):
            rows.append(f"<tr class='fail'><td>{r.get('file')}</td><td colspan='8'>{r.get('error')}</td></tr>")
            continue
        rows.append(
            "<tr>"
            f"<td>{r['file']}</td><td>{r['materials']}</td><td>{r['net_kg']}</td>"
            f"<td>{r['boxes']}</td><td>{r['can_fit']}</td><td>{r['containers_used']}</td>"
            f"<td>{r['space_utilization']}</td><td>{r.get('space_best')}</td>"
            f"<td>{r['weight_utilization']}</td><td>{r['risk_level']}/{r['risk_score']}</td>"
            f"<td>{','.join(r.get('box_types') or [])}</td></tr>"
        )
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"/><title>Excel 钢结构测试</title>
<style>
body{{font-family:Microsoft YaHei,sans-serif;margin:24px;background:#0f1419;color:#e7ecf3}}
table{{border-collapse:collapse;width:100%;font-size:12px}}
th,td{{border:1px solid #2a3a52;padding:6px}} th{{background:#1a2332;color:#93c5fd}}
.fail{{color:#fca5a5}} .note{{background:#1e293b;padding:12px;border-radius:8px}}
</style></head><body>
<h1>钢结构 Excel 测试集跑批 ok={summary['ok']} fail={summary['fail']}</h1>
<div class="note">数据来自远东项目 Excel 拆分 + 合成用例（非网上通用装箱单模板）。路径：材料→铁架箱型→40HQ 拼柜。</div>
<table><thead><tr>
<th>文件</th><th>材料</th><th>净重kg</th><th>箱数</th><th>can_fit</th><th>用柜</th>
<th>容积率</th><th>最满柜</th><th>重量率</th><th>风险</th><th>箱型</th>
</tr></thead><tbody>{''.join(rows)}</tbody></table>
</body></html>"""
    (out_dir / "report.html").write_text(html, encoding="utf-8")
    print("=" * 64)
    print(f"DONE ok={summary['ok']} fail={summary['fail']}")
    print(f"  {out_dir / 'summary.json'}")
    print(f"  {out_dir / 'report.html'}")
    return 0 if summary["fail"] == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
