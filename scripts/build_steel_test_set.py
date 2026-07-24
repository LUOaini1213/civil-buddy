#!/usr/bin/env python3
"""
从REDACTED-CLIENT项目 Excel 拆分业务测试集 + 生成合成用例。

网上几乎没有「钢结构材料→铁架→拼柜」完整 Excel；本脚本以现有多 sheet 为主：

  报价单           → test_materials_01.xlsx（材料清单）
  7-20装货单1/2柜  → 按铁架段拆成 test_boxes_*.xlsx
  合成             → 短件 / 超长 / 近限重 / 超重风险 等

用法:
  python scripts/build_steel_test_set.py
  python scripts/build_steel_test_set.py --source "xxx.xlsx"
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from copy import deepcopy
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

# 标准列（材料级 / 装货明细级 / 全流程）
MATERIAL_HEADERS = [
    "id",
    "name",
    "spec",
    "quantity",
    "weight_kg",
    "total_weight_kg",
    "length_mm",
    "width_mm",
    "height_mm",
    "part_no",
    "source_sheet",
    "note",
]

BOX_LINE_HEADERS = [
    "box_group",
    "box_type",
    "seq",
    "name",
    "part_no",
    "drawing_no",
    "length_mm",
    "width_mm",
    "height_mm",
    "weight_kg",
    "quantity",
    "total_weight_kg",
    "source_sheet",
    "note",
]

FULL_HEADERS = [
    "row_type",  # material | box_line | box_meta | container
    "id",
    "name",
    "box_group",
    "box_type",
    "container_type",
    "quantity",
    "weight_kg",
    "total_weight_kg",
    "length_mm",
    "width_mm",
    "height_mm",
    "part_no",
    "note",
]


def _find_source(explicit: str = "") -> Path:
    if explicit:
        p = Path(explicit)
        if not p.is_absolute():
            p = ROOT / p
        if p.exists():
            return p
        raise FileNotFoundError(explicit)
    cands = list(ROOT.glob("REDACTED-CODE*.xlsx")) + list(ROOT.glob("*REDACTED-CLIENT*.xlsx"))
    if not cands:
        raise FileNotFoundError("未找到REDACTED-CLIENT Excel（REDACTED-CODE*.xlsx）")
    return cands[0]


def _num(v: Any, default: float = 0.0) -> float:
    if v is None or v == "":
        return default
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        m = re.search(r"[-+]?\d*\.?\d+", s)
        return float(m.group()) if m else default


def _is_seq(v: Any) -> bool:
    try:
        n = int(float(v))
        return 1 <= n <= 9999
    except Exception:
        return False


def parse_quote_materials(ws) -> List[Dict[str, Any]]:
    """报价单 → 材料行。"""
    rows = list(ws.iter_rows(values_only=True))
    header_i = None
    col: Dict[str, int] = {}
    for i, row in enumerate(rows[:30]):
        cells = [str(c or "").replace("\n", "") for c in row]
        joined = "".join(cells)
        if "序号" in joined and "名称" in joined and ("数量" in joined or "重量" in joined):
            header_i = i
            for j, h in enumerate(cells):
                if h == "序号" or h.startswith("序号"):
                    col["seq"] = j
                elif "名称" in h and "存货" not in h:
                    col.setdefault("name", j)
                elif "尺寸" in h or "规格" in h:
                    col.setdefault("spec", j)
                elif "数量" in h:
                    col.setdefault("qty", j)
                elif "单件重量" in h or (h.strip() in ("单重", "单件重量(KG）", "单件重量(KG)")):
                    col.setdefault("unit_wt", j)
                elif "总重量" in h or "總重量" in h:
                    col.setdefault("total_wt", j)
                elif "存货" in h:
                    col.setdefault("stock", j)
            break
    if header_i is None:
        return []

    # 若 unit_wt 没找到，扫表头模糊
    if "unit_wt" not in col:
        cells = [str(c or "").replace("\n", "") for c in rows[header_i]]
        for j, h in enumerate(cells):
            if "重量" in h and "总" not in h and "總" not in h and "龙绅" not in h:
                col["unit_wt"] = j
                break

    materials: List[Dict[str, Any]] = []
    mid = 0
    for row in rows[header_i + 1 :]:
        if not row or not _is_seq(row[col.get("seq", 0)] if col.get("seq") is not None else row[0]):
            # 合计行
            name0 = str(row[col["name"]] if "name" in col and row[col["name"]] else "")
            if "合计" in name0 or "總" in name0:
                break
            continue
        name = str(row[col["name"]] if "name" in col else "").strip()
        if not name or name in ("名称",):
            continue
        qty = max(int(_num(row[col["qty"]]) if "qty" in col else 1), 1)
        unit_wt = _num(row[col["unit_wt"]]) if "unit_wt" in col else 0
        total_wt = _num(row[col["total_wt"]]) if "total_wt" in col else unit_wt * qty
        if unit_wt <= 0 and total_wt > 0:
            unit_wt = total_wt / qty
        spec = str(row[col["spec"]] if "spec" in col else "" or "").strip()
        # 尺寸规格常空；从名称估长度关键词留给 pipeline
        L, W, H = _guess_dims_from_name(name, spec)
        mid += 1
        materials.append(
            {
                "id": f"M{mid:03d}",
                "name": name,
                "spec": spec,
                "quantity": qty,
                "weight_kg": round(unit_wt, 4),
                "total_weight_kg": round(total_wt or unit_wt * qty, 4),
                "length_mm": L,
                "width_mm": W,
                "height_mm": H,
                "part_no": str(row[col["stock"]] if "stock" in col else "" or "").strip(),
                "source_sheet": "报价单",
                "note": "dims_estimated" if L and not _parse_lwh(spec) else "",
            }
        )
    return materials


def _parse_lwh(spec: str) -> Optional[Tuple[float, float, float]]:
    if not spec:
        return None
    nums = re.findall(r"(\d+(?:\.\d+)?)", spec.replace("×", "x").replace("*", "x").replace("X", "x"))
    if len(nums) >= 3:
        a, b, c = map(float, nums[:3])
        # 最长当 length
        dims = sorted([a, b, c], reverse=True)
        return dims[0], dims[1], dims[2]
    if len(nums) == 1:
        return float(nums[0]), 100.0, 100.0
    return None


def _guess_dims_from_name(name: str, spec: str) -> Tuple[float, float, float]:
    p = _parse_lwh(spec)
    if p:
        return p
    # 业务启发式（与 dims_override / 装货单经验对齐）
    n = name
    if any(k in n for k in ("垫片", "螺栓", "胶", "垫块")):
        return 200.0, 100.0, 50.0
    if "预埋" in n:
        return 800.0, 400.0, 300.0
    if "圆通" in n:
        return 800.0, 200.0, 200.0
    if any(k in n for k in ("钢通", "铁通", "H型", "钢梁", "支撑")):
        return 4000.0, 250.0, 250.0
    if "铁板" in n or "钢板" in n:
        return 2000.0, 1000.0, 10.0
    if "幕墙" in n or "铝" in n:
        return 2200.0, 800.0, 150.0
    return 1500.0, 300.0, 200.0


def parse_packing_sheet(ws, sheet_name: str) -> List[Dict[str, Any]]:
    """
    装货单：按「X米铁架N号 / 铁笼N号」分段，提取明细。
    注意：表中 长/宽/高 列对钢通常是截面+长度混排，按数值最大维作 length。
    """
    rows = list(ws.iter_rows(values_only=True))
    box_re = re.compile(
        r"(?P<type>\d+(?:\.\d+)?米铁架|铁笼|木箱|铁框|6米框)(?P<no>\d+)?号?"
    )
    current_group = ""
    current_type = ""
    header_map: Dict[str, int] = {}
    lines: List[Dict[str, Any]] = []

    for row in rows:
        cells = list(row)
        text = " ".join(str(c) for c in cells[:6] if c is not None)
        m = box_re.search(text.replace(" ", ""))
        # 标题行：含项目名的完整铁架标题
        if m and ("REDACTED-CLIENT" in text or "REDACTED-CODE" in text or "号" in text):
            current_type = m.group("type")
            no = m.group("no") or ""
            current_group = f"{current_type}{no}号" if no else current_type
            continue
        # 短标题如「1.1米铁架」「2米铁架」作类型确认
        if re.fullmatch(r"\d+(\.\d+)?米铁架|铁笼|木箱", text.strip().replace(" ", "")):
            current_type = text.strip()
            continue

        # 表头
        first = str(cells[0] or "").strip()
        if first in ("序号", "序 号"):
            headers = [str(c or "").replace("\n", "") for c in cells]
            header_map = {}
            for j, h in enumerate(headers):
                if h in ("序号",):
                    header_map["seq"] = j
                elif h in ("名称",):
                    header_map["name"] = j
                elif "加工件编号" in h or h == "加工件编号":
                    header_map["part"] = j
                elif "加工图号" in h:
                    header_map["drawing"] = j
                elif h in ("长", "长度", "长度(mm)"):
                    header_map["L"] = j
                elif h in ("宽", "宽度") and "W" not in h:
                    header_map["W"] = j
                elif h in ("高", "高度") and "H" not in h:
                    header_map["H"] = j
                elif "单件重量" in h:
                    header_map["wt"] = j
                elif h in ("数量",):
                    header_map["qty"] = j
                elif "总重" in h:
                    header_map["tw"] = j
            continue

        if not header_map or not _is_seq(cells[0]):
            continue
        if not current_group:
            current_group = current_type or "未分组"

        def g(key: str, default=None):
            j = header_map.get(key)
            if j is None or j >= len(cells):
                return default
            return cells[j]

        name = str(g("name") or "").strip()
        if not name:
            continue
        a, b, c = _num(g("L")), _num(g("W")), _num(g("H"))
        # 装货单里「长宽高」有时截面在前、长度在某一维；取最大为 length
        dims = sorted([a, b, c], reverse=True)
        L, W, H = (dims + [0, 0, 0])[:3]
        # 业务修正：钢通长度常在「高」列（如 1421），截面 250×250
        if a > 0 and b > 0 and c > 0:
            # 若两维接近（截面）且一维明显长 → 长件
            sorted_d = sorted([a, b, c])
            if sorted_d[2] >= sorted_d[0] * 2:
                L, W, H = sorted_d[2], sorted_d[1], sorted_d[0]
            else:
                L, W, H = a, b, c

        qty = max(int(_num(g("qty"), 1)), 1)
        wt = _num(g("wt"))
        tw = _num(g("tw"), wt * qty)
        lines.append(
            {
                "box_group": current_group,
                "box_type": current_type or _type_from_group(current_group),
                "seq": int(float(cells[0])),
                "name": name,
                "part_no": str(g("part") or "").strip(),
                "drawing_no": str(g("drawing") or "").strip(),
                "length_mm": round(L, 2),
                "width_mm": round(W, 2),
                "height_mm": round(H, 2),
                "weight_kg": round(wt, 4),
                "quantity": qty,
                "total_weight_kg": round(tw or wt * qty, 4),
                "source_sheet": sheet_name,
                "note": "",
            }
        )
    return lines


def _type_from_group(g: str) -> str:
    m = re.search(r"(\d+(?:\.\d+)?米铁架|铁笼|木箱|6米框)", g)
    return m.group(1) if m else g


def lines_to_materials(lines: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """装货明细汇总为材料清单（按 名称+尺寸 聚合）。"""
    key_map: Dict[Tuple, Dict[str, Any]] = {}
    for ln in lines:
        key = (
            ln["name"],
            ln["length_mm"],
            ln["width_mm"],
            ln["height_mm"],
            round(ln["weight_kg"], 2),
        )
        if key not in key_map:
            key_map[key] = {
                "id": "",
                "name": ln["name"],
                "spec": f"{ln['length_mm']}x{ln['width_mm']}x{ln['height_mm']}",
                "quantity": 0,
                "weight_kg": ln["weight_kg"],
                "total_weight_kg": 0.0,
                "length_mm": ln["length_mm"],
                "width_mm": ln["width_mm"],
                "height_mm": ln["height_mm"],
                "part_no": ln.get("part_no") or "",
                "source_sheet": ln.get("source_sheet") or "",
                "note": f"from_box:{ln['box_group']}",
            }
        key_map[key]["quantity"] += int(ln["quantity"])
        key_map[key]["total_weight_kg"] += float(ln["total_weight_kg"])
    mats = list(key_map.values())
    for i, m in enumerate(mats, 1):
        m["id"] = f"M{i:03d}"
        m["total_weight_kg"] = round(m["total_weight_kg"], 4)
    return mats


def write_xlsx(path: Path, headers: List[str], rows: List[Dict[str, Any]], sheet: str = "data") -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet[:31]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_multi_sheet(path: Path, sheets: Dict[str, Tuple[List[str], List[Dict[str, Any]]]]) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    first = True
    for name, (headers, rows) in sheets.items():
        if first:
            ws = wb.active
            ws.title = name[:31]
            first = False
        else:
            ws = wb.create_sheet(name[:31])
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E3A5F")
        for r in rows:
            ws.append([r.get(h, "") for h in headers])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def synthetic_cases() -> Dict[str, List[Dict[str, Any]]]:
    """
    2～5 套合成材料用例（列与材料标准一致）。
    """
    def M(i, name, q, w, L, W, H, note=""):
        return {
            "id": f"S{i:03d}",
            "name": name,
            "spec": f"{L}x{W}x{H}",
            "quantity": q,
            "weight_kg": w,
            "total_weight_kg": round(w * q, 2),
            "length_mm": L,
            "width_mm": W,
            "height_mm": H,
            "part_no": f"SYN-{i:03d}",
            "source_sheet": "synthetic",
            "note": note,
        }

    return {
        "syn_short_frames": [
            # 短件 → 偏 1.1/2 米框
            M(1, "镀锌短钢通", 40, 25, 900, 120, 120, "短件"),
            M(2, "连接板", 80, 8, 400, 300, 20, "板件"),
            M(3, "角码", 120, 2.5, 200, 150, 80, "小件"),
            M(4, "镀锌圆通短料", 30, 12, 1100, 80, 80, "短圆通"),
        ],
        "syn_long_6m": [
            # 大量超长
            M(1, "热镀锌空心铁通6m", 24, 48, 5800, 200, 200, "超长"),
            M(2, "幕墙支撑钢构件6m", 12, 95, 6000, 250, 180, "超长重"),
            M(3, "长杆件", 8, 60, 5500, 150, 150, "超长"),
        ],
        "syn_near_payload": [
            # 总重逼近 40HQ ~26t 有效载重
            M(1, "重型钢梁", 10, 1200, 4200, 350, 250, "近限重"),
            M(2, "重型钢柱", 8, 1100, 3800, 400, 400, "近限重"),
            M(3, "配重块钢件", 6, 900, 1500, 800, 600, "近限重"),
        ],
        "syn_overweight_risk": [
            # 故意超重/超尺 → 测风险
            M(1, "超重钢柱", 4, 8000, 4000, 500, 500, "单件极重"),
            M(2, "超长超宽件", 2, 3000, 12000, 2400, 1000, "超柜尺"),
            M(3, "普通连接件", 20, 15, 600, 400, 100, "正常对照"),
        ],
        "syn_mixed_realistic": [
            # 贴近REDACTED-CLIENT：长短混
            M(1, "镀锌钢通", 60, 45, 2500, 250, 250, "中长"),
            M(2, "镀锌钢通长件", 20, 85, 4200, 250, 250, "4米级"),
            M(3, "幕墙支撑", 15, 70, 3800, 300, 200, "支撑"),
            M(4, "铁垫片", 500, 0.2, 150, 100, 10, "小件"),
            M(5, "短支撑", 40, 18, 800, 150, 150, "短件"),
        ],
    }


def build_manifest(out_dir: Path, meta: Dict[str, Any]) -> None:
    (out_dir / "MANIFEST.json").write_text(
        json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    readme = """# 钢结构业务测试集（Excel）

> 网上几乎没有同类型「材料→铁架→拼柜」真实数据集。
> 本目录由 `scripts/build_steel_test_set.py` 从REDACTED-CLIENT项目 Excel 拆分 + 合成生成。

## 标准列

### materials（材料清单）
`id | name | spec | quantity | weight_kg | total_weight_kg | length_mm | width_mm | height_mm | part_no | source_sheet | note`

### boxes（已装铁架明细）
`box_group | box_type | seq | name | part_no | drawing_no | length_mm | width_mm | height_mm | weight_kg | quantity | total_weight_kg | source_sheet | note`

### full_flow
多 sheet：`materials` + `box_lines` + `containers` + `meta`

## 文件说明

| 文件 | 用途 |
|------|------|
| test_materials_01.xlsx | 报价单材料 → Team A 材料解析/选箱 |
| test_boxes_1.1m.xlsx 等 | 装货单按框型拆分 → 对照真实合箱 |
| test_full_flow.xlsx | 材料+箱子+建议柜型 |
| syn_*.xlsx | 合成：短件/超长/近限重/超重风险/混装 |

## 跑法

```bash
# 仅生成
python scripts/build_steel_test_set.py

# 用材料 Excel 跑拼柜
python scripts/run_excel_tests.py
```
"""
    (out_dir / "README.md").write_text(readme, encoding="utf-8")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", default="", help="REDACTED-CLIENT Excel 路径")
    ap.add_argument("--out", default=str(ROOT / "test" / "excel"))
    args = ap.parse_args()

    import openpyxl

    src = _find_source(args.source)
    out_dir = Path(args.out)
    if not out_dir.is_absolute():
        out_dir = ROOT / out_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"SOURCE: {src}")
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)

    # —— 报价单材料 ——
    mats_quote: List[Dict[str, Any]] = []
    if "报价单" in wb.sheetnames:
        mats_quote = parse_quote_materials(wb["报价单"])
        write_xlsx(out_dir / "test_materials_01.xlsx", MATERIAL_HEADERS, mats_quote, "materials")
        print(f"  test_materials_01.xlsx  materials={len(mats_quote)}")

    # —— 装货单按框拆分 ——
    all_lines: List[Dict[str, Any]] = []
    for sn in wb.sheetnames:
        if "装货单" in sn:
            lines = parse_packing_sheet(wb[sn], sn)
            all_lines.extend(lines)
            print(f"  sheet {sn}: lines={len(lines)} groups={len({x['box_group'] for x in lines})}")

    by_type: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    by_group: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for ln in all_lines:
        by_type[ln["box_type"] or "unknown"].append(ln)
        by_group[ln["box_group"]].append(ln)

    type_file_map = {
        "1.1米铁架": "test_boxes_1.1m.xlsx",
        "2米铁架": "test_boxes_2m.xlsx",
        "4米铁架": "test_boxes_4m.xlsx",
        "6米铁架": "test_boxes_6m.xlsx",
        "铁笼": "test_boxes_cage.xlsx",
    }
    for t, fname in type_file_map.items():
        rows = by_type.get(t) or []
        if rows:
            write_xlsx(out_dir / fname, BOX_LINE_HEADERS, rows, t[:31])
            print(f"  {fname}  lines={len(rows)} groups={len({r['box_group'] for r in rows})}")

    # 全量装货明细
    if all_lines:
        write_xlsx(out_dir / "test_boxes_all.xlsx", BOX_LINE_HEADERS, all_lines, "box_lines")
        mats_from_boxes = lines_to_materials(all_lines)
        write_xlsx(
            out_dir / "test_materials_from_boxes.xlsx",
            MATERIAL_HEADERS,
            mats_from_boxes,
            "materials",
        )
        print(f"  test_materials_from_boxes.xlsx  materials={len(mats_from_boxes)}")

    # —— full flow ——
    container_rows = [
        {
            "row_type": "container",
            "id": "C1",
            "name": "建议柜1",
            "box_group": "",
            "box_type": "",
            "container_type": "40HQ",
            "quantity": 1,
            "weight_kg": "",
            "total_weight_kg": "",
            "length_mm": "",
            "width_mm": "",
            "height_mm": "",
            "part_no": "",
            "note": "装货单1柜对应",
        },
        {
            "row_type": "container",
            "id": "C2",
            "name": "建议柜2",
            "box_group": "",
            "box_type": "",
            "container_type": "40HQ",
            "quantity": 1,
            "weight_kg": "",
            "total_weight_kg": "",
            "length_mm": "",
            "width_mm": "",
            "height_mm": "",
            "part_no": "",
            "note": "装货单2柜对应",
        },
    ]
    mat_rows = []
    src_mats = mats_from_boxes if all_lines else mats_quote
    for m in src_mats:
        mat_rows.append(
            {
                "row_type": "material",
                "id": m["id"],
                "name": m["name"],
                "box_group": "",
                "box_type": "",
                "container_type": "40HQ",
                "quantity": m["quantity"],
                "weight_kg": m["weight_kg"],
                "total_weight_kg": m["total_weight_kg"],
                "length_mm": m["length_mm"],
                "width_mm": m["width_mm"],
                "height_mm": m["height_mm"],
                "part_no": m.get("part_no") or "",
                "note": m.get("note") or "",
            }
        )
    box_meta = []
    for g, glines in by_group.items():
        tw = sum(float(x["total_weight_kg"]) for x in glines)
        box_meta.append(
            {
                "row_type": "box_meta",
                "id": g,
                "name": g,
                "box_group": g,
                "box_type": _type_from_group(g),
                "container_type": "40HQ",
                "quantity": 1,
                "weight_kg": round(tw, 2),
                "total_weight_kg": round(tw, 2),
                "length_mm": "",
                "width_mm": "",
                "height_mm": "",
                "part_no": "",
                "note": f"lines={len(glines)}",
            }
        )

    write_multi_sheet(
        out_dir / "test_full_flow.xlsx",
        {
            "materials": (FULL_HEADERS, mat_rows),
            "box_lines": (
                BOX_LINE_HEADERS,
                all_lines,
            ),
            "box_meta": (FULL_HEADERS, box_meta),
            "containers": (FULL_HEADERS, container_rows),
            "meta": (
                ["key", "value"],
                [
                    {"key": "project", "value": "REDACTED-PROJECT"},
                    {"key": "source_file", "value": src.name},
                    {"key": "pipeline", "value": "materials→frame boxes→container load"},
                    {"key": "default_container", "value": "40HQ"},
                ],
            ),
        },
    )
    # fix meta sheet simple write - write_multi_sheet expects dict rows with headers
    # recreate meta properly
    import openpyxl as ox

    fwb = ox.load_workbook(out_dir / "test_full_flow.xlsx")
    if "meta" in fwb.sheetnames:
        del fwb["meta"]
    mws = fwb.create_sheet("meta")
    mws.append(["key", "value"])
    for k, v in [
        ("project", "REDACTED-PROJECT"),
        ("source_file", src.name),
        ("pipeline", "materials→frame boxes→container load"),
        ("default_container", "40HQ"),
        ("materials_count", len(mat_rows)),
        ("box_groups", len(by_group)),
        ("box_lines", len(all_lines)),
    ]:
        mws.append([k, v])
    fwb.save(out_dir / "test_full_flow.xlsx")
    print(f"  test_full_flow.xlsx  mats={len(mat_rows)} groups={len(by_group)} lines={len(all_lines)}")

    # —— 合成 ——
    syn_dir = out_dir / "synthetic"
    syn_dir.mkdir(exist_ok=True)
    syn = synthetic_cases()
    for name, rows in syn.items():
        write_xlsx(syn_dir / f"{name}.xlsx", MATERIAL_HEADERS, rows, "materials")
        print(f"  synthetic/{name}.xlsx  n={len(rows)} net={sum(r['total_weight_kg'] for r in rows):.0f}kg")

    # schema 说明
    schema = {
        "description": "钢结构 材料→铁架→拼柜 测试 Excel 列定义",
        "material_columns": MATERIAL_HEADERS,
        "box_line_columns": BOX_LINE_HEADERS,
        "full_flow_columns": FULL_HEADERS,
        "units": {"length": "mm", "weight": "kg"},
        "synthetic_cases": {
            "syn_short_frames": "只有短件，期望偏 1.1/2 米铁架",
            "syn_long_6m": "大量 4–6 米超长件",
            "syn_near_payload": "总重接近 40HQ 限重",
            "syn_overweight_risk": "故意超重/超尺，测风险合规",
            "syn_mixed_realistic": "长短混装，贴近REDACTED-CLIENT",
        },
        "source_excel": src.name,
        "note": "网上同业务完整 Excel 几乎没有；以本项目REDACTED-CLIENT表 + 合成为主",
    }
    (out_dir / "SCHEMA.json").write_text(
        json.dumps(schema, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    manifest = {
        "source": str(src),
        "out_dir": str(out_dir),
        "files": sorted([p.name for p in out_dir.glob("*.xlsx")]),
        "synthetic": sorted([p.name for p in syn_dir.glob("*.xlsx")]),
        "stats": {
            "quote_materials": len(mats_quote),
            "box_lines": len(all_lines),
            "box_groups": len(by_group),
            "box_types": {k: len(v) for k, v in by_type.items()},
        },
    }
    build_manifest(out_dir, manifest)
    wb.close()

    print("=" * 60)
    print(f"DONE → {out_dir}")
    print(f"  SCHEMA: {out_dir / 'SCHEMA.json'}")
    print(f"  README: {out_dir / 'README.md'}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
